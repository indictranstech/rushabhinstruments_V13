from __future__ import unicode_literals
import frappe
from PIL import Image, ImageDraw
import requests
import io
import base64
import pyqrcode
import re


import json
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


import os
from frappe.utils import get_files_path
from pathlib import Path
from frappe.utils import getdate,time
import glob
from pathlib import Path
import pandas as pd
from frappe.utils import flt, cstr, nowdate, nowtime
def after_insert(doc,method):
	start_string = doc.name 
	end_string = '.png'
	label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Stock Entry' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
	if label_files:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Stock Entry' and attached_to_name=%s and file_name != %s""",
		(doc.name,label_files[0].file_name))
	else:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Stock Entry' and attached_to_name=%s""",
		(doc.name))
	# file_name = doc.name + '.pdf'
	# frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Stock Entry' and attached_to_name=%s and file_name = %s""",
	# 	(doc.name,file_name))
	pdf_data=frappe.attach_print('Stock Entry',doc.name, print_format='Stock Entry')
	
	_file = frappe.get_doc({
	"doctype": "File",
	"file_name": pdf_data.get('fname'),
	"attached_to_doctype": "Stock Entry",
	"attached_to_name": doc.name,
	"is_private": 1,
	"content": pdf_data.get('fcontent')
	})
	_file.save()
def validate(doc,method):
	if doc.work_order:
		if doc.items:
			for item in doc.items:
				if item.s_warehouse:
					engineering_revision = frappe.db.get_value("Work Order Item",{'item_code':item.item_code,'parent':doc.work_order},'engineering_revision')
					manufacturing_package = frappe.db.get_value("Work Order Item",{'item_code':item.item_code,'parent':doc.work_order},'manufacturing_package')
					item.engineering_revision = engineering_revision
					item.manufacturing_package = manufacturing_package
				if item.t_warehouse:
					engineering_revision = frappe.db.get_value("Work Order",{'production_item':item.item_code,'name':doc.work_order},'engineering_revision')
					manufacturing_package = frappe.db.get_value("Work Order",{'production_item':item.item_code,'name':doc.work_order},'manufacturing_package_name')
					item.engineering_revision = engineering_revision
					item.manufacturing_package = manufacturing_package
	if not doc.get("__islocal"):
		start_string = doc.name 
		end_string = '.png'
		label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Stock Entry' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
		if label_files:
			frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Stock Entry' and attached_to_name=%s and file_name != %s""",
			(doc.name,label_files[0].file_name))
		else:
			frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Stock Entry' and attached_to_name=%s""",
			(doc.name))
		# file_name = doc.name + '.pdf'
		# frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Stock Entry' and attached_to_name=%s and file_name = %s""",
		# 	(doc.name,file_name))
		pdf_data=frappe.attach_print('Stock Entry',doc.name, print_format='Stock Entry')
		
		_file = frappe.get_doc({
		"doctype": "File",
		"file_name": pdf_data.get('fname'),
		"attached_to_doctype": "Stock Entry",
		"attached_to_name": doc.name,
		"is_private": 1,
		"content": pdf_data.get('fcontent')
		})
		_file.save()

@frappe.whitelist()
def on_submit(doc,method):
	#update wo_status on Production Planning With Lead Time
	if doc.work_order:
		d = frappe.get_doc("Work Order", doc.work_order)
		frappe.db.set_value("Final Work Orders", {'item':d.production_item, 'sales_order':d.sales_order}, "wo_status", d.get_status())
		frappe.db.commit()

	if doc.work_order:
		frappe.db.set_value("Pick Orders", {"stock_entry":doc.name}, "stock_entry_status", "Submitted")
		frappe.db.commit()
	
	if doc.purchase_order:
		frappe.db.set_value("Pick List Purchase Order Table", {"stock_entry":doc.name}, "stock_entry_status", "Submitted")
		frappe.db.commit()

	status = []
	doc_status = []
	if doc.consolidated_pick_list:
		con_doc = frappe.get_doc("Consolidated Pick List", doc.consolidated_pick_list) 
		for row in con_doc.work_orders:
			if row.qty_of_finished_goods>0 and not row.stock_entry and not status:
				frappe.db.set_value("Pick Orders", row.name, "stock_entry_status", "Not Created")
				status.append("Not Created")
				frappe.db.commit()

		po_orders_count=frappe.db.get_list('Pick Orders', filters={'qty_of_finished_goods': ['>', 0], 'parent':doc.consolidated_pick_list}, fields=['count(name) as count'])
		po_orders_status=frappe.db.get_list('Pick Orders', filters={'qty_of_finished_goods': ['>', 0], "stock_entry_status":"Submitted", 'parent':doc.consolidated_pick_list}, fields=['count(name) as count'])
		if po_orders_count and po_orders_status and po_orders_count[0].get("count")==po_orders_status[0].get("count"):
			frappe.db.set_value("Consolidated Pick List", doc.consolidated_pick_list, "status", "Completed")

	# frappe.db.set_value("Final Work Orders", {'item':doc.production_item, 'sales_order':doc.sales_order}, "wo_status", doc.status)
	# frappe.db.commit()

@frappe.whitelist()
def on_cancel(doc,method):
	frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Stock Entry' and attached_to_name=%s""",
		(doc.name))

	#update wo_status on Production Planning With Lead Time
	if doc.work_order:
		d = frappe.get_doc("Work Order", doc.work_order)
		frappe.db.set_value("Final Work Orders", {'item':d.production_item, 'sales_order':d.sales_order}, "wo_status", d.get_status())
		frappe.db.commit()

@frappe.whitelist()
def get_items_from_pick_list(pick_list,work_order):
	if pick_list:
		qty_of_finish_good = frappe.db.get_value("Pick Orders",{'parent':pick_list,'work_order':work_order},'qty_of_finished_goods')
		items = frappe.db.sql("""SELECT item_code,warehouse as s_warehouse,picked_qty,work_order,stock_uom,engineering_revision,batch_no from `tabWork Order Pick List Item` where parent = '{0}' and work_order = '{1}' and picked_qty > 0""".format(pick_list,work_order),as_dict =1)
		# work_order_doc = frappe.get_doc("Work Order",work_order)
		# items.append({'item_code':work_order_doc.production_item,'t_warehouse':work_order_doc.fg_warehouse,'picked_qty':qty_of_finish_good,'stock_uom':work_order_doc.stock_uom,'engineering_revision':work_order_doc.engineering_revision})
		return items,qty_of_finish_good

# @frappe.whitelist()
# def before_save(doc):
# 	if doc.work_order_pick_list:
# 		if doc.items:
# 			qty_of_finish_good = frappe.db.get_value("Pick Orders",{'parent':doc.work_order_pick_list,'work_order':doc.work_order},'qty_of_finished_goods')
# 			doc.fg_completed_qty = qty_of_finish_good

@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_warehouse_for_query(doctype, txt, searchfield, start, page_len, filters):
	item_doc = frappe.get_cached_doc('Item', filters.get('parent'))
	return frappe.db.sql(""" SELECT warehouse FROM `tabItem Locations` where parent = '{0}' """.format(filters.get("parent")))
@frappe.whitelist()
def get_target_warehouse(work_order):
	item = frappe.db.get_value("Work Order",{'name':work_order},'production_item')
	if item:
		item_data = frappe.db.sql(""" SELECT warehouse FROM `tabItem Locations` where parent = '{0}' """.format(item),as_dict=1)
		item_list = [item.warehouse for item in item_data]
		return item_list
		
def label_img(doc,method):
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Stock Entry'},'url')
	final_string = url + doc.name
	img = Image.new('RGB', (384,192), color='white')
	qrc = pyqrcode.create(final_string)
	inmf = io.BytesIO()
	qrc.png(inmf,scale=6)
	qrcimg = Image.open(inmf)
	qrcimg.thumbnail((72,72))
	img.paste(qrcimg,(26,30))
	d = ImageDraw.Draw(img)
	d.multiline_text((120,35), "{0}\n\n{1}\n\nWork Order: {2}\nTarget Warehouse: {3}".format(doc.name,doc.purpose,doc.work_order,doc.to_warehouse), fill=(0,0,0), spacing=2)
	d.text((35,160), "Stock Entry Traveler", fill=(0,0,0))
	imgbuffer = io.BytesIO()
	img.save(imgbuffer, format='PNG')
	b64str = base64.b64encode(imgbuffer.getvalue())
	fname = frappe.db.get_list('File',filters={'attached_to_name':doc.name},fields=['name','file_name'])
	count=0
	if fname:
		for filedoc in fname:
			if "label" in filedoc.file_name:
				lnum = re.search("label(.*).png",filedoc.file_name)
				frappe.delete_doc('File',filedoc.name)
				count = lnum.group(1)
	namestr = doc.name + "-label{0}".format(count) + ".png"
	imgfile = frappe.get_doc({'doctype':'File','file_name':namestr,'attached_to_doctype':"Stock Entry",'attached_to_name':doc.name,"content":b64str,"decode":1})
	imgfile.insert()
@frappe.whitelist()
def print_label(data,doc):
	data = json.loads(data)
	data1 = json.loads(doc)
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Stock Entry'},'url')
	public_file_path = url.split('app')
	public_file_path = public_file_path[0] + 'files/stock_entry.xlsx'
	final_url = url + data1.get('name')
	file_path = os.path.realpath(get_files_path(is_private=0))
	file = file_path + '/' + 'stock_entry.xlsx'
	if os.path.exists(file):
		wb = openpyxl.load_workbook(filename=file)
		ws = wb['Sheet']
		row = ws.max_row +1
		# row = ws.get_highest_row() + 1
		col = 1
		cell = ws.cell(row=row,column=col)
		cell.value = ws.max_row -1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+1)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+2)
		cell.value = data1.get('stock_entry_type') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+3)
		cell.value = data1.get('work_order') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+4)
		cell.value = data1.get('to_warehouse')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+5)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = ws.cell(row=row,column=col+6)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+7)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		wb.save(file)
		frappe.msgprint(f'''Stock Entry Traveller Updated,You can download it from here
                        <a href = {public_file_path}><b>{public_file_path}</b></a>''')
		return public_file_path

	else:
		header = ['Record ID','Name','Stock Entry Type','Work Order','Target Warehouse','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1

		cell = sheet.cell(row=row,column=col)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+1)
		cell.value = data1.get('name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = data1.get('stock_entry_type') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value = data1.get('work_order') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+4)
		cell.value = data1.get('to_warehouse')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+5)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+6)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+7)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	
		file = frappe.utils.get_site_path("public")
		fname = file_path+'/' + 'stock_entry'+'.xlsx'
		book.save(fname)
		frappe.msgprint("Stock Entry Traveller Created,You can download it from here {0}".format(public_file_path))
		return public_file_path




 