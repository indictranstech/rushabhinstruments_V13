from __future__ import unicode_literals
import frappe
import json
import pyqrcode
from PIL import Image, ImageDraw
import io
import requests
import base64
import textwrap
import re
from frappe import _
import os, shutil
from frappe.utils import call_hook_method, cint, cstr, encode, get_files_path, get_hook_method, random_string, strip
from zipfile import ZipFile

import json
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


import os
from frappe.utils import get_files_path
from pathlib import Path
from frappe.utils import getdate,time
import glob
from pathlib import Path
import pandas as pd
from frappe.utils import flt, cstr, nowdate, nowtime
@frappe.whitelist()
def check_stock(doc,method):
	frappe.db.set_value("Final Work Orders", {'item':doc.production_item, 'sales_order':doc.sales_order}, "wo_status", doc.status)
	frappe.db.commit()
	if doc.get('__islocal')!= 1:
		final_item_status = []
		final_item_percent = []
		ohs = get_current_stock()
		for item in doc.required_items:
			if item.item_code in ohs:
				if item.required_qty <= ohs.get(item.item_code):
					final_item_status.append('Full Qty Available')
					percent_stock = 100
					final_item_percent.append(percent_stock)
				# elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0:
				elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0:
					final_item_status.append('Partial Qty Available')
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					final_item_percent.append(percent_stock)

				else : 
					final_item_status.append('Qty Not Available')
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					final_item_percent.append(percent_stock)

		status_list = ['Full Qty Available']
		status_list_pa = ['Partial Qty Available']
		status_list_na = ['Qty Not Available']
		check =  all(item in status_list for item in final_item_status)
		check_pa = all(item in status_list_pa for item in final_item_status)
		check_na = all(item in status_list_na for item in final_item_status)
		min_value = min(final_item_percent) if len(final_item_percent) > 1 else 0
		if check == True:
			frappe.db.set_value("Work Order",doc.name,'item_stock_status','Full Qty Available')
			frappe.db.set_value("Work Order",doc.name,'stock_percentage',min_value)
			frappe.db.commit()
			doc.reload()
		elif check_pa == True:
			frappe.db.set_value("Work Order",doc.name,'item_stock_status','Partial Qty Available')
			frappe.db.set_value("Work Order",doc.name,'stock_percentage',min_value)
			frappe.db.commit()
			doc.reload()
		elif check_na == True : 
			frappe.db.set_value("Work Order",doc.name,'item_stock_status','Qty Not Available')
			frappe.db.set_value("Work Order",doc.name,'stock_percentage',min_value)
			frappe.db.commit()
			doc.reload()
		elif 'Qty Not Available' in final_item_status and 'Partial Qty Available' in final_item_status: 
			frappe.db.set_value("Work Order",doc.name,'item_stock_status','Qty Available For Some Items')
			frappe.db.set_value("Work Order",doc.name,'stock_percentage',min_value)
			frappe.db.commit()
		else: 
			frappe.db.set_value("Work Order",doc.name,'item_stock_status','Partial Qty Available')
			frappe.db.set_value("Work Order",doc.name,'stock_percentage',min_value)
			frappe.db.commit()
			doc.reload()
	doc.reload()
def get_current_stock():
	# 1.get wip warehouse
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", 'default_wip_warehouse')
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse != '{0}' group by item_code """.format(wip_warehouse),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict

@frappe.whitelist()
def add_bom_level(doc,method):
	if doc.bom_no:
		bom_level = frappe.db.get_value("BOM",{'name' : doc.bom_no},'bom_level')
		if bom_level:
			doc.bom_level = bom_level
			
			# frappe.db.set_value("Work Order",doc.name,'bom_level',bom_level)
			# frappe.db.commit()
			# doc.reload()

@frappe.whitelist()
def on_submit(doc,method):
	prepare_zip_attachment_for_po(doc, method)
	if doc.required_items:
		for item in doc.required_items:
			if item.engineering_revision:
				er_rev = frappe.get_doc("Engineering Revision",item.engineering_revision)
				if er_rev :
					if not (er_rev.start_date and er_rev.start_transaction and er_rev.document_type):
						er_rev.start_date = doc.planned_start_date
						er_rev.document_type = "Work Order"
						er_rev.start_transaction = doc.name
					er_rev.last_date = doc.planned_start_date
					er_rev.end_document_type = "Work Order"
					er_rev.end_transaction = doc.name
					er_rev.save(ignore_permissions = True)

@frappe.whitelist()
def get_prod_engineering_revision(item_code,bom_no):
	if item_code:
		engineering_revision = frappe.db.sql("""SELECT engineering_revision from `tabItem` where item_code = '{0}'""".format(item_code),as_dict=1)
		engineering_revision[0]['use_specific_engineering_revision'] = 0
		er_from_bom = frappe.db.sql("""SELECT boi.engineering_revision ,boi.use_specific_engineering_revision from `tabBOM` bo join `tabBOM Item` boi on boi.parent = bo.name where bo.name = '{0}' and boi.item_code = '{1}' and boi.engineering_revision != ''""".format(bom_no,item_code),as_dict=1)
		if len(er_from_bom) > 0 and er_from_bom[0].get("engineering_revision") != None:
			return er_from_bom
		else:
			return engineering_revision

@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_engineering_revisions_for_filter(doctype, txt, searchfield, start, page_len, filters):
	return frappe.db.sql(""" SELECT name FROM `tabEngineering Revision` where item_code = '{0}' """.format(filters.get("item_code")))
@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_consolidated_pick_list(doctype, txt, searchfield, start, page_len, filters):
	return frappe.db.sql(""" SELECT po.parent,poi.purpose FROM `tabPick Orders` po join `tabConsolidated Pick List` poi on poi.name = po.parent where po.work_order = '{0}' """.format(filters.get("work_order")))
def after_insert(doc,method):
	start_string = doc.name 
	end_string = '.png'
	label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Work Order' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
	if label_files:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Work Order' and attached_to_name=%s and file_name != %s""",
		(doc.name,label_files[0].file_name))
	else:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Work Order' and attached_to_name=%s""",
		(doc.name))
	pdf_data=frappe.attach_print('Work Order',doc.name, print_format='Work Order')
	
	_file = frappe.get_doc({
	"doctype": "File",
	"file_name": pdf_data.get('fname'),
	"attached_to_doctype": "Work Order",
	"attached_to_name": doc.name,
	"is_private": 1,
	"content": pdf_data.get('fcontent')
	})
	_file.save()

def validate(doc,method):
	# doc.skip_transfer =1
	prod_item_engineering_revision = get_engineering_revision(doc.production_item)
	doc.engineering_revision = prod_item_engineering_revision
	if doc.engineering_revision:
		manufacturing_package = frappe.db.get_value("Manufacturing Package Table",{'parent':doc.engineering_revision},'manufacturing_package_name')
		doc.manufacturing_package_name = manufacturing_package
	for item in doc.required_items:
		engineering_revision = get_prod_engineering_revision(item.item_code,doc.bom_no)
		item.engineering_revision = engineering_revision[0].get("engineering_revision")
		item.use_specific_engineering_revision = engineering_revision[0].get("use_specific_engineering_revision")
		if item.engineering_revision:
			manufacturing_package = frappe.db.get_value("Manufacturing Package Table",{'parent':item.engineering_revision},'manufacturing_package_name')
			item.manufacturing_package = manufacturing_package
	if not doc.get("__islocal"):
		start_string = doc.name 
		end_string = '.png'
		label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Work Order' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
		if label_files:
			frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Work Order' and attached_to_name=%s and file_name != %s""",
			(doc.name,label_files[0].file_name))
		else:
			frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Work Order' and attached_to_name=%s""",
			(doc.name))
		pdf_data=frappe.attach_print('Work Order',doc.name, print_format='Work Order')
		
		_file = frappe.get_doc({
		"doctype": "File",
		"file_name": pdf_data.get('fname'),
		"attached_to_doctype": "Work Order",
		"attached_to_name": doc.name,
		"is_private": 1,
		"content": pdf_data.get('fcontent')
		})
		_file.save()
	if doc.status == 'Draft':
		unstock_items_details(doc.bom_no)

@frappe.whitelist()
def get_engineering_revision(item_code):
	if item_code:
		engineering_revision = frappe.db.get_value("Item",{'name':item_code},'engineering_revision')
		return engineering_revision

def disable_bom(doc,method):
	bom = frappe.get_doc('BOM',doc.bom_no)
	wos_for_bom = frappe.db.sql("""SELECT COUNT(name) as wo_num FROM `tabWork Order` WHERE bom_no='{}' AND status IN ('Submitted','Not Started','In Process','Draft') GROUP BY bom_no""".format(doc.bom_no), as_dict=True)
	if not wos_for_bom:
		if bom.to_be_disabled and frappe.db.get_value("Item",{'name':bom.item},'auto_disable_old_active_boms'):
			any_mboms = frappe.db.sql("""SELECT name FROM `tabMapped BOM Item` WHERE bom_no='{0}'""".format(bom.name))
			if not any_mboms:
				bom.is_active = 0
				bom.save()
				bom.submit()
				
def label_img(doc,method):
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Work Order'},'url')
	final_string = url + doc.name
	img = Image.new('RGB', (384,192), color='white')
	qrc = pyqrcode.create(final_string)
	inmf = io.BytesIO()
	qrc.png(inmf,scale=6)
	qrcimg = Image.open(inmf)
	qrcimg.thumbnail((72,72))
	img.paste(qrcimg,(26,30))
	d = ImageDraw.Draw(img)
	itemname = textwrap.fill(doc.item_name,width=35)
	d.multiline_text((120,30), "{0}\n\nItem to Manufacture: {1}\n\nQty to Manufacture: {2} \nSales Order: {3}\nWIPWarehouse: {4}\nTarget Warehouse: {5}\nItem Name: {6}".format(doc.name,doc.production_item,doc.qty,doc.sales_order,doc.wip_warehouse,doc.fg_warehouse,itemname), fill=(0,0,0), spacing=1)
	d.text((35,160), "Work Order Traveler", fill=(0,0,0))
	imgbuffer = io.BytesIO()
	img.save(imgbuffer, format='PNG')
	b64str = base64.b64encode(imgbuffer.getvalue())
	fname = frappe.db.get_list('File',filters={'attached_to_name':doc.name},fields=['name','file_name'])
	count=0
	if fname:
		for filedoc in fname:
			if "label" in filedoc.file_name:
				lnum = re.search("label(.*).png",filedoc.file_name)
				count = int(lnum.group(1))+1
				frappe.delete_doc('File',filedoc.name)
	namestr = doc.name + "-label{0}".format(count) + ".png"
	imgfile = frappe.get_doc({'doctype':'File','file_name':namestr,'attached_to_doctype':"Work Order",'attached_to_name':doc.name,"content":b64str,"decode":1})
	imgfile.insert()
	


# Overriding Methods
@frappe.whitelist()
def stop_unstop(work_order, status):
	"""Called from client side on Stop/Unstop event"""
	if not frappe.has_permission("Work Order", "write"):
		frappe.throw(_("Not permitted"), frappe.PermissionError)

	pro_order = frappe.get_doc("Work Order", work_order)

	if pro_order.status == "Closed":
		frappe.throw(_("Closed Work Order can not be stopped or Re-opened"))

	pro_order.update_status(status)
	pro_order.update_planned_qty()
	frappe.msgprint(_("Work Order has been {0}").format(status))
	pro_order.notify_update()
	frappe.db.set_value("Final Work Orders", {'item':pro_order.production_item, 'sales_order':pro_order.sales_order}, "wo_status", pro_order.status)
	frappe.db.commit()
	return pro_order.status	
	

def update_status_on_production_planning_with_lead_time(doc, method=None):
	frappe.db.set_value("Final Work Orders", {'item':doc.production_item, 'sales_order':doc.so_reference}, "wo_status", doc.status)
	frappe.db.commit()


def on_trash(doc, method=None):
	frappe.db.set_value("Final Work Orders", {'item':doc.production_item, 'sales_order':doc.sales_order}, "wo_status", "")
	frappe.db.commit()


@frappe.whitelist()
def unstock_items_details(bom_no):
	unstock_items = []
	if bom_no:
		bom_doc = frappe.get_doc("BOM", bom_no)
		for row in bom_doc.items:
			if not frappe.db.get_value("Item", row.item_code, "is_stock_item"):
				unstock_items.append({"item_code":row.item_code, "item_name":row.item_name, "description":row.description, "qty":row.qty})				
	return unstock_items


#Create ZIP File and Attach

def prepare_zip_attachment_for_po(doc, method):
	all_files = []
	item_dict = []
	engineering_revision=""
	if not doc.engineering_revision:
		engineering_revision=frappe.db.get_value("Engineering Revision", {"item_code":doc.production_item, "is_default":True, "is_active":True}, "name")
	else:
		engineering_revision = doc.engineering_revision
	item_dict.append({"engineering_revision":engineering_revision, "item_code":doc.production_item, "name":doc.name})
	
	for row in doc.required_items:
		if not row.engineering_revision:
			engineering_revision=frappe.db.get_value("Engineering Revision", {"item_code":row.item_code, "is_default":True, "is_active":True}, "name")
		else:
			engineering_revision = row.engineering_revision
		
		item_dict.append({"engineering_revision":engineering_revision, "item_code":row.item_code, "name":doc.name})

	for row in item_dict:
		if row.get('item_code') and row.get('engineering_revision'):
			engineering_revision_doc(row, all_files)
	if all_files:
		create_zip_file(doc, all_files)
		doc.reload()

def engineering_revision_doc(row, all_files): 
	purchasing_package_list = []
	item_folder = []
	purchasing_package = frappe.db.sql("""SELECT purchasing_package_name from `tabPurchasing Package Table` a join `tabEngineering Revision` b on a.parent = b.name where b.name = '{0}'""".format(row.get('engineering_revision')),as_dict=1,debug=False)
	purchasing_package_list = [item.get('purchasing_package_name') for item in purchasing_package]
	# file_path = os.path.realpath(get_files_path(is_private=1))
	file_path = get_files_path(is_private=1)
	if row.get('item_code') not in item_folder:
		full_path = file_path+ "/"+row.get('item_code')
		os.mkdir(full_path)
		all_files.append(full_path)
		item_folder.append(row.get('item_code'))

		for col in purchasing_package_list:
			package_doc = frappe.get_doc("Package Document",col)
			from frappe.desk.form.load import get_attachments
			attachments = get_attachments(package_doc.doctype, package_doc.name)
			create_folder_with_file(attachments, full_path)

		if row.get('engineering_revision'):
			er_doc = frappe.get_doc("Engineering Revision",row.get('engineering_revision'))
			other_dict={}
			for i in er_doc.other_engineering_revision:
				if i.revision and not i.exclude_purchasing_package:
					other_dict["item_code"]=i.item_code
					other_dict["parent"]=row.get("name")
					other_dict["engineering_revision"]=i.revision
				elif i.purchase_package_name and not i.exclude_purchasing_package:
					other_dict["item_code"]=i.item_code
					other_dict["parent"]=row.get("name")
					other_dict["engineering_revision"]="_".join(i.purchase_package_name.split("_")[0:2])
				else:
					if not i.exclude_purchasing_package:
						other_dict["item_code"]=i.item_code
						other_dict["parent"]=row.get("name")
						other_dict["engineering_revision"]=frappe.db.get_value("Engineering Revision", {"item_code":i.item_code, "is_default":True, "is_active":True}, "name")
				engineering_revision_doc(other_dict, all_files)


def create_folder_with_file(attachments, full_path):
	for i in attachments:
		file_doc = frappe.get_doc("File",{'file_url':i.file_url})
		with open(full_path+f"/{i.file_name}","wb") as f:
			try:
				f.write(file_doc.get_content())
			except Exception as e:
				print(e)

def create_zip_file(doc, all_files):
	file_path = get_files_path(is_private=1)
	zip_full_path = file_path+ "/"+doc.engineering_revision+"_"+doc.name+".zip"
	file_name = doc.engineering_revision+"_"+doc.name+".zip"
	file_url = '/private/files/'+file_name

	with ZipFile(zip_full_path,'w') as zip:
		for full_path in all_files:
			for dirname, subdirs, files in os.walk(full_path):
				zip.write(dirname)
				for filename in files:
					zip.write(os.path.join(dirname, filename))
			shutil.rmtree(full_path)
	file_doc = frappe.new_doc("File")
	file_doc.file_name =file_name
	file_doc.folder = "Home/Attachments"
	file_doc.attached_to_doctype = doc.doctype
	file_doc.attached_to_name = doc.name
	file_doc.file_url = file_url
	file_doc.insert(ignore_permissions=True)
	frappe.db.commit()


@frappe.whitelist()
def get_fg_item_for_consolidated_pick_list(item):
	fg_item_groups = frappe.db.sql("""SELECT item_group from `tabTable For Item Group`""",as_dict=1)
	fg_item_groups = [item.get('item_group') for item in fg_item_groups]
	item_group = frappe.db.get_value("Item", {"name":item}, "item_group")
	return True if item_group in fg_item_groups else False


@frappe.whitelist()
def make_consolidated_pick_list(source_name, target_doc=None):
	from frappe.model.mapper import get_mapped_doc
	target_doc = get_mapped_doc("Work Order", source_name, {
			"Work Order": {
				"doctype": "Consolidated Pick List",
				"field_map": {
					"name": "work_order",
				},
			}
		})
	return target_doc


@frappe.whitelist()
def print_label(data,doc):
	data = json.loads(data)
	data1 = json.loads(doc)
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Work Order'},'url')
	public_file_path = url.split('app')
	public_file_path = public_file_path[0] + 'files/work_order.xlsx'
	final_url = url + data1.get('name')
	file_path = os.path.realpath(get_files_path(is_private=0))
	file = file_path + '/' + 'work_order.xlsx'
	barcode_doc = frappe.db.get_value("Barcodes",{'doctype_name':'Work Order'},'name')
	if barcode_doc:
		barcode_doc = frappe.get_doc("Barcodes",barcode_doc)
		if barcode_doc:
			barcode_doc.append('table_for_label_print',{
				'work_order':data1.get('name')  ,
				'part_number':data1.get('production_item'),
				'item_name':data1.get('item_name') ,
				'qty_to_manufacture':data1.get('qty'),
				'sales_order':data1.get('sales_order'),
				'wip_warehouse':data1.get('wip_warehouse'),
				'target_warehouse':data1.get('fg_warehouse'),
				'no_of_copies':data.get('no_of_copies') ,
				'printer':data.get('printer_name'),
				'url':final_url 
				})
			barcode_doc.save()
	else:
		b_doc = frappe.new_doc("Barcodes")
		if b_doc:
			b_doc.doctype_name = "Work Order"
			b_doc.append('table_for_label_print',{
				'work_order':data1.get('name')  ,
				'part_number':data1.get('production_item'),
				'item_name':data1.get('item_name') ,
				'qty_to_manufacture':data1.get('qty'),
				'sales_order':data1.get('sales_order'),
				'wip_warehouse':data1.get('wip_warehouse'),
				'target_warehouse':data1.get('fg_warehouse'),
				'no_of_copies':data.get('no_of_copies') ,
				'printer':data.get('printer_name'),
				'url':final_url 
				})
		b_doc.save()
	if os.path.exists(file):
		wb = openpyxl.load_workbook(filename=file)
		ws = wb['Sheet']
		row = ws.max_row +1
		# row = ws.get_highest_row() + 1
		col = 1
		cell = ws.cell(row=row,column=col)
		cell.value = ws.max_row -1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+1)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+2)
		cell.value = data1.get('production_item') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+3)
		cell.value = data1.get('item_name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+4)
		cell.value = data1.get('qty')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+5)
		cell.value = data1.get('sales_order')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+6)
		cell.value = data1.get('wip_warehouse')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+7)
		cell.value = data1.get('fg_warehouse')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+8)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = ws.cell(row=row,column=col+9)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+10)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		wb.save(file)
		frappe.msgprint(f'''Work Order Traveller Updated,You can download it from here
                        <a href = {public_file_path}><b>{public_file_path}</b></a>''')
		return public_file_path

	else:
		header = ['Record ID','Work Order Name','Item to Manufacture','Item Name','Qty To Manufacture','Sales Order','WIP Warehouse','Target Warehouse','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1

		cell = sheet.cell(row=row,column=col)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+1)
		cell.value = data1.get('name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = data1.get('production_item') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value = data1.get('item_name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+4)
		cell.value = data1.get('qty')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell = sheet.cell(row=row,column=col+5)
		cell.value = data1.get('sales_order')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+6)
		cell.value = data1.get('wip_warehouse')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+7)
		cell.value = data1.get('fg_warehouse')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+8)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+9)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+10)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	
		file = frappe.utils.get_site_path("public")
		fname = file_path+'/' + 'work_order'+'.xlsx'
		book.save(fname)
		frappe.msgprint("Work Order Traveller Created,You can download it from here {0}".format(public_file_path))
		return public_file_path




 