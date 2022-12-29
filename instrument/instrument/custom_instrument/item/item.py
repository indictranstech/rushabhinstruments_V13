from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from PIL import Image, ImageDraw
import pyqrcode
import io
import base64
import requests
import textwrap
import re

import json
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


import os
from frappe.utils import get_files_path
from pathlib import Path
from frappe.utils import getdate,time
import glob
from pathlib import Path
import pandas as pd
from frappe.utils import flt, cstr, nowdate, nowtime
def after_insert(doc,method):
	start_string = doc.name 
	end_string = '.png'
	label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Item' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
	if label_files:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Item' and attached_to_name=%s and file_name != %s""",
		(doc.name,label_files[0].file_name))
	else:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Item' and attached_to_name=%s""",
		(doc.name))
	pdf_data=frappe.attach_print('Item',doc.name, print_format='Item Print 8.5" x 11"')
	
	_file = frappe.get_doc({
	"doctype": "File",
	"file_name": pdf_data.get('fname'),
	"attached_to_doctype": "Item",
	"attached_to_name": doc.name,
	"is_private": 1,
	"content": pdf_data.get('fcontent')
	})
	_file.save()
def validate(doc,method):
	# if not doc.get("__islocal"):
	start_string = doc.name 
	end_string = '.png'
	label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Item' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
	if label_files:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Item' and attached_to_name=%s and file_name != %s""",
		(doc.name,label_files[0].file_name))
	else:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Item' and attached_to_name=%s""",
		(doc.name))
	pdf_data=frappe.attach_print('Item',doc.name, print_format='Item Print 8.5" x 11"')
	
	_file = frappe.get_doc({
	"doctype": "File",
	"file_name": pdf_data.get('fname'),
	"attached_to_doctype": "Item",
	"attached_to_name": doc.name,
	"is_private": 1,
	"content": pdf_data.get('fcontent')
	})
	_file.save()
	if doc.item_attribute_table:
		attribute_list = [item.attribute for item in doc.item_attribute_table]
		attribute_set = set(attribute_list)
		if len(attribute_set) != len(attribute_list):
			frappe.throw("Duplicate Attribute Not Allowed")

	warehouse_list  = []
	if doc.warehouses:
		for row in doc.warehouses:
			warehouse_list.append(row.warehouse)

	if doc.item_defaults:
		for row in doc.item_defaults:
			if row.default_warehouse not in warehouse_list:
				doc.append('warehouses',{
					'warehouse':row.default_warehouse
					})
				# frappe.throw("Please Add Default Warehouse {0} in Item Locations".format(row.default_warehouse))

def disable_old_boms(doc,method):
	if doc.auto_disable_old_active_boms:
		old_boms = frappe.db.sql("""SELECT name from `tabBOM` WHERE item='{0}' and name<(SELECT name FROM `tabBOM` WHERE item='{0}' AND is_default=1) """.format(doc.name), as_dict=1)
		for bom in old_boms:
			bom_doc = frappe.get_doc('BOM', bom)
			if bom_doc.is_active:
				any_wos = frappe.db.sql("""SELECT name FROM `tabWork Order` WHERE bom_no='{0}' AND status IN ('Submitted','Not Started', 'In Process','Draft')""".format(bom['name']))
				any_mboms = frappe.db.sql("""SELECT name FROM `tabMapped BOM Item` WHERE bom_no='{0}'""".format(bom['name']))
				if not any_wos and not any_mboms:
					bom_doc.is_active = 0
				if any_wos:
					bom_doc.to_be_disabled = 1
				bom_doc.save()
				bom_doc.submit()

def label_img(doc,method):
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Item'},'url')
	final_string = url + doc.name
	img = Image.new('RGB', (600, 600), color='white')
	qrc = pyqrcode.create(final_string)
	inmf = io.BytesIO()
	qrc.png(inmf,scale=6)
	qrcimg = Image.open(inmf)
	qrcimg.thumbnail((72,72))
	img.paste(qrcimg,(12,12))
	d = ImageDraw.Draw(img)
	itemname = textwrap.fill(text = doc.item_name,width=18,max_lines=5,placeholder="...")
	d.text((90,10), itemname, fill=(0,0,0))
	d.text((12,96), doc.name, fill=(0,0,0))
	item_locations = frappe.db.get_list('Item Locations',{'parent':doc.name},pluck='warehouse')
	locs_str = ""
	for loc in item_locations:
		locs_str += loc
	locs_str = textwrap.fill(text=locs_str,width=40,max_lines=4,placeholder="...")
	d.text((12,110), "Item Locations: {0}".format(locs_str), fill=(0,0,0))
	barcode = requests.get('https://barcode.tec-it.com/barcode.ashx?data={0}&code=Code128&translate-esc=true'.format(doc.item_code))
	barc = Image.open(io.BytesIO(barcode.content))
	barc = barc.resize((180,20))
	img.paste(barc,(6,160))
	imgbuffer = io.BytesIO()
	img.save(imgbuffer, format='PNG')
	b64str = base64.b64encode(imgbuffer.getvalue())
	fname = frappe.db.get_list('File',filters={'attached_to_name':doc.name},fields=['name','file_name'])
	count=0
	if fname:
		for filedoc in fname:
			if "label" in filedoc.file_name:
				lnum = re.search("label(.*).png",filedoc.file_name)
				frappe.delete_doc('File',filedoc.name)
				count = lnum.group(1)
	namestr = doc.name + "-label{0}".format(count) + ".png"
	imgfile = frappe.get_doc({'doctype':'File','file_name':namestr,'attached_to_doctype':"Item",'attached_to_name':doc.name,"content":b64str,"decode":1})
	imgfile.insert()

@frappe.whitelist()
def get_label_details(template_name):
	if template_name:
		data = frappe.db.sql("""SELECT a.label from `tabItem Additional Label Info Template Table` a join `tabItem Additional Label Info Template` b on b.name =a.parent where b.template_name = '{0}' order by a.idx""".format(template_name),as_dict=1)
		return data

@frappe.whitelist()
def print_label(data,doc):
	data = json.loads(data)
	data1 = json.loads(doc)
	item_locations = [item.get('warehouse') for item in data1.get('warehouses')]
	warehouse_list = ''
	if len(item_locations) >0:
		for i in item_locations:
			warehouse_list += i
			warehouse_list += ','
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Item'},'url')
	public_file_path = url.split('app')
	public_file_path = public_file_path[0] + 'files/item_traveler.xlsx'
	final_url = url + data1.get('name')
	file_path = os.path.realpath(get_files_path(is_private=0))
	file = file_path + '/' + 'item_traveler.xlsx'
	if os.path.exists(file):
		wb = openpyxl.load_workbook(filename=file)
		ws = wb['Sheet']
		row = ws.max_row +1
		# row = ws.get_highest_row() + 1
		col = 1
		cell = ws.cell(row=row,column=col)
		cell.value = ws.max_row -1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+1)
		cell.value = data1.get('item_code') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+2)
		cell.value = data1.get('item_name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+3)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+4)
		cell.value = warehouse_list
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+5)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = ws.cell(row=row,column=col+6)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+7)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		wb.save(file)
		frappe.msgprint(f'''Item Traveller Updated,You can download it from here
                        <a href = {public_file_path}><b>{public_file_path}</b></a>''')
		return public_file_path

	else:
		header = ['Record ID','Part Number','Part Description','Item Code','Locations','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1

		cell = sheet.cell(row=row,column=col)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+1)
		cell.value = data1.get('item_code')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = data1.get('item_name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+4)
		cell.value = warehouse_list
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+5)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+6)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+7)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	
		file = frappe.utils.get_site_path("public")
		fname = file_path+'/' + 'item_traveler'+'.xlsx'
		book.save(fname)
		frappe.msgprint("Item Traveller Created,You can download it from here {0}".format(public_file_path))
		return public_file_path




 