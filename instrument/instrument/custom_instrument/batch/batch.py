import frappe
from frappe.model.naming import make_autoname
from frappe.utils import nowdate, cstr, flt, cint, now, getdate,get_datetime,time_diff_in_seconds,add_to_date,time_diff_in_seconds,add_days,today
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import io
import base64
import pyqrcode
import requests
import textwrap
import re

import json
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


import os
from frappe.utils import get_files_path
from pathlib import Path
from frappe.utils import getdate,time
import glob
from pathlib import Path
import pandas as pd
from frappe.utils import flt, cstr, nowdate, nowtime
# import xlsxwriter
def autoname(doc, method):
	if doc.item:
		now = datetime.now()
		currentMonth = datetime.now().month
		currentMonth = '{:02d}'.format(currentMonth)
		currentYear = datetime.now().year
		
		engineering_revision = frappe.db.get_value("Item",{'item_code':doc.item},'engineering_revision')
		# get_naming_prefix(doc,naming_series)
		if doc.reference_doctype == 'Purchase Receipt':
			if doc.reference_name:
				pr_doc = frappe.get_doc("Purchase Receipt",doc.reference_name)
				for item in pr_doc.items:
					if item.item_code == doc.item:
						naming_series = frappe.db.get_value("Item",{'item_code':doc.item},'batch_number_series')
						naming_prefix = get_naming_prefix(doc,item.engineering_revision,currentYear,naming_series)

						doc.name = make_autoname(naming_prefix+'-'+".#####")
						# doc.name = make_autoname('BN-' + str(item.engineering_revision) + '-'+str(currentYear) +'-'+str(currentMonth) + '-' + '.#####')
						doc.batch_id = doc.name
						return doc.name
		elif doc.reference_doctype == 'Stock Entry':
			if doc.reference_name:
				se_doc = frappe.get_doc("Stock Entry",doc.reference_name)
				for item in se_doc.items:
					if item.item_code == doc.item:
						naming_series = frappe.db.get_value("Item",{'item_code':doc.item},'batch_number_series')
						naming_prefix = get_naming_prefix(doc,item.engineering_revision,currentYear,naming_series)

						doc.name = make_autoname(naming_prefix+'-'+".#####")
						# doc.name = make_autoname('BN-' + str(item.engineering_revision) + '-'+str(currentYear) +'-'+str(currentMonth) + '-' + '.#####')
						doc.batch_id = doc.name
						return doc.name
		elif engineering_revision:
			naming_series = frappe.db.get_value("Item",{'item_code':doc.item},'batch_number_series')
			naming_prefix = get_naming_prefix(doc,engineering_revision,currentYear,naming_series)

			doc.name = make_autoname(naming_prefix+'-'+".#####")
			# doc.name = make_autoname('BN-' + str(item.engineering_revision) + '-'+str(currentYear) +'-'+str(currentMonth) + '-' + '.#####')
			doc.batch_id = doc.name
			return doc.name
		else:
			doc.name = make_autoname('BN-' + '-'+str(currentYear) +'-'+str(currentMonth) + '-' + '.#####')
			doc.batch_id = doc.name
			return doc.name

def get_naming_prefix(doc,engineering_revision,currentYear,naming_series):
	now = datetime.now()
	week_no = datetime.date(now).isocalendar()[1]
	naming_series = naming_series.split('-')
	naming_series = naming_series[0]
	new_naming = naming_series
	if 'ENG' in naming_series:
		new_naming = naming_series.replace('ENG',str(engineering_revision))
	if 'YYYY' in naming_series:
		new_naming = new_naming.replace('YYYY',str(currentYear))
	if 'WW' in naming_series:
		new_naming = new_naming.replace('WW',str(week_no))
	return new_naming



def label_img(doc, method):
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Batch'},'url')
	final_string = url +  doc.name
	warehouse = frappe.db.get_value('Stock Ledger Entry',{'batch_no':doc.name,'item_code':doc.item,'posting_date':doc.manufacturing_date},'warehouse')
	if not warehouse:
		warehouse = ""
	img = Image.new('RGB', (384,192), color='white')
	qrc = pyqrcode.create(final_string)
	inmf = io.BytesIO()
	qrc.png(inmf,scale=6)
	qrcimg = Image.open(inmf)
	qrcimg.thumbnail((72,72))
	img.paste(qrcimg,(26,30))
	d = ImageDraw.Draw(img)
	d.multiline_text((120,35), "{0}\n\n{1}\n\nTotal Qty: {2}\nBatch: {3}\nBatch Name: {4}\nLocation: {5}".format(doc.item,textwrap.fill(text=doc.item_name,width=40),doc.batch_qty,doc.batch_id,doc.name,warehouse), fill=(0,0,0), spacing=1)
	d.text((30,160), "Batch Traveler", fill=(0,0,0))
	barcode = requests.get('https://barcode.tec-it.com/barcode.ashx?data={0}&code=Code128&translate-esc=true'.format(doc.item))
	barc = Image.open(io.BytesIO(barcode.content))
	barc = barc.resize((220,15))
	img.paste(barc,(140,160))
	imgbuffer = io.BytesIO()
	img.save(imgbuffer, format='PNG')
	b64str = base64.b64encode(imgbuffer.getvalue())
	fname = frappe.db.get_list('File',filters={'attached_to_name':doc.name},fields=['name','file_name'])
	count=0
	if fname:
		for filedoc in fname:
			if "label" in filedoc.file_name:
				lnum = re.search("label(.*).png",filedoc.file_name)
				count = int(lnum.group(1))+1
				frappe.delete_doc('File',filedoc.name)
	namestr = doc.name + "-label{0}".format(count) + ".png"
	imgfile = frappe.get_doc({'doctype':'File','file_name':namestr,'attached_to_doctype':"Batch",'attached_to_name':doc.name,"content":b64str,"decode":1})
	imgfile.insert()
def after_insert(doc,method):
	# if not doc.get("__islocal"):
	# file_url = '/private/files/' + doc.name + '.pdf'
	# frappe.db.sql("""delete from `tabFile` where attached_to_doctype='Batch' and attached_to_name=%s and file_url = %s""",
	# (doc.name,file_url))
	start_string = doc.name 
	end_string = '.png'
	label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Batch' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
	if label_files:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Batch' and attached_to_name=%s and file_name != %s""",
		(doc.name,label_files[0].file_name))
	else:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Batch' and attached_to_name=%s""",
		(doc.name))
	pdf_data=frappe.attach_print('Batch',doc.name, print_format='Batch Print 8.5" x 11"')
	
	_file = frappe.get_doc({
	"doctype": "File",
	"file_name": pdf_data.get('fname'),
	"attached_to_doctype": "Batch",
	"attached_to_name": doc.name,
	"is_private": 1,
	"content": pdf_data.get('fcontent')
	})
	_file.save()
def validate(doc,method):
	# if not doc.get("__islocal"):
	# file_url = '/private/files/' + doc.name + '.pdf'
	# frappe.db.sql("""delete from `tabFile` where attached_to_doctype='Batch' and attached_to_name=%s and file_url = %s""",
	# (doc.name,file_url))
	start_string = doc.name 
	end_string = '.png'
	label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Batch' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
	if label_files:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Batch' and attached_to_name=%s and file_name != %s""",
		(doc.name,label_files[0].file_name))
	else:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Batch' and attached_to_name=%s""",
		(doc.name))
	pdf_data=frappe.attach_print('Batch',doc.name, print_format='Batch Print 8.5" x 11"')
	
	_file = frappe.get_doc({
	"doctype": "File",
	"file_name": pdf_data.get('fname'),
	"attached_to_doctype": "Batch",
	"attached_to_name": doc.name,
	"is_private": 1,
	"content": pdf_data.get('fcontent')
	})
	_file.save()


@frappe.whitelist()
def print_label(data,doc):
	data = json.loads(data)
	data1 = json.loads(doc)
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Batch'},'url')
	public_file_path = url.split('app')
	public_file_path = public_file_path[0] + 'files/batch_traveller.xlsx'
	final_url = url + data1.get('name')
	file_path = os.path.realpath(get_files_path(is_private=0))
	file = file_path + '/' + 'batch_traveller.xlsx'
	if os.path.exists(file):
		wb = openpyxl.load_workbook(filename=file)
		ws = wb['Sheet']
		row = ws.max_row +1
		col = 1
		cell = ws.cell(row=row,column=col)
		cell.value = ws.max_row -1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+1)
		cell.value = data1.get('item') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+2)
		cell.value = data1.get('item_name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+3)
		cell.value = data1.get('batch_qty') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+4)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+5)
		cell.value = data1.get('batch_id') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+6)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+7)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+8)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = ws.cell(row=row,column=col+9)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		wb.save(file)
		frappe.msgprint(f'''Batch Traveller Updated,You can download it from here
                        <a href = {public_file_path}><b>{public_file_path}</b></a>''')
		return public_file_path

	else:
		header = ['Record ID','Part Number','Part Description','Total Qty','Batch','Batch Name','Locations','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1

		cell = sheet.cell(row=row,column=col)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+1)
		cell.value = data1.get('item')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = data1.get('item_name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value = data1.get('batch_qty')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+4)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+5)
		cell.value = data1.get('batch_id') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+6)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+7)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+8)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+9)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	
		file = frappe.utils.get_site_path("public")
		fname = file_path+'/' + 'batch_traveller'+'.xlsx'
		book.save(fname)
		frappe.msgprint("Batch Traveller Created,You can download it from here {0}".format(public_file_path))
		return public_file_path




 