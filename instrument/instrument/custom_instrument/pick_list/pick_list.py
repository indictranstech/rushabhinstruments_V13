from __future__ import unicode_literals
import frappe
from PIL import Image, ImageDraw
import io
import base64
import pyqrcode
import re

import json
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


import os
from frappe.utils import get_files_path
from pathlib import Path
from frappe.utils import getdate,time
import glob
from pathlib import Path
# import pandas as pd
from frappe.utils import flt, cstr, nowdate, nowtime
def label_img(doc,method):
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Pick List'},'url')
	final_string = url + doc.name
	img = Image.new('RGB', (384,192), color='white')
	qrc = pyqrcode.create(final_string)
	inmf = io.BytesIO()
	qrc.png(inmf,scale=6)
	qrcimg = Image.open(inmf)
	qrcimg.thumbnail((72,72))
	img.paste(qrcimg,(26,30))
	d = ImageDraw.Draw(img)
	d.multiline_text((120,35), "{0}\n\nWork Order: {1}\nQty of Finished Goods Items: {2}".format(doc.name,doc.work_order,doc.for_qty), fill=(0,0,0), spacing=2)
	d.text((35,160), "Pick List Traveler", fill=(0,0,0))
	imgbuffer = io.BytesIO()
	img.save(imgbuffer, format='PNG')
	b64str = base64.b64encode(imgbuffer.getvalue())
	fname = frappe.db.get_list('File',filters={'attached_to_name':doc.name},fields=['name','file_name'])
	count=0
	if fname:
		for filedoc in fname:
			if "label" in filedoc.file_name:
				lnum = re.search("label(.*).png",filedoc.file_name)
				count = int(lnum.group(1))+1
				frappe.delete_doc('File',filedoc.name)
	namestr = doc.name + "-label{0}".format(count) + ".png"
	imgfile = frappe.get_doc({'doctype':'File','file_name':namestr,'attached_to_doctype':"Pick List",'attached_to_name':doc.name,"content":b64str,"decode":1})
	imgfile.insert()
def after_insert(doc,method):
	start_string = doc.name 
	end_string = '.png'
	label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Pick List' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
	if label_files:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Pick List' and attached_to_name=%s and file_name != %s""",
		(doc.name,label_files[0].file_name))
	else:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Pick List' and attached_to_name=%s""",
	(doc.name))
	# file_url = '/private/files/' + doc.name + '.pdf'
	# frappe.db.sql("""delete from `tabFile` where attached_to_doctype='Pick List' and attached_to_name=%s and file_url = %s""",
	# (doc.name,file_url))
	pdf_data=frappe.attach_print('Pick List',doc.name, print_format='Pick List Print With QR')
	
	_file = frappe.get_doc({
	"doctype": "File",
	"file_name": pdf_data.get('fname'),
	"attached_to_doctype": "Pick List",
	"attached_to_name": doc.name,
	"is_private": 1,
	"content": pdf_data.get('fcontent')
	})
	_file.save()

def validate(doc,method):
	if not doc.get("__islocal"):
		start_string = doc.name 
		end_string = '.png'
		label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Pick List' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
		if label_files:
			frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Pick List' and attached_to_name=%s and file_name != %s""",
			(doc.name,label_files[0].file_name))
		else:
			frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Pick List' and attached_to_name=%s""",
		(doc.name))
		# file_url = '/private/files/' + doc.name + '.pdf'
		# frappe.db.sql("""delete from `tabFile` where attached_to_doctype='Pick List' and attached_to_name=%s and file_url = %s""",
		# (doc.name,file_url))
		pdf_data=frappe.attach_print('Pick List',doc.name, print_format='Pick List Print With QR')
		
		_file = frappe.get_doc({
		"doctype": "File",
		"file_name": pdf_data.get('fname'),
		"attached_to_doctype": "Pick List",
		"attached_to_name": doc.name,
		"is_private": 1,
		"content": pdf_data.get('fcontent')
		})
		_file.save()

@frappe.whitelist()
def print_label(data,doc):
	data = json.loads(data)
	data1 = json.loads(doc)
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Pick List'},'url')
	public_file_path = url.split('app')
	public_file_path = public_file_path[0] + 'files/pick_list.xlsx'
	final_url = url + data1.get('name')
	file_path = os.path.realpath(get_files_path(is_private=0))
	file = file_path + '/' + 'pick_list.xlsx'
	barcode_doc = frappe.db.get_value("Barcodes",{'doctype_name':'Pick List'},'name')
	if barcode_doc:
		barcode_doc = frappe.get_doc("Barcodes",barcode_doc)
		if barcode_doc:
			barcode_doc.append('table_for_label_print',{
				'pick_list_name':data1.get('name')  ,
				'qty_of_finish_good_items':data1.get('for_qty'),
				'work_order':data1.get('work_order'),
				'no_of_copies':data.get('no_of_copies') ,
				'printer':data.get('printer_name'),
				'url':final_url 
				})
			barcode_doc.save()
	else:
		b_doc = frappe.new_doc("Barcodes")
		if b_doc:
			b_doc.doctype_name = "Pick List"
			b_doc.append('table_for_label_print',{
				'pick_list_name':data1.get('name')  ,
				'qty_of_finish_good_items':data1.get('for_qty'),
				'work_order':data1.get('work_order'),
				'no_of_copies':data.get('no_of_copies') ,
				'printer':data.get('printer_name'),
				'url':final_url 
				})
		b_doc.save()
	if os.path.exists(file):
		wb = openpyxl.load_workbook(filename=file)
		ws = wb['Sheet']
		row = ws.max_row +1
		# row = ws.get_highest_row() + 1
		col = 1
		cell = ws.cell(row=row,column=col)
		cell.value = ws.max_row -1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+1)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+2)
		cell.value = data1.get('work_order') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+3)
		cell.value = data1.get('for_qty') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+4)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = ws.cell(row=row,column=col+5)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+6)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		wb.save(file)
		frappe.msgprint(f'''Pick List Traveller Updated,You can download it from here
                        <a href = {public_file_path}><b>{public_file_path}</b></a>''')
		return public_file_path

	else:
		header = ['Record ID','Pick List Name','Work Order','Qty Of Finished Goods Items','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1

		cell = sheet.cell(row=row,column=col)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+1)
		cell.value = data1.get('name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = data1.get('work_order') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value = data1.get('for_qty') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+4)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+5)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+6)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	
		file = frappe.utils.get_site_path("public")
		fname = file_path+'/' + 'pick_list'+'.xlsx'
		book.save(fname)
		frappe.msgprint("Pick Traveler Created,You can download it from here {0}".format(public_file_path))
		return public_file_path