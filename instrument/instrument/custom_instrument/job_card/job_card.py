from __future__ import unicode_literals
import frappe
import json
import requests
from PIL import Image, ImageDraw
import io
import base64
import pyqrcode
import re

import json
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


import os
from frappe.utils import get_files_path
from pathlib import Path
from frappe.utils import getdate,time
import glob
from pathlib import Path
# import pandas as pd
from frappe.utils import flt, cstr, nowdate, nowtime
def after_insert(doc,method):
	start_string = doc.name 
	end_string = '.png'
	label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Job Card' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
	if label_files:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Job Card' and attached_to_name=%s and file_name != %s""",
		(doc.name,label_files[0].file_name))
	else:
		frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Job Card' and attached_to_name=%s""",
		(doc.name))
	# file_name = doc.name + '.pdf'
	# frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Job Card' and attached_to_name=%s and file_name = %s""",
	# 	(doc.name,file_name))
	pdf_data=frappe.attach_print('Job Card',doc.name, print_format='Job Card Print With QR')
	
	_file = frappe.get_doc({
	"doctype": "File",
	"file_name": pdf_data.get('fname'),
	"attached_to_doctype": "Job Card",
	"attached_to_name": doc.name,
	"is_private": 1,
	"content": pdf_data.get('fcontent')
	})
	_file.save()
def validate(doc,method):
	if not doc.get("__islocal"):
		start_string = doc.name 
		end_string = '.png'
		label_files = frappe.db.sql("""SELECT file_name from `tabFile` where attached_to_name = '{0}' and attached_to_doctype = 'Job Card' and file_name like '{1}%{2}'""".format(doc.name,start_string,end_string),as_dict=1)
		if label_files:
			frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Job Card' and attached_to_name=%s and file_name != %s""",
			(doc.name,label_files[0].file_name))
		else:
			frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Job Card' and attached_to_name=%s""",
			(doc.name))
		# file_name = doc.name + '.pdf'
		# frappe.db.sql("""DELETE from `tabFile` where attached_to_doctype='Job Card' and attached_to_name=%s and file_name = %s""",
		# 	(doc.name,file_name))
		pdf_data=frappe.attach_print('Job Card',doc.name, print_format='Job Card Print With QR')
		
		_file = frappe.get_doc({
		"doctype": "File",
		"file_name": pdf_data.get('fname'),
		"attached_to_doctype": "Job Card",
		"attached_to_name": doc.name,
		"is_private": 1,
		"content": pdf_data.get('fcontent')
		})
		_file.save()
@frappe.whitelist()
def get_engineering_revision(item_code,work_order):
	if item_code:
		engineering_revision = frappe.db.get_value("Item",{'name':item_code},'engineering_revision')
		er_from_wo = frappe.db.get_value("Work Order Item",{'parent':work_order,'item_code':item_code},'engineering_revision')
		if er_from_wo:
			return er_from_wo
		else:
			return engineering_revision
			
def label_img(doc,method):
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Job Card'},'url')
	final_string = url + doc.name
	img = Image.new('RGB', (384,192), color='white')
	qrc = pyqrcode.create(final_string)
	inmf = io.BytesIO()
	qrc.png(inmf,scale=6)
	qrcimg = Image.open(inmf)
	qrcimg.thumbnail((72,72))
	img.paste(qrcimg,(26,30))
	d = ImageDraw.Draw(img)
	d.multiline_text((120,35), "{0}\n\nProduction Item: {1}\nQty to Manufacture: {2}\n\nOperation: {3}\nWorkstation: {4}\nWork Order: {5}\nItem Name: {6}".format(doc.name,doc.production_item,doc.for_quantity,doc.operation,doc.workstation,doc.work_order,doc.item_name), fill=(0,0,0), spacing=1)
	d.text((30,160), "Job Card Traveler", fill=(0,0,0))
	barcode = requests.get('https://barcode.tec-it.com/barcode.ashx?data={0}&code=Code128&translate-esc=true'.format(doc.production_item))
	barc = Image.open(io.BytesIO(barcode.content))
	barc = barc.resize((220,15))
	img.paste(barc,(140,160))
	imgbuffer = io.BytesIO()
	img.save(imgbuffer, format='PNG')
	b64str = base64.b64encode(imgbuffer.getvalue())
	fname = frappe.db.get_list('File',filters={'attached_to_name':doc.name},fields=['name','file_name'])
	count=0
	if fname:
		for filedoc in fname:
			if "label" in filedoc.file_name:
				lnum = re.search("label(.*).png",filedoc.file_name)
				count = int(lnum.group(1))+1
				frappe.delete_doc('File',filedoc.name)
	namestr = doc.name + "-label{0}".format(count) + ".png"
	imgfile = frappe.get_doc({'doctype':'File','file_name':namestr,'attached_to_doctype':"Job Card",'attached_to_name':doc.name,"content":b64str,"decode":1})
	imgfile.insert()

@frappe.whitelist()
def print_label(data,doc):
	data = json.loads(data)
	data1 = json.loads(doc)
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Job Card'},'url')
	public_file_path = url.split('app')
	public_file_path = public_file_path[0] + 'files/job_card.xlsx'
	final_url = url + data1.get('name')
	file_path = os.path.realpath(get_files_path(is_private=0))
	file = file_path + '/' + 'job_card.xlsx'
	barcode_doc = frappe.db.get_value("Barcodes",{'doctype_name':'Job Card'},'name')
	if barcode_doc:
		barcode_doc = frappe.get_doc("Barcodes",barcode_doc)
		if barcode_doc:
			barcode_doc.append('table_for_label_print',{
				'job_card_name':data1.get('name')   ,
				'part_number':data1.get('production_item') ,
				'item_name':data1.get('item_name') ,
				'qty_to_manufacture':data1.get('for_quantity'),
				'operation':data1.get('operation'),
				'workstation':data1.get('workstation'),
				'work_order':data1.get('work_order'),
				'no_of_copies':data.get('no_of_copies') ,
				'printer':data.get('printer_name'),
				'url':final_url 
				})
			barcode_doc.save()
	else:
		b_doc = frappe.new_doc("Barcodes")
		if b_doc:
			b_doc.doctype_name = "Job Card"
			b_doc.append('table_for_label_print',{
				'job_card_name':data1.get('name')   ,
				'part_number':data1.get('production_item') ,
				'item_name':data1.get('item_name') ,
				'qty_to_manufacture':data1.get('for_quantity'),
				'operation':data1.get('operation'),
				'workstation':data1.get('workstation'),
				'work_order':data1.get('work_order'),
				'no_of_copies':data.get('no_of_copies') ,
				'printer':data.get('printer_name'),
				'url':final_url 
				})
		b_doc.save()
	if os.path.exists(file):
		wb = openpyxl.load_workbook(filename=file)
		ws = wb['Sheet']
		row = ws.max_row +1
		# row = ws.get_highest_row() + 1
		col = 1
		cell = ws.cell(row=row,column=col)
		cell.value = ws.max_row -1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+1)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+2)
		cell.value = data1.get('production_item') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+3)
		cell.value = data1.get('item_name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+4)
		cell.value = data1.get('for_quantity')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+5)
		cell.value = data1.get('operation')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+6)
		cell.value = data1.get('workstation')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+7)
		cell.value = data1.get('work_order')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+8)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = ws.cell(row=row,column=col+9)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+10)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		wb.save(file)
		frappe.msgprint(f'''Job Card Traveller Updated,You can download it from here
                        <a href = {public_file_path}><b>{public_file_path}</b></a>''')
		return public_file_path

	else:
		header = ['Record ID','Job Card Name','Production Item','Item Name','Qty To Manufacture','Operation','Workstation','Work Order','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1

		cell = sheet.cell(row=row,column=col)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+1)
		cell.value = data1.get('name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = data1.get('production_item') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value = data1.get('item_name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+4)
		cell.value = data1.get('for_quantity')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell = sheet.cell(row=row,column=col+5)
		cell.value = data1.get('operation')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+6)
		cell.value = data1.get('workstation')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+7)
		cell.value = data1.get('work_order')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+8)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+9)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+10)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	
		file = frappe.utils.get_site_path("public")
		fname = file_path+'/' + 'job_card'+'.xlsx'
		book.save(fname)
		frappe.msgprint("Job Card Traveller Created,You can download it from here {0}".format(public_file_path))
		return public_file_path




 