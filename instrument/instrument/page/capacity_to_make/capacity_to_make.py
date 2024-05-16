from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from frappe.utils import nowdate, cstr, flt, cint, now, getdate,get_datetime,time_diff_in_seconds,add_to_date,time_diff_in_seconds,add_days,today
from datetime import datetime,date, timedelta
from frappe.model.naming import make_autoname
import time
import json
# import pandas as pd
from frappe import _
from erpnext.manufacturing.doctype.bom.bom import get_bom_items_as_dict
import operator
import itertools
import datetime
from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from erpnext.manufacturing.doctype.bom.bom import get_children, validate_bom_no


@frappe.whitelist(allow_guest=True)
def get_capacity_data(filters=None):
	filters = json.loads(filters)
	if filters.get('delivery_date') <= today():
		frappe.throw("Expected Delivery Date Must be Greater Than Or Equal To Today")
	data = dict()
	precision=frappe.db.get_singles_value('System Settings', 'float_precision')
	production_item = filters.get("production_item")

	bom = frappe.db.get_value("BOM",{'is_default':1,'is_active':1,'item':filters.get('production_item')},'name')
	if bom:
		main_item_data = frappe.db.sql("""SELECT item as production_item,quantity as qty,uom ,name from `tabBOM` where name = '{0}'""".format(bom),as_dict=1)
		bom_data = []
		ohs_dict = get_ohs()
		for row in main_item_data:
			total_operation_time = frappe.db.sql("""SELECT sum(time_in_mins) as total_time from `tabBOM Operation` where parent = '{0}'""".format(bom),as_dict=1)
			if row.get('item_code') in ohs_dict:
				row.update({'ohs':ohs_dict.get(row.get('item_code'))})
			else:
				row.update({'ohs':0})
			if total_operation_time[0].get('total_time') :
				row.update({'total_operation_time':flt(total_operation_time[0].get('total_time')/60)})
			else:
				row.update({'total_operation_time':0})
			get_sub_assembly_item(row.get("name"), bom_data)
		
		for row in bom_data:
			if row.get('item_code') in ohs_dict:
				row.update({'ohs':ohs_dict.get(row.get('item_code'))})
			else:
				row.update({'ohs':0})
			main_item_data.append(row)
		for row in main_item_data:
			lead_time_days = frappe.db.get_value("Item",{'item_code':row.get('production_item')},'lead_time_days')
			row.update({'lead_time_days':lead_time_days})
			on_order_stock = get_on_order_stock(row.get('production_item'),filters.get('delivery_date'))
			if row.get('production_item') in on_order_stock:
				row.update({'on_order_stock':on_order_stock.get(row.get('production_item'))})
			else:
				row.update({'on_order_stock':0})


		
	start_dt = date.today()
	end_dt = filters.get('delivery_date')
	date_data = get_date_data(start_dt,end_dt)

	data.update({'date_data':date_data})
	data.update({'production_item_data':main_item_data})
	final_data = data
	path = 'instrument/instrument/page/capacity_to_make/capacity_to_make.html'
	html=frappe.render_template(path,{'data':data})
	return {'html':html,'data':final_data}

def get_date_data(start_dt,end_dt):
	date_list = []
	for dt in pd.date_range(start_dt, end_dt):
		date_list.append(dt.strftime("%Y-%m-%d"))
	return date_list

@frappe.whitelist()
def get_sub_assembly_item(bom_no, bom_data,indent=0):
	data = get_children('BOM', parent = bom_no)
	for d in data:
		if d.expandable:
			if d.value:
				total_operation_time = frappe.db.sql("""SELECT sum(time_in_mins) as total_time from `tabBOM Operation` where parent = '{0}'""".format(d.value),as_dict=1)
			bom_data.append(frappe._dict({
				'description': d.description,
				'production_item': d.item_code,
				'item_name': d.item_name,
				'stock_uom': d.stock_uom,
				'qty':d.stock_qty,
				'uom': d.stock_uom,
				'bom_no': d.value,
				'is_sub_contracted_item': d.is_sub_contracted_item,
				'bom_level': indent,
				'indent': indent,
				'is_subassembly':1,
				'total_operation_time':flt(total_operation_time[0].get('total_time')/60) if total_operation_time[0].get('total_time') else 0
			}))
			
			if d.value:
				get_sub_assembly_item(d.value, bom_data, indent=indent+1)
		else:
			bom_data.append(frappe._dict({
				'description': d.description,
				'production_item': d.item_code,
				'item_name': d.item_name,
				'stock_uom': d.stock_uom,
				'qty':d.stock_qty,
				'uom': d.stock_uom,
				'bom_no': d.value,
				'is_sub_contracted_item': d.is_sub_contracted_item,
				'bom_level': indent,
				'indent': indent,
				'is_subassembly':0,
				'total_operation_time':0
			}))
def get_ohs():
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse != '{0}' group by item_code """.format(wip_warehouse),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict


def get_on_order_stock(item_code,required_date):
	planned_po = frappe.db.sql("""SELECT poi.item_code,if(sum(poi.qty-poi.received_qty)>0,sum(poi.qty-poi.received_qty),0) as qty,poi.schedule_date from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.schedule_date <= '{0}' and qty > 0""".format(required_date),as_dict=1)
	# Manipulate in order to show in dict format
	if planned_po:
		on_order_stock = {item.item_code : item.qty for item in planned_po if item.item_code!=None}
	else:
		on_order_stock = dict()
	schedule_date_dict = {item.item_code : item.schedule_date for item in planned_po if item.item_code!=None}
	planned_mr = frappe.db.sql("""SELECT mri.item_code,if(sum(mri.qty)>0,sum(mri.qty),0) as qty ,mri.schedule_date from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.schedule_date <= '{0}' and mr.material_request_type = 'Purchase' and not exists (SELECT po.name from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.material_request = mr.name)""".format(required_date),as_dict=1)
	if planned_mr:
		for row in planned_mr:
			if row.get('item_code') != None:
				if row.get('item_code') in on_order_stock:
					qty = flt(on_order_stock.get(row.get('item_code'))) + flt(row.get('qty'))
					on_order_stock.update({row.get('item_code'):qty})
				else:
					on_order_stock.update({row.get('item_code'):flt(row.get('qty'))})
				if row.get('item_code') in schedule_date_dict:
					if row.get('schedule_date') > schedule_date_dict.get(row.get('item_code')):
						schedule_date_dict.update({row.get('item_code'):row.schedule_date})
				else:
					schedule_date_dict.update({row.get('item_code'):row.schedule_date})
	# Get planned qty from work order
	planned_wo = frappe.db.sql("""SELECT wo.production_item,(wo.qty-wo.produced_qty) as qty from `tabWork Order` wo where wo.planned_start_date <= '{0}'""".format(required_date),as_dict=1)
	if planned_wo:
		for row in planned_wo:
			if row.get('production_item') != None:
				if row.get('production_item') in on_order_stock:
					qty = flt(on_order_stock.get(row.get('production_item'))) + flt(row.get('qty'))
					on_order_stock.update({row.get('production_item'):qty})
				else:
					on_order_stock.update({row.get('production_item'):flt(row.get('qty'))})

	return on_order_stock

