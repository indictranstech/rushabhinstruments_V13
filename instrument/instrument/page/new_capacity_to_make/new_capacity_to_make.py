from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from frappe.utils import nowdate, cstr, flt, cint, now, getdate,get_datetime,time_diff_in_seconds,add_to_date,time_diff_in_seconds,add_days,today
from datetime import datetime,date, timedelta
from frappe.utils import nowdate,nowtime, today, flt,now

from frappe.model.naming import make_autoname
import time
import json
import pandas as pd
from frappe import _
from erpnext.manufacturing.doctype.bom.bom import get_bom_items_as_dict
import operator
import itertools
import datetime
from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from erpnext.manufacturing.doctype.bom.bom import get_children, validate_bom_no


@frappe.whitelist()
def get_sales_items():
	sales_items = frappe.get_all("Item",{'is_sales_item':1},'name')
	if sales_items:
		for item in sales_items:
			print("==========",item)
			filters = {'production_item':'B11844'}
			data = get_capacity_data(filters)
			break

@frappe.whitelist()
def get_capacity_data(filters=None):
	try:
		# filters = json.loads(filters)
		if 'delivery_date' in filters:
			if filters.get('delivery_date') <= today():
				frappe.throw("Expected Delivery Date Must be Greater Than Or Equal To Today")
		# Get all the in hand stock of all the warehouses excluding wip warehouse
		ohs_dict = get_ohs()
		bom = frappe.db.get_value("BOM",{'is_default':1,'is_active':1,'item':filters.get('production_item')},'name')
		std_lead_time = frappe.db.get_value("Item",{'item_code':filters.get('production_item')},'lead_time_days')
		if bom:
			calulated_lead_time_in_days = calculate_lead_time(bom)
			end_date_of_lead_time = date.today() + timedelta(calulated_lead_time_in_days)
		else:
			frappe.throw("There Is No Any Active & Default BOM Available for item {0}".format(filters.get('production_item')))
		# Case 1 : How much qty can be deliver today(Available stock not allocated for other work orders)
		# manipulate in order to show in erpnext and web
		case_1 = dict()
		case_1['type'] = 'In Stock'
		case_1['std_lead_time'] = std_lead_time
		case_1['calulated_lead_time_in_days'] =calulated_lead_time_in_days
		case_1['date_available'] = today()
		case_1['remark'] = '{0} Qty Can Be Ready To Deliver'.format(ohs_dict.get(filters.get('production_item'))) if ohs_dict.get(filters.get('production_item')) else 'There is no Current Stock Available'
		case_1['qty'] = ohs_dict.get(filters.get('production_item')) if ohs_dict.get(filters.get('production_item')) else 0
		data = []
		data.append(case_1)
		
		bom_childs = []
		bom_child_list = get_child_boms(bom,bom_childs)
		bom_child_list.append({'bom' : bom})
		
		# find all the bom in descending order (bom level)
		final_bom_list = get_all_boms_in_order(bom_child_list)
		final_bom_list = [item.get('bom') for item in final_bom_list]
		# Case 2
		# Earliest Delivery date and qty(Post quick assembly)
		expected_delivery = date.today() + timedelta(calulated_lead_time_in_days)
		if bom:
			# all_bom_items = get_exploded_items(bom)
			bom_doc = frappe.get_doc("BOM",bom)
			main_bom_item = frappe.db.get_value("BOM",bom,'item')
			all_bom_items = get_all_bom_items(final_bom_list)
			all_bom_items = [i.get('item_code') for i in all_bom_items]
			final_item_dict = dict()
			parent_bom_items = get_raw_bom_data(bom)
			assembly_time = get_assembly_time(bom)
			post_quick_assembly = []
			remaining_qty = dict()
			for row in parent_bom_items:
				if row.item_code in ohs_dict:
					ohs_qty = ohs_dict.get(row.item_code)
					qty = flt(ohs_qty/row.raw_qty)
					post_quick_assembly.append(qty)
			if post_quick_assembly:
				post_quick_qty = min(post_quick_assembly)
			else:
				post_quick_qty = 0
			if post_quick_qty > 0:
				for row in parent_bom_items:
					if row.item_code in ohs_dict:
						ohs_qty = ohs_dict.get(row.item_code)
						rem_qty = flt(ohs_qty - (row.raw_qty*qty))
						ohs_dict[row.item_code] = rem_qty
			case_2_1 = dict()
			case_2_1['type'] = 'Earliast Delivery Date & Qty(Post Quick Assembly)'
			case_2_1['std_lead_time'] = std_lead_time
			case_2_1['calulated_lead_time_in_days'] =calulated_lead_time_in_days
			case_2_1['date_available'] = date.today() + timedelta(assembly_time)
			case_2_1['qty'] = flt(post_quick_qty)
			case_2_1['remark'] = '{0} Qty Can Be Assembed Quick'.format(flt(post_quick_qty)) if flt(post_quick_qty) > 0 else 'There is No Any Current Stock for Quick Assembly'
			data.append(case_2_1)
			# Case 2
			# Earliest Delivery date and qty(Production)
			date_range = date.today() + timedelta(calulated_lead_time_in_days)
			todays_date = date.today()
			all_bom_items = "', '".join(all_bom_items)
			on_order_stock = get_on_order_stock_for_rm(date_range,all_bom_items)
			date_qty_dict = dict()
			max_date = date.today()
			if on_order_stock :
				for i in on_order_stock:
					if on_order_stock.get(i).get('schedule_date') > max_date:
						max_date = on_order_stock.get(i).get('schedule_date')
			final_item_dict = dict()
			for bom in final_bom_list:
				qty_can_be_produced = []
				bom_item = frappe.db.get_value("BOM",bom,'item')
				bom_items = get_raw_bom_data(bom)
				for item in bom_items:
					if ohs_dict and item.item_code in ohs_dict :
						available_quantity = ohs_dict.get(item.item_code)
					else:
						available_quantity = 0
					if on_order_stock and  item.item_code in on_order_stock:
						available_quantity = available_quantity + flt(on_order_stock.get(item.item_code).get('qty'))	
					if item.item_code in final_item_dict:
						available_quantity = final_item_dict.get(item.item_code).get('qty')
					proportion_qty = flt((available_quantity)/item.get('raw_qty'))
					qty_can_be_produced.append(proportion_qty)
				if qty_can_be_produced :
					qty_can_be_produced = min(qty_can_be_produced)
					final_item_dict[bom_item] = {'schedule_date':max_date,'qty':qty_can_be_produced if qty_can_be_produced else 0}
			final_qty = (flt(final_item_dict.get(main_bom_item).get('qty')) if flt(final_item_dict.get(main_bom_item).get('qty')) > 0 else 0)
			
			case_3 = dict()
			case_3['type'] = 'Qty Can Be Manufactured Before Calculated Lead Time'
			case_3['std_lead_time'] = std_lead_time
			case_3['calulated_lead_time_in_days'] =calulated_lead_time_in_days
			case_3['date_available'] = max_date if flt(final_item_dict.get(main_bom_item).get('qty')) > 0 else end_date_of_lead_time
			case_3['remark'] = '{0} Can Be Manufactured Before {1}'.format(final_qty,end_date_of_lead_time) if final_qty > 0 else 'There is no enough material in stock and currently ordered which will available before {0}'.format(end_date_of_lead_time)
			case_3['qty'] = final_qty
			data.append(case_3)

			# Case 4
			total_qty = flt(ohs_dict.get(filters.get('production_item'))) + flt(post_quick_qty) + final_qty
			case_4 = dict()
			case_4['type'] = 'Max Qty Before Calculated Lead Time'
			case_4['std_lead_time'] = std_lead_time
			case_4['calulated_lead_time_in_days'] =calulated_lead_time_in_days
			case_4['date_available'] = max_date if flt(final_item_dict.get(main_bom_item).get('qty')) > 0 else end_date_of_lead_time
			case_4['remark'] = '{0} Can Be Ready To Deliver Before {1}'.format(flt(total_qty),end_date_of_lead_time) if flt(total_qty) > 0 else 'There is no enough material in stock and currently ordered which will available before {0}'.format(end_date_of_lead_time)
			case_4['qty'] = flt(total_qty)
			
			if calulated_lead_time_in_days:
				date_list = get_date_list(calulated_lead_time_in_days)
				case_4['date_list'] = date_list
				new_table_data = get_new_table_data(bom,std_lead_time,calulated_lead_time_in_days)
				case_4['new_table_data'] = new_table_data
			data.append(case_4)
			return data
	except Exception as e:
		traceback = frappe.get_traceback()
		frappe.log_error(
			title=_("Error in Capacity To Make"),
			message=traceback,
		)
		raise e

@frappe.whitelist()
def capacity_data(filters=None):
	filters = json.loads(filters)
	data = get_capacity_data(filters)
	path = 'instrument/instrument/page/new_capacity_to_make/new_capacity_to_make.html'
	html=frappe.render_template(path,{'data':data})
	return {'html':html,'data':data}

def get_date_list(calulated_lead_time_in_days):
	start_date = date.today()
	end_date = date.today() + timedelta(calulated_lead_time_in_days)
	date_list = [start_date + datetime.timedelta(days=x) for x in range(calulated_lead_time_in_days)]
	final_date_list = []
	for d in date_list:
		final_date_list.append(str(d)) 
	return final_date_list
@frappe.whitelist()
def get_sub_assembly_items_final(bom_no, bom_data_new,indent=1):
	data = get_children('BOM', parent = bom_no)
	item_dict = {item.item_code:item.stock_qty for item in data}
	final_data = []
	count = 0
	for item in item_dict:
		total = 0
		count = count + 1
		for info in data:
			if item ==  info.get('item_code'):
				total = total + info.get('stock_qty')
				item_info = info
		item_info.stock_qty = total
		final_data.append(item_info)
	for d in final_data:
		if d.expandable:
			operation_time = frappe.db.sql("""SELECT sum(bo.time_in_mins) as lead_time from `tabBOM Operation` bo join `tabBOM` b on b.name = bo.parent where b.name = '{0}' """.format(bom_no),as_dict=1)
			operation_time = flt(operation_time[0].get('lead_time'))/flt(60)
			bom_data_new.append(frappe._dict({	
				'item': d.item_code,
				'item_name': d.item_name,
				'uom': d.stock_uom,
				'indent': indent,
				'qty': d.stock_qty,
				'type':"Assembly Manufactured Via Work Order",
				'is_subassembly':1,
				'operation_time':operation_time if operation_time else 0,
				'bom_no':d.value

			}))
			if d.value:
				get_sub_assembly_items_final(d.value, bom_data_new, indent=indent+1)
		else:
			bom_data_new.append(frappe._dict({	
				'item': d.item_code,
				'item_name': d.item_name,
				'uom': d.stock_uom,
				'indent': indent,
				'qty': d.stock_qty,
				'type':'Raw Material Purchased from Supplier',
				'is_subassembly':0,
				'operation_time':0

			}))
def get_new_table_data(bom,std_lead_time,calulated_lead_time_in_days):
	bom_data_new = []
	doc = frappe.get_doc("BOM",bom)
	operation_time = frappe.db.sql("""SELECT sum(bo.time_in_mins) as lead_time from `tabBOM Operation` bo join `tabBOM` b on b.name = bo.parent where b.name = '{0}' """.format(bom),as_dict=1)
	operation_time = flt(operation_time[0].get('lead_time'))/flt(60)
	if doc:
		bom_data_new.append(frappe._dict({	
				'item': doc.item,
				'item_name': doc.item_name,
				'uom': doc.uom,
				'indent': 0,
				'qty': doc.quantity,
				'type':"Assembly Manufactured Via Work Order",
				'is_subassembly':1,
				'operation_time':operation_time if operation_time else 0,
				'bom_no':bom
			}))
	get_sub_assembly_items_final(bom,bom_data_new)
	ohs_dict = get_ohs()
	for row in bom_data_new:
		std_lead_time = frappe.db.get_values("Item",{'item_code':row.get('item')},['lead_time_days','item_name'],as_dict=1,debug=0)
		row['std_lead_time'] = std_lead_time[0].get('lead_time_days') if std_lead_time[0].get('lead_time_days') else 0
		row['item_name'] = std_lead_time[0].get('item_name') if std_lead_time[0].get('item_name') else ''
		if row.get('item') in ohs_dict:
			row['ohs'] = ohs_dict.get(row.get('item'))
		else:
			row['ohs'] = 0
		date_range = date.today() + timedelta(calulated_lead_time_in_days)
		item_list = [row.get('item')]
		item_list = "', '".join(item_list)
		ordered_dict = get_on_order_stock_for_rm(date_range,item_list)
		row['ordered_qty'] = ordered_dict.get(row.get('item')).get('qty') if ordered_dict else 0
		max_days =  (ordered_dict.get(row.get('item')).get('schedule_date')-date.today()).days if ordered_dict else 0
		row['max_days'] = max_days
	calculate_day_wise_qty(bom,ohs_dict,bom_data_new,calulated_lead_time_in_days)
	return bom_data_new

def calculate_day_wise_qty(bom,ohs_dict,bom_data_new,calulated_lead_time_in_days):
	date_list = get_date_list(calulated_lead_time_in_days)
	date_wise_data = dict()
	all_bom_items = [i.get('item') for i in bom_data_new]
	all_bom_items = "', '".join(all_bom_items)
	on_order_stock_before = get_on_order_stock_before(today(),all_bom_items)
	bom_childs = []
	bom_child_list = get_child_boms(bom,bom_childs)
	bom_child_list.append({'bom' : bom})
	final_bom_list = get_all_boms_in_order_desc(bom_child_list)
	final_bom_list = [item.get('bom') for item in final_bom_list]
	if date_list:
		for date in date_list:
			sub_assembly_item_dict= dict()
			qty_can_be_produced = []
			raw_item_dict= dict()
			for row in reversed(bom_data_new):
				if row.bom_no:
					bom_items = get_bom_items(row.bom_no)
					sub_item_list = [row.get('item')]
					sub_item_list = "', '".join(sub_item_list)
					qty_can_be_produced_raw = []
					for item in bom_items:
						if item.get('item') in raw_item_dict:
							qty_can_be_produced_raw.append(raw_item_dict.get(item.get('item')))
						else:
							qty_can_be_produced_raw.append(0)
					
					final_qty = (min(qty_can_be_produced_raw) + (ohs_dict.get(row.get('item')) if row.get('item') in ohs_dict else 0))
					ohs_dict[row.get('item')] = 0
					row[str(date)] = final_qty
					raw_item_dict[row.get('item')] = final_qty
					if final_qty > 0:
						ohs_dict[row.get('item')] = 0
						if on_order_stock_before and  row.get('item') in on_order_stock_before: 
							on_order_stock_before[row.get('item')]['qty'] = 0
						if on_order_stock and  row.get('item') in on_order_stock: 
							on_order_stock[row.get('item')]['qty']=0 
				else:
					item_list = [row.get('item')]
					item_list = "', '".join(item_list)
					on_order_stock = get_on_order_stock_day(date,item_list)
					if ohs_dict and row.get('item') in ohs_dict :
						available_quantity = ohs_dict.get(row.get('item'))
					else:
						available_quantity = 0
					if on_order_stock_before and  row.get('item') in on_order_stock_before:
						available_quantity = available_quantity + flt(on_order_stock_before.get(row.get('item')).get('qty'))
					if on_order_stock and  row.get('item') in on_order_stock:
						available_quantity = available_quantity + flt(on_order_stock.get(row.get('item')).get('qty'))
					if row.get('item') in sub_assembly_item_dict:
						available_quantity = available_quantity + sub_assembly_item_dict.get(row.get('item'))
					proportion_qty = flt(available_quantity)/flt(row.get('qty'))
					row[str(date)] = proportion_qty
					raw_item_dict[row.get('item')] = proportion_qty
					if proportion_qty > 0:
						ohs_dict[row.get('item')] = 0
						if on_order_stock_before and  row.get('item') in on_order_stock_before: 
							on_order_stock_before[row.get('item')]['qty'] = 0
						if on_order_stock and  row.get('item') in on_order_stock: 
							on_order_stock[row.get('item')]['qty']=0 
			# calculate_remaining_qty(date,bom_data_new,ohs_dict,on_order_stock,on_order_stock_before)
	return bom_data_new
def calculate_remaining_qty(date,bom_data_new,ohs_dict,on_order_stock,on_order_stock_before):
	# print("================bom_data_new",bom_data_new[0])
	if bom_data_new[0]:
		if bom_data_new[0].get(date):
			qty = flt(bom_data_new[0].get(date)) - flt(bom_data_new[0].get('ohs'))
			if qty > 0:
				bom_items = get_bom_items(bom_data_new[0].get('bom_no'))
			else:
				pass

def get_bom_items(bom):
	if bom:
		bom_items = frappe.db.sql("""SELECT i.item_code as item ,sum(i.qty) as qty,i.bom_no from `tabBOM Item` i join `tabBOM` b on b.name=i.parent where b.name = '{0}' group by item_code""".format(bom),as_dict=1)
		return bom_items
def get_parent_bom_detials(bom,raw_data):
	doc = frappe.get_doc("BOM",{'name':bom})
	operation_time = frappe.db.sql("""SELECT sum(bo.time_in_mins) as lead_time from `tabBOM Operation` bo join `tabBOM` b on b.name = bo.parent where b.name = '{0}' """.format(bom),as_dict=1)
	operation_time = flt(operation_time[0].get('lead_time'))/flt(60)
	if doc:
		raw_data.append({
			'item':doc.item,
			'qty':doc.quantity,
			'uom':doc.uom,
			'operation_time':operation_time if operation_time else 0,
			'type':"Assembly Manufactured Via Work Order",
			'is_subassembly':1
		})
		return raw_data

@frappe.whitelist()
def get_raw_items(bom,raw_data):
	doc = frappe.get_doc("BOM",{'name':bom})
	if doc:
		for row in doc.items:
			# raw_data.append({
			# 		'item':row.item_code,
			# 		'qty':row.qty,
			# 		'type':"Assembly Manufactured Via Work Order" if row.bom_no else 'Raw Material Purchased from Supplier'
			# 	})
			# stock_qty = (row.qty*qty)
			if row.bom_no:
				operation_time = frappe.db.sql("""SELECT sum(bo.time_in_mins) as lead_time from `tabBOM Operation` bo join `tabBOM` b on b.name = bo.parent where b.name = '{0}' """.format(bom),as_dict=1)
				operation_time = flt(operation_time[0].get('lead_time'))/flt(60)
				raw_data.append({
					'item':row.item_code,
					'qty':row.qty,
					'uom':row.stock_uom,
					'operation_time':operation_time if operation_time else 0,
					'type':"Assembly Manufactured Via Work Order" if row.bom_no else 'Raw Material Purchased from Supplier',
					'is_subassembly':1
				})

			else:
				raw_data.append({
					'item':row.item_code,
					'qty':row.qty,
					'uom':row.stock_uom,
					'operation_time':0,
					'type':'Raw Material Purchased from Supplier',
					'is_subassembly':0
				})
		# print("=============raw_data",raw_data)
		return raw_data	
def calculate_lead_time(bom):
	# Considered following factors to calculate_lead_time
	# 1) MAx lead time for all the raw material item to purchase
	# 2) Total operation time required for all the subassembly and fg items
	# 3) Total Makeup days required to make ready subasembly or FG item.
	if bom:
		bom_childs = []
		bom_child_list = get_child_boms(bom,bom_childs)
		bom_child_list.append({'bom' : bom})
		final_bom_list = get_all_boms_in_order(bom_child_list)
		final_bom_list = [item.get('bom') for item in final_bom_list]
		joined_bom_list = "', '".join(final_bom_list)
		raw_items = frappe.db.sql("""SELECT boi.item_code  from `tabBOM Item` boi where boi.parent in  ('{0}') and boi.bom_no = ''""".format(joined_bom_list),as_dict=1)
		raw_items = [item.item_code for item in raw_items]
		joined_item_list = "', '".join(raw_items)
		raw_item_lead_time = frappe.db.sql("""SELECT max(lead_time_days) as raw_lead_time from `tabItem` where item_code in ('{0}')""".format(joined_item_list),as_dict=1)
		lead_time = frappe.db.sql("""SELECT sum(bo.time_in_mins) as lead_time from `tabBOM Operation` bo join `tabBOM` b on b.name = bo.parent where b.name in ('{0}') """.format(joined_bom_list),as_dict=1)
		make_up_days = frappe.db.sql("""SELECT sum(b.makeup_days) as makeup_days from `tabBOM` b where b.name in ('{0}') """.format(joined_bom_list),as_dict=1)

		operation_time = flt(lead_time[0].get('lead_time'))/flt(60)
		calculate_lead_time = flt(flt(make_up_days[0].get('makeup_days')) + flt(raw_item_lead_time[0].get('raw_lead_time')),0)
		return calculate_lead_time

def get_assembly_time(bom):
	assembly_time = frappe.db.get_value("BOM",{'name':bom},'makeup_days')
	return assembly_time
# Get all child mapped_bom
def get_child_boms(bom,bom_childs):
	if bom:
		childs = frappe.db.sql("""SELECT bom_no as bom from `tabBOM Item` where parent = '{0}' and bom_no is not null """.format(bom),as_dict=1)

		bom_childs += childs

		if len(childs)>0:
			for c in childs:
				get_child_boms(c.get('bom'),bom_childs)
		return bom_childs
# Get all the boms in order
def get_all_boms_in_order(bom_childs):
	if len(bom_childs)>1:
		final_list = [row.get('bom') for row in bom_childs if row.get("bom")]

		final_list = '(' + ','.join("'{}'".format(i) for i in final_list) + ')'
		childs = frappe.db.sql("""SELECT b.name as bom,b.bom_level from `tabBOM` b where b.name in {0} order by b.bom_level asc""".format(final_list),as_dict=1)
		return childs
# Get all the boms in order
def get_all_boms_in_order_desc(bom_childs):
	if len(bom_childs)>1:
		final_list = [row.get('bom') for row in bom_childs if row.get("bom")]

		final_list = '(' + ','.join("'{}'".format(i) for i in final_list) + ')'
		childs = frappe.db.sql("""SELECT b.name as bom,b.bom_level from `tabBOM` b where b.name in {0} order by b.bom_level desc""".format(final_list),as_dict=1)
		return childs
def get_all_bom_items(final_bom_list):
	all_bom_items = []
	if final_bom_list:
		for bom in final_bom_list:
			bom_data = get_raw_bom_data(bom)
			for item in bom_data:
				all_bom_items.append(item)
	return all_bom_items
def get_raw_bom_data(bom):
	if bom:
		raw_bom_data = frappe.db.sql("""SELECT b.item as subassembly_item,b.quantity,boi.item_code,boi.qty as raw_qty from `tabBOM` b join `tabBOM Item` boi on b.name = boi.parent where b.name = '{0}' """.format(bom),as_dict=1)
		if raw_bom_data:
			return raw_bom_data
def get_on_order_stock_for_rm(date_range,all_bom_items):
	planned_po = frappe.db.sql("""SELECT p.name,poi.item_code,sum(poi.qty-poi.received_qty) as qty ,max(poi.schedule_date) as schedule_date from `tabPurchase Order Item` poi join `tabPurchase Order` p on p.name=poi.parent where (poi.qty-poi.received_qty) > 0 and p.docstatus =1 and poi.schedule_date <= '{0}' and poi.item_code in ('{1}') group by poi.item_code""".format(date_range,all_bom_items),as_dict=1)
	ordered_dict = dict()
	if planned_po != []:
		for row in planned_po:
			ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':row.get('qty')}
		# for row in planned_po:
		# 	if row.get("item_code") in ordered_dict:
		# 		updated_data = ordered_dict.get(row.get("item_code"))
		# 		if row.get("schedule_date") in updated_data:
		# 			updated_data[row.get("schedule_date")] = updated_data[row.get("schedule_date")] + row.get("qty",0)
		# 		else:
		# 			updated_data[row.get("schedule_date")] = row.get("qty",0)
		# 	else:
		# 		ordered_dict[row.get("item_code")] = {
		# 			row.get("schedule_date"): row.get("qty", 0)
				
		# 		}
	planned_mr = frappe.db.sql("""SELECT mri.item_code,if(sum(mri.qty)>0,sum(mri.qty),0) as qty ,mri.schedule_date from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.schedule_date <= '{0}' and mr.material_request_type in ('Purchase','Manufacture') and not exists (SELECT po.name from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.material_request = mr.name) and not exists (SELECT wo.name from `tabWork Order` wo where wo.material_request = mr.name) and mri.item_code in ('{0}')  group by mri.item_code""".format(date_range,all_bom_items),as_dict=1)
	if planned_mr:
		for row in planned_mr:
			if row.get('item_code') in ordered_dict:
				check_date = ordered_dict.get(row.get("item_code")).get('schedule_date')
				qty = flt(row.get('qty')) + flt(ordered_dict.get(row.get("item_code")).get('qty'))
				if row.get('schedule_date') > check_date:
					ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':qty}
				else:
					ordered_dict[row.get("item_code")] = {'schedule_date':check_date,'qty':qty}
			else:
				ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':row.get('qty')}

	planned_wo = frappe.db.sql("""SELECT wo.name,wo.production_item as item_code,sum(wo.qty-wo.produced_qty) as qty ,wo.planned_end_date from `tabWork Order` wo  where  (wo.qty-wo.produced_qty) > 0 and wo.docstatus =1 and wo.production_item in  ('{0}') group by wo.production_item,wo.planned_end_date <= '{0}'""".format(date_range,all_bom_items),as_dict=1)

	if planned_wo != []:
		for row in planned_wo:
			if row.get('planned_end_date'):
				planned_end_date = row.get("planned_end_date").date()
				# print("============planned_end_date",row.get("planned_end_date").date())
				if row.get("item_code") in ordered_dict:
					check_date = ordered_dict.get(row.get("item_code")).get('schedule_date')
					qty = flt(row.get('qty')) + flt(ordered_dict.get(row.get("item_code")).get('qty'))
					if row.get("planned_end_date").date() > check_date:
						ordered_dict[row.get("item_code")] = {'schedule_date':row.get("planned_end_date").date(),'qty':qty}
					else:
						ordered_dict[row.get("item_code")] = {'schedule_date':check_date,'qty':qty}
				else:
					ordered_dict[row.get("item_code")] = {'schedule_date':row.get("planned_end_date").date(),'qty':row.get('qty')}

			# if row.get("item_code") in ordered_dict:
			# 	updated_data = ordered_dict.get(row.get("item_code"))
			# 	if planned_end_date in updated_data:
			# 		updated_data[planned_end_date] = updated_data[planned_end_date] + row.get("qty",0)
			# 	else:
			# 		updated_data[planned_end_date] = row.get("qty",0)
			# else:
			# 	ordered_dict[row.get("item_code")] = {
			# 		planned_end_date: row.get("qty", 0)
				
				# }	
	if ordered_dict:
		return ordered_dict
def get_on_order_stock_day(date_range,all_bom_items):
	planned_po = frappe.db.sql("""SELECT p.name,poi.item_code,sum(poi.qty-poi.received_qty) as qty ,max(poi.schedule_date) as schedule_date from `tabPurchase Order Item` poi join `tabPurchase Order` p on p.name=poi.parent where (poi.qty-poi.received_qty) > 0 and p.docstatus =1 and poi.schedule_date = '{0}' and poi.item_code in ('{1}') group by poi.item_code""".format(date_range,all_bom_items),as_dict=1)
	ordered_dict = dict()
	if planned_po != []:
		for row in planned_po:
			ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':row.get('qty')}
		# for row in planned_po:
		# 	if row.get("item_code") in ordered_dict:
		# 		updated_data = ordered_dict.get(row.get("item_code"))
		# 		if row.get("schedule_date") in updated_data:
		# 			updated_data[row.get("schedule_date")] = updated_data[row.get("schedule_date")] + row.get("qty",0)
		# 		else:
		# 			updated_data[row.get("schedule_date")] = row.get("qty",0)
		# 	else:
		# 		ordered_dict[row.get("item_code")] = {
		# 			row.get("schedule_date"): row.get("qty", 0)
				
		# 		}
	planned_mr = frappe.db.sql("""SELECT mri.item_code,if(sum(mri.qty)>0,sum(mri.qty),0) as qty ,mri.schedule_date from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.schedule_date = '{0}' and mr.material_request_type in ('Purchase','Manufacture') and not exists (SELECT po.name from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.material_request = mr.name) and not exists (SELECT wo.name from `tabWork Order` wo where wo.material_request = mr.name) and mri.item_code in ('{0}')  group by mri.item_code""".format(date_range,all_bom_items),as_dict=1)
	if planned_mr:
		for row in planned_mr:
			if row.get('item_code') in ordered_dict:
				check_date = ordered_dict.get(row.get("item_code")).get('schedule_date')
				qty = flt(row.get('qty')) + flt(ordered_dict.get(row.get("item_code")).get('qty'))
				if row.get('schedule_date') > check_date:
					ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':qty}
				else:
					ordered_dict[row.get("item_code")] = {'schedule_date':check_date,'qty':qty}
			else:
				ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':row.get('qty')}

	planned_wo = frappe.db.sql("""SELECT wo.name,wo.production_item as item_code,sum(wo.qty-wo.produced_qty) as qty ,wo.planned_end_date from `tabWork Order` wo  where  (wo.qty-wo.produced_qty) > 0 and wo.docstatus =1 and wo.production_item in  ('{0}') group by wo.production_item,wo.planned_end_date = '{0}'""".format(date_range,all_bom_items),as_dict=1)

	if planned_wo != []:
		for row in planned_wo:
			if row.get('planned_end_date'):
				planned_end_date = row.get("planned_end_date").date()
				# print("============planned_end_date",row.get("planned_end_date").date())
				if row.get("item_code") in ordered_dict:
					check_date = ordered_dict.get(row.get("item_code")).get('schedule_date')
					qty = flt(row.get('qty')) + flt(ordered_dict.get(row.get("item_code")).get('qty'))
					if row.get("planned_end_date").date() > check_date:
						ordered_dict[row.get("item_code")] = {'schedule_date':row.get("planned_end_date").date(),'qty':qty}
					else:
						ordered_dict[row.get("item_code")] = {'schedule_date':check_date,'qty':qty}
				else:
					ordered_dict[row.get("item_code")] = {'schedule_date':row.get("planned_end_date").date(),'qty':row.get('qty')}

			# if row.get("item_code") in ordered_dict:
			# 	updated_data = ordered_dict.get(row.get("item_code"))
			# 	if planned_end_date in updated_data:
			# 		updated_data[planned_end_date] = updated_data[planned_end_date] + row.get("qty",0)
			# 	else:
			# 		updated_data[planned_end_date] = row.get("qty",0)
			# else:
			# 	ordered_dict[row.get("item_code")] = {
			# 		planned_end_date: row.get("qty", 0)
				
				# }	
	if ordered_dict:
		return ordered_dict
def get_on_order_stock_before(date_range,all_bom_items):
	planned_po = frappe.db.sql("""SELECT p.name,poi.item_code,sum(poi.qty-poi.received_qty) as qty ,max(poi.schedule_date) as schedule_date from `tabPurchase Order Item` poi join `tabPurchase Order` p on p.name=poi.parent where (poi.qty-poi.received_qty) > 0 and p.docstatus =1 and poi.schedule_date < '{0}' and poi.item_code in ('{1}') group by poi.item_code""".format(date_range,all_bom_items),as_dict=1)
	ordered_dict = dict()
	if planned_po != []:
		for row in planned_po:
			ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':row.get('qty')}
		# for row in planned_po:
		# 	if row.get("item_code") in ordered_dict:
		# 		updated_data = ordered_dict.get(row.get("item_code"))
		# 		if row.get("schedule_date") in updated_data:
		# 			updated_data[row.get("schedule_date")] = updated_data[row.get("schedule_date")] + row.get("qty",0)
		# 		else:
		# 			updated_data[row.get("schedule_date")] = row.get("qty",0)
		# 	else:
		# 		ordered_dict[row.get("item_code")] = {
		# 			row.get("schedule_date"): row.get("qty", 0)
				
		# 		}
	planned_mr = frappe.db.sql("""SELECT mri.item_code,if(sum(mri.qty)>0,sum(mri.qty),0) as qty ,mri.schedule_date from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.schedule_date < '{0}' and mr.material_request_type in ('Purchase','Manufacture') and not exists (SELECT po.name from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.material_request = mr.name) and not exists (SELECT wo.name from `tabWork Order` wo where wo.material_request = mr.name) and mri.item_code in ('{0}')  group by mri.item_code""".format(date_range,all_bom_items),as_dict=1)
	if planned_mr:
		for row in planned_mr:
			if row.get('item_code') in ordered_dict:
				check_date = ordered_dict.get(row.get("item_code")).get('schedule_date')
				qty = flt(row.get('qty')) + flt(ordered_dict.get(row.get("item_code")).get('qty'))
				if row.get('schedule_date') > check_date:
					ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':qty}
				else:
					ordered_dict[row.get("item_code")] = {'schedule_date':check_date,'qty':qty}
			else:
				ordered_dict[row.get("item_code")] = {'schedule_date':row.get('schedule_date'),'qty':row.get('qty')}

	planned_wo = frappe.db.sql("""SELECT wo.name,wo.production_item as item_code,sum(wo.qty-wo.produced_qty) as qty ,wo.planned_end_date from `tabWork Order` wo  where  (wo.qty-wo.produced_qty) > 0 and wo.docstatus =1 and wo.production_item in  ('{0}') group by wo.production_item,wo.planned_end_date < '{0}'""".format(date_range,all_bom_items),as_dict=1)

	if planned_wo != []:
		for row in planned_wo:
			if row.get('planned_end_date'):
				planned_end_date = row.get("planned_end_date").date()
				# print("============planned_end_date",row.get("planned_end_date").date())
				if row.get("item_code") in ordered_dict:
					check_date = ordered_dict.get(row.get("item_code")).get('schedule_date')
					qty = flt(row.get('qty')) + flt(ordered_dict.get(row.get("item_code")).get('qty'))
					if row.get("planned_end_date").date() > check_date:
						ordered_dict[row.get("item_code")] = {'schedule_date':row.get("planned_end_date").date(),'qty':qty}
					else:
						ordered_dict[row.get("item_code")] = {'schedule_date':check_date,'qty':qty}
				else:
					ordered_dict[row.get("item_code")] = {'schedule_date':row.get("planned_end_date").date(),'qty':row.get('qty')}

			# if row.get("item_code") in ordered_dict:
			# 	updated_data = ordered_dict.get(row.get("item_code"))
			# 	if planned_end_date in updated_data:
			# 		updated_data[planned_end_date] = updated_data[planned_end_date] + row.get("qty",0)
			# 	else:
			# 		updated_data[planned_end_date] = row.get("qty",0)
			# else:
			# 	ordered_dict[row.get("item_code")] = {
			# 		planned_end_date: row.get("qty", 0)
				
				# }	
	if ordered_dict:
		return ordered_dict
def get_date_data(start_dt,end_dt):
	date_list = []
	for dt in pd.date_range(start_dt, end_dt):
		date_list.append(dt.strftime("%Y-%m-%d"))
	return date_list

@frappe.whitelist()
def get_sub_assembly_item(bom_no, bom_data,indent=0):
	data = get_children('BOM', parent = bom_no)
	for d in data:
		if d.expandable:
			if d.value:
				total_operation_time = frappe.db.sql("""SELECT sum(time_in_mins) as total_time from `tabBOM Operation` where parent = '{0}'""".format(d.value),as_dict=1)
			bom_data.append(frappe._dict({
				'description': d.description,
				'production_item': d.item_code,
				'item_name': d.item_name,
				'stock_uom': d.stock_uom,
				'qty':d.stock_qty,
				'uom': d.stock_uom,
				'bom_no': d.value,
				'is_sub_contracted_item': d.is_sub_contracted_item,
				'bom_level': indent,
				'indent': indent,
				'is_subassembly':1,
				'total_operation_time':flt(total_operation_time[0].get('total_time')/60) if total_operation_time[0].get('total_time') else 0
			}))
			
			if d.value:
				get_sub_assembly_item(d.value, bom_data, indent=indent+1)
		else:
			bom_data.append(frappe._dict({
				'description': d.description,
				'production_item': d.item_code,
				'item_name': d.item_name,
				'stock_uom': d.stock_uom,
				'qty':d.stock_qty,
				'uom': d.stock_uom,
				'bom_no': d.value,
				'is_sub_contracted_item': d.is_sub_contracted_item,
				'bom_level': indent,
				'indent': indent,
				'is_subassembly':0,
				'total_operation_time':0
			}))
def get_ohs():
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse != '{0}' and actual_qty > 0 group by item_code """.format(wip_warehouse),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict
def get_ordered_stock(required_date):
	planned_po = frappe.db.sql("""SELECT poi.item_code,if(sum(poi.qty-poi.received_qty)>0,sum(poi.qty-poi.received_qty),0) as qty,poi.schedule_date from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.schedule_date = '{0}' and qty > 0  and poi.item_code = '{1}' group by poi.item_code""".format(required_date,item_code),as_dict=1)
	if planned_po:
		ordered_dict = {item.item_code:item.qty for item in planned_po}
	else:
		ordered_dict = dict()

	planned_mr = frappe.db.sql("""SELECT mri.item_code,if(sum(mri.qty)>0,sum(mri.qty),0) as qty ,mri.schedule_date from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.schedule_date = '{0}' and mr.material_request_type in ('Purchase','Manufacture') and not exists (SELECT po.name from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.material_request = mr.name) and not exists (SELECT wo.name from `tabWork Order` wo where wo.material_request = mr.name) and mri.item_code = '{1}' group by mri.item_code""".format(required_date,item_code),as_dict=1)
	if planned_mr:
		for row in planned_mr:
			if row.item_code in ordered_dict:
				ordered_dict[row.item_code] = ordered_dict.get('item_code') + row.qty
			else:
				ordered_dict[row.item_code] = row.qty
	planned_wo = frappe.db.sql("""SELECT wo.production_item as item_code,sum(wo.qty-wo.produced_qty) as qty from `tabWork Order` wo where wo.planned_end_date = '{0}' and wo.production_item = '{1}' group by item_code""".format(todays_date,item_code),as_dict=1)
	if planned_wo:
		for row in planned_wo:
			if row.item_code in ordered_dict:
				ordered_dict[row.item_code] = ordered_dict.get('item_code') + row.qty
			else:
				ordered_dict[row.item_code] = row.qty
	return ordered_dict

def get_on_order_stock(item_code,required_date):
	planned_po = frappe.db.sql("""SELECT poi.item_code,if(sum(poi.qty-poi.received_qty)>0,sum(poi.qty-poi.received_qty),0) as qty,poi.schedule_date from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.schedule_date <= '{0}' and qty > 0""".format(required_date),as_dict=1)
	# Manipulate in order to show in dict format
	if planned_po:
		on_order_stock = {item.item_code : item.qty for item in planned_po if item.item_code!=None}
	else:
		on_order_stock = dict()
	schedule_date_dict = {item.item_code : item.schedule_date for item in planned_po if item.item_code!=None}
	planned_mr = frappe.db.sql("""SELECT mri.item_code,if(sum(mri.qty)>0,sum(mri.qty),0) as qty ,mri.schedule_date from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.schedule_date <= '{0}' and mr.material_request_type = 'Purchase' and not exists (SELECT po.name from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.material_request = mr.name)""".format(required_date),as_dict=1)
	if planned_mr:
		for row in planned_mr:
			if row.get('item_code') != None:
				if row.get('item_code') in on_order_stock:
					qty = flt(on_order_stock.get(row.get('item_code'))) + flt(row.get('qty'))
					on_order_stock.update({row.get('item_code'):qty})
				else:
					on_order_stock.update({row.get('item_code'):flt(row.get('qty'))})
				if row.get('item_code') in schedule_date_dict:
					if row.get('schedule_date') > schedule_date_dict.get(row.get('item_code')):
						schedule_date_dict.update({row.get('item_code'):row.schedule_date})
				else:
					schedule_date_dict.update({row.get('item_code'):row.schedule_date})
	# Get planned qty from work order
	planned_wo = frappe.db.sql("""SELECT wo.production_item,(wo.qty-wo.produced_qty) as qty from `tabWork Order` wo where wo.planned_start_date <= '{0}'""".format(required_date),as_dict=1)
	if planned_wo:
		for row in planned_wo:
			if row.get('production_item') != None:
				if row.get('production_item') in on_order_stock:
					qty = flt(on_order_stock.get(row.get('production_item'))) + flt(row.get('qty'))
					on_order_stock.update({row.get('production_item'):qty})
				else:
					on_order_stock.update({row.get('production_item'):flt(row.get('qty'))})

	return on_order_stock

@frappe.whitelist()
def make_xlsx_file(renderd_data):
	data =json.loads(renderd_data)
	
	header = ['Item Code','BOM Qty','In Stock','Ordered Qty','Max Days','Lead Time (days)','Operation Time (hours)According to BOM Qty','Type']

	if data[3].get('date_list'):
		for date in data[3].get('date_list'):
			header.append(date)
	
	book = Workbook()
	sheet = book.active
	
	row = 1
	col = 1

	for item in header:
		cell = sheet.cell(row=row,column=col)
		cell.value = item
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
		
		col+=1

	row = 2
	col = 1

	for data_row in data[3].get('new_table_data'):
		cell = sheet.cell(row=row,column=col)
		cell.value = data_row.get('item')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)


		cell = sheet.cell(row=row,column=col+1)
		cell.value = data_row.get('qty')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+2)
		cell.value = data_row.get('ohs')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+3)
		cell.value = data_row.get('ordered_qty')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+4)
		cell.value = data_row.get('max_days')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+5)
		cell.value = data_row.get('std_lead_time')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+6)
		cell.value = data_row.get('operation_time')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+7)
		cell.value = data_row.get('type')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)


		cols = 9
		for date in data[3].get('date_list'):
			cell = sheet.cell(row=row,column=cols)
			cell.value = data_row.get(date)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

			cols+=1
		row += 1

	file_path = frappe.utils.get_site_path("public")
	fname = "Capacity_To_Make" + nowdate() + ".xlsx"
	book.save(file_path+fname)
	return  file_path+fname

@frappe.whitelist()
def download_xlsx():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	# now = datetime.now()
	fname = "Capacity_To_Make" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname


@frappe.whitelist(allow_guest=True)
def capacity_to_make_api(production_item=None):
    filters= {'production_item':production_item}
    data = get_capacity_data(filters)
    final_data = []
    instock = {'in_stock':data[0].get('qty'),'date':data[0].get('date_available')}
    qty = data[3].get('new_table_data')[0].get(data[0].get('date_available'))-data[0].get('qty')
    data[3]['new_table_data'][0][data[0].get('date_available')] = qty
    final_data.append(instock)
    planned_wo  = frappe.db.sql("""SELECT sum(wo.qty) as qty,date(wo.planned_end_date) as planned_end_date from `tabWork Order` wo where  wo.planned_end_date <= '{0}' and wo.production_item = '{1}' and wo.status not in ('Completed','Stopped') group by wo.planned_end_date""".format(data[2].get('date_available'),production_item),as_dict =1)
    in_progress = []
    for row in planned_wo:
    	in_progress.append({'in_progress_qty':row.get('qty'),'expected_completion_date':row.get('planned_end_date')})
    final_data.append(in_progress)
    in_built = []
    for row in data[3].get('new_table_data')[0]:
    	for date in data[3].get('date_list'):
    		if row == date and data[3].get('new_table_data')[0].get(row) > 0:
    			in_built.append({'qty_can_be_built':data[3].get('new_table_data')[0].get(row),'expected_date':date})
    final_data.append(in_built)
    final_data.append({
    	"lead_time_in_calender_days":data[0].get('calulated_lead_time_in_days')
    	})
    return(final_data)