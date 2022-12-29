# Copyright (c) 2021, instrument and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from erpnext.manufacturing.doctype.bom.bom import get_bom_items_as_dict
from erpnext.stock.doctype.item.item import get_item_defaults
# from erpnext.stock.doctype.batch.batch import get_batch_no
from frappe import _
from frappe.utils import cint, flt, get_link_to_form

import json
from datetime import datetime
from frappe.utils import nowdate,nowtime, today, flt
from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from erpnext.manufacturing.doctype.bom.bom import get_children, validate_bom_no
from frappe.utils import nowdate,nowtime, today, flt
from datetime import timedelta,date
import datetime
import calendar
import time



import os
from frappe.utils import get_files_path
from pathlib import Path
from frappe.utils import getdate,time
import glob
from pathlib import Path
import pandas as pd
from frappe.utils import flt, cstr, nowdate, nowtime
class WorkOrderPickList(Document):
	@frappe.whitelist()
	def get_fg_work_orders(self):
		self.work_order_table = ''
		fg_item_groups = frappe.db.sql("""SELECT item_group from `tabTable For Item Group`""",as_dict=1)
		todays_date = datetime.date.today()
		fg_item_groups = [item.get('item_group') for item in fg_item_groups]
		if fg_item_groups:
			if self.purpose == "Material Transfer for Manufacture":
				fg_work_orders = frappe.db.sql("""SELECT wo.name,(wo.qty - wo.material_transferred_for_manufacturing) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where i.item_group in {0} and wo.planned_start_date >= '{1}' and wo.status in ('Not Started','In Process') and wo.skip_transfer = 0""".format(tuple(fg_item_groups),todays_date),as_dict=1,debug=1)
				if fg_work_orders:
					for row in fg_work_orders:
						doc = frappe.get_doc("Work Order",row.get("name"))
						ohs = get_current_stock()
						qty_will_be_produced_list = []
						for item in doc.required_items:
							if item.item_code in ohs:
								if item.required_qty <= ohs.get(item.item_code):
									percent_stock = 100
									qty_will_be_produced = item.required_qty
									qty_will_be_produced_list.append(qty_will_be_produced)
								elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0: 
									percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
									qty_will_be_produced = (percent_stock/100*item.required_qty)
									qty_will_be_produced_list.append(qty_will_be_produced)
								else : 
									percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
									qty_will_be_produced = 0
									qty_will_be_produced_list.append(qty_will_be_produced)
						
						row['qty_will_be_produced'] =min(qty_will_be_produced_list)
						self.append('work_order_table',{
							'work_order':row.get('name'),
							'pending_qty':row.get('pending_qty'),
							'qty_can_be_pulledmanufactured' :row.get('qty_will_be_produced'),
							'total_qty_of_finish_good_on_work_order':row.get('qty'),
							'qty_of_finished_goods_already_completed':row.get('produced_qty')
							})
			else:
				fg_work_orders = frappe.db.sql("""SELECT wo.name,(wo.qty - wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where i.item_group in {0} and wo.planned_start_date >= '{1}' and wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and wo.produced_qty < wo.material_transferred_for_manufacturing and wo.material_transferred_for_manufacturing > 0""".format(tuple(fg_item_groups),todays_date),as_dict=1,debug=1)
				fg_work_orders_2 = frappe.db.sql("""SELECT wo.name,(wo.qty - wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where i.item_group in {0} and wo.planned_start_date >= '{1}' and wo.status in ('Not Started','In Process') and wo.skip_transfer = 1 and wo.produced_qty < wo.qty """.format(tuple(fg_item_groups),todays_date),as_dict=1,debug=1)
				if fg_work_orders_2:
					for item in fg_work_orders_2:
						fg_work_orders.append(item)
				if fg_work_orders:
					for row in fg_work_orders:
						doc = frappe.get_doc("Work Order",row.get("name"))
						if doc.skip_transfer == 1 and doc.from_wip_warehouse == 1:
							ohs = get_current_stock_for_manufacture()
						elif doc.skip_transfer == 1:
							ohs = get_current_stock()
						else:
							ohs = get_current_stock()
						qty_will_be_produced_list = []
						for item in doc.required_items:
							if item.item_code in ohs:
								if item.required_qty <= ohs.get(item.item_code):
									percent_stock = 100
									qty_will_be_produced = item.required_qty
									qty_will_be_produced_list.append(qty_will_be_produced)
								elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0: 
									percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
									qty_will_be_produced = (percent_stock/100*item.required_qty)
									qty_will_be_produced_list.append(qty_will_be_produced)
								else : 
									percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
									qty_will_be_produced = 0
									qty_will_be_produced_list.append(qty_will_be_produced)
						
						row['qty_will_be_produced'] =min(qty_will_be_produced_list)
						self.append('work_order_table',{
							'work_order':row.get('name'),
							'pending_qty':row.get('pending_qty'),
							'qty_can_be_pulledmanufactured' :row.get('qty_will_be_produced'),
							'total_qty_of_finish_good_on_work_order':row.get('qty'),
							'qty_of_finished_goods_already_completed':row.get('produced_qty')
							})
		else:
			frappe.throw("Please Enter Item Groups for FG in Rushabh Settings")

	@frappe.whitelist()
	def get_sub_assembly_items(self, manufacturing_type=None):
		self.sub_assembly_items = []
		for row in self.work_order_table:
			bom = frappe.db.get_value("Work Order",{'name':row.work_order},'bom_no')
			planned_start_date = frappe.db.get_value("Work Order",{'name':row.work_order},'planned_start_date')
			bom_data = []
			get_sub_assembly_items(bom, bom_data, row.pending_qty)
			bom_qty_dict = {item.production_item : item.stock_qty for item in bom_data}
			sub_assembly_items = [item.production_item for item in bom_data]
			final_subassembly_items = []
			if self.consider_current_stock == 1:
				wip_warehouse = frappe.db.get_value("Work Order",{'name':row.work_order},'wip_warehouse')
				current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse = '{0}' and actual_qty > 0 group by item_code """.format(wip_warehouse),as_dict=1)
				ohs_dict = {item.item_code : item.qty for item in current_stock}
				for col in sub_assembly_items:
					if col not in ohs_dict:
						final_subassembly_items.append(col)
					elif (ohs_dict.get(col) < bom_qty_dict.get(col)):
						final_subassembly_items.append(col)
				self.get_work_orders_for_subassembly(final_subassembly_items,planned_start_date)
			else:
				self.get_work_orders_for_subassembly(sub_assembly_items,planned_start_date)
			
			self.append('work_orders',{
				'work_order':row.work_order,
				'qty_of_finished_goods_to_pull':row.pending_qty,
				'total_qty_to_of_finished_goods_on_work_order':row.total_qty_of_finish_good_on_work_order,
				'qty_of_finished_goods_already_completed':row.qty_of_finished_goods_already_completed,
				'qty_of_finished_goods':row.qty_can_be_pulledmanufactured
				})


	def get_work_orders_for_subassembly(self,sub_assembly_items,planned_start_date):
		todays_date = datetime.date.today()
		if self.purpose == 'Material Transfer for Manufacture':
			work_orders = frappe.db.sql("""SELECT name,(qty-produced_qty) as pending_qty ,qty,produced_qty from `tabWork Order` where production_item in {0} and planned_start_date <= '{1}' and docstatus =1 and status in ('Not Started','In Process') order by name desc""".format(tuple(sub_assembly_items),planned_start_date.date()),as_dict=1,debug=1)
		else:
			work_orders =frappe.db.sql("""SELECT wo.name,(wo.qty-wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo where wo.production_item in {0} and wo.planned_start_date <= '{1}' and wo.docstatus =1 and wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and wo.produced_qty < wo.material_transferred_for_manufacturing and wo.material_transferred_for_manufacturing > 0""".format(tuple(sub_assembly_items),planned_start_date.date()),as_dict=1,debug=1)
			work_order_1 =frappe.db.sql("""SELECT wo.name,(wo.qty-wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo where wo.production_item in {0} and wo.planned_start_date <= '{1}' and wo.docstatus =1 and wo.status in ('Not Started','In Process') and wo.skip_transfer = 1 and wo.produced_qty < wo.qty """.format(tuple(sub_assembly_items),planned_start_date.date()),as_dict=1,debug=1)
			if work_order_1:
				for item in work_order_1:
					work_orders.append(item)
		if work_orders:
			for row in work_orders:
				doc = frappe.get_doc("Work Order",row.get("name"))
				if doc.skip_transfer == 1 and doc.from_wip_warehouse == 1:
					ohs = get_current_stock_for_manufacture()
				elif doc.skip_transfer == 1:
					ohs = get_current_stock()
				else:
					ohs = get_current_stock()
				qty_will_be_produced_list = []
				for item in doc.required_items:
					if item.item_code in ohs:
						if item.required_qty <= ohs.get(item.item_code):
							percent_stock = 100
							qty_will_be_produced = item.required_qty
							qty_will_be_produced_list.append(qty_will_be_produced)
						elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0: 
							percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
							qty_will_be_produced = (percent_stock/100*item.required_qty)
							qty_will_be_produced_list.append(qty_will_be_produced)
						else : 
							percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
							qty_will_be_produced = 0
							qty_will_be_produced_list.append(qty_will_be_produced)
					else:
						qty_will_be_produced = 0
						qty_will_be_produced_list.append(qty_will_be_produced)
				row['qty_will_be_produced'] =min(qty_will_be_produced_list)
				self.append('work_orders',{
					'work_order':row.get('name'),
					'qty_of_finished_goods_to_pull':row.get('pending_qty'),
					'total_qty_to_of_finished_goods_on_work_order':row.get('qty'),
					'qty_of_finished_goods_already_completed':row.get('produced_qty'),
					'qty_of_finished_goods':row.get('qty_will_be_produced')
					})

	def validate(self):
		if self.get("__islocal"):
			final_item_list = []
			for row in self.work_orders:
				final_item_list.append({
				'work_order' :row.get("work_order"),
				'total_qty_to_of_finished_goods_on_work_order':row.get("total_qty_to_of_finished_goods_on_work_order"),
				'qty_of_finished_goods_to_pull':row.get("qty_of_finished_goods_to_pull"),
				'qty_of_finished_goods_already_completed':row.get("qty_of_finished_goods_already_completed"),
				'qty_of_finished_goods':row.get("qty_of_finished_goods")
				})
			final_data = sorted(final_item_list,key = lambda x:x["work_order"],reverse=True)
			self.work_orders = ''
			count = 1
			for col in final_data:
				col.update({'idx' : count})
				self.append("work_orders",col)
				count = count + 1


	@frappe.whitelist()
	def get_work_order_items(self):
		final_raw_item_list = []
		
		final_data = dict()
		if self.work_orders:
			for item in self.work_orders:
				item_list = []
				bom_no = frappe.db.get_value("Work Order",{'name':item.work_order},'bom_no')
				wo_doc = frappe.get_doc("Work Order",item.work_order)
				if bom_no:
					# Get all the required raw materials with required qty according to qty of qty_of_finished_goods
					raw_materials = get_raw_material(bom_no,self.company,item.qty_of_finished_goods_to_pull,item.work_order)
					final_raw_item_list.append(raw_materials)
					i_list = [item for item in raw_materials]
					for i in i_list:
						item_list.append(i)
					final_item_list= list(set(item_list))
					# Get all the avaialble locations for required items
					if self.purpose == 'Material Transfer for Manufacture':
						item_locations_dict = get_item_locations(self,final_item_list,self.company)
					if self.purpose == 'Manufacture':
						if wo_doc.skip_transfer == 1 and wo_doc.from_wip_warehouse == 0:
							item_locations_dict = get_item_locations(self,final_item_list,self.company)
						else:
							item_locations_dict = get_item_locations_for_manufacture(self,final_item_list,self.company)

					
					# Manipulate in order to show in table
					for row in raw_materials:
						engineering_revision = frappe.db.get_value("Work Order Item",{'parent':item.work_order,'item_code':row},'engineering_revision')

						if row in item_locations_dict:
							col = item_locations_dict.get(row)
							for i in col:
								# Fetch batch  FIFO basis
								# batch_no = get_batch_no(row, i.get("warehouse"),raw_materials.get(row).get("qty"), throw=False)
								transferred_qty = frappe.db.get_value("Work Order Item",{'parent':item.work_order,'item_code':row},'transferred_qty')
								i['required_qty'] = (flt(raw_materials.get(row).get("qty"))-flt(transferred_qty)) 
								i['picked_qty'] = 0
								i['stock_uom'] = raw_materials.get(row).get('stock_uom')
								i['engineering_revision'] = engineering_revision
								# i['batch_no'] = batch_no
								# Prefill the picked qty
								# if batch_no:
								# 	batch_qty = frappe.db.get_value("Batch",{'name':batch_no},'batch_qty')
								# 	if batch_qty >= (flt(raw_materials.get(row).get("qty"))-flt(transferred_qty)) :
								# 		i['picked_qty'] = (flt(raw_materials.get(row).get("qty"))-flt(transferred_qty)) 
								# 	else:
								# 		i['picked_qty'] = batch_qty	
					final_data[item.work_order] = item_locations_dict
			# add items in item_locations
			work_order_list = [item.work_order for item in self.work_orders]
			for work_order in work_order_list:
				for row in final_data:
					if row == work_order:
						final_row = final_data.get(row)
						for i in final_row:
							item_data = get_item_defaults(i, self.company)
							j = final_row.get(i)
							for d in j:
								if (d.get("required_qty")) > 0:
									self.append("work_order_pick_list_item",{
										"item_code": i,
										"item_name": item_data.get("item_name"),
										"warehouse":d.get("warehouse"),
										"required_qty":d.get("required_qty"),
										"stock_qty": d.get("qty"),
										"work_order" : row,
										"picked_qty" : d.get('picked_qty'),
										"uom" : d.get('stock_uom'),
										"stock_uom":d.get('stock_uom'),
										"description" :item_data.get('description'),
										"item_group" : item_data.get("item_group"),
										"engineering_revision" : d.get("engineering_revision"),
										# "batch_no" : d.get('batch_no')
									})
			return self.batch_assignment_fifo()
		# self.save()
	# def batch_assignment_fifo(self):
	# 	item_list = [item.item_code for item in self.work_order_pick_list_item]
	# 	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
		
	# 	batch_data = frappe.db.sql("""SELECT b.name,b.item,`tabStock Ledger Entry`.warehouse, sum(`tabStock Ledger Entry`.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` ignore index (item_code, warehouse) on (b.name = `tabStock Ledger Entry`.batch_no ) where `tabStock Ledger Entry`.item_code in {0}  and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) and `tabStock Ledger Entry`.warehouse != '{1}'  group by batch_id order by b.expiry_date ASC, b.creation ASC""".format(tuple(item_list),wip_warehouse),as_dict=1,debug=1)
	# 	allocated_item_list = []
	# 	allocated_item_dict = dict()
	# 	for row in self.work_order_pick_list_item:
	# 		if row.item_code not in allocated_item_dict:
	# 			for col in batch_data:	
	# 				if row.item_code == col.item and row.warehouse == col.warehouse :
						
	# 					if col.qty >= row.required_qty:
	# 						row.batch_no = col.name
	# 						row.picked_qty = row.required_qty
	# 						allocated_item_list.append(row.item_code)
	# 						allocated_item_dict.update({row.item_code:row.work_order})
	# 						if frappe.get_cached_value('Item', row.item_code, 'has_serial_no') == 1:
	# 							serial_nos = get_serial_no_batchwise(row.item_code,col.name,col.warehouse,row.required_qty)
	# 							row.serial_no = serial_nos 
	# 						break
	# 					elif col.qty < row.required_qty:
	# 						row.batch_no = col.name
	# 						row.picked_qty = col.qty
	# 						allocated_item_list.append(row.item_code)
	# 						# allocated_item_dict.update({row.item_code:row.work_order})
	# 						if frappe.get_cached_value('Item', row.item_code, 'has_serial_no') == 1:
	# 							serial_nos = get_serial_no_batchwise(row.item_code,col.name,col.warehouse,row.required_qty)
	# 							row.serial_no = serial_nos
	# 						break

	def batch_assignment_fifo(self):
		item_list = [item.item_code for item in self.work_order_pick_list_item]
		wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')

		batch_data = frappe.db.sql("""SELECT b.name,b.item, `tabStock Ledger Entry`.warehouse, sum(`tabStock Ledger Entry`.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` ignore index (item_code, warehouse) on (b.name = `tabStock Ledger Entry`.batch_no ) where `tabStock Ledger Entry`.item_code in {0}  and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) and `tabStock Ledger Entry`.warehouse != '{1}'  group by batch_id order by b.expiry_date ASC, b.creation ASC""".format(tuple(item_list),wip_warehouse),as_dict=1,debug=0)

		for row in batch_data:
			calculate_remaining_batch_qty(row)

		work_list = []
		for row in self.work_order_pick_list_item:
			work_list.append({"item_code":row.item_code, "uom":row.uom, "uom_conversion_factor":row.uom_conversion_factor, "stock_uom":row.stock_uom, "serial_nos":row.serial_no, "warehouse":row.warehouse, "required_qty":row.required_qty, "work_order":row.work_order, "stock_qty":0, "picked_qty":0})
		
		new_list=[]
		allocated_item_dict = dict()
		allocated_war_item_dict = dict()
		for row in work_list:
		    wo_item = row.get("item_code")+row.get("work_order")
		    war_item = row.get("work_order")+row.get("item_code")+row.get("warehouse")
		    for col in batch_data:
		        if row.get("item_code")==col.get("item") and row.get("warehouse")==col.get("warehouse") and wo_item not in allocated_item_dict:
		            item_qty=sum([r.get("picked_qty") for r in new_list if new_list and row.get("item_code")==r.get("item_code") and r.get("item_code")+r.get("work_order")==wo_item])            
		            rem_reqd_qty = abs(item_qty-row.get("required_qty"))
		            row.update({"required_qty":rem_reqd_qty})
		            
		            if col.get("qty")>= row.get("required_qty"):
		                batch_qty = col.get("qty") - row.get("required_qty") 
		                new_list.append({
		                    "item_code":row.get("item_code"), 
		                    "warehouse":row.get("warehouse"), 
		                    "required_qty":row.get("required_qty"),
		                    "work_order":row.get("work_order"),
		                    "stock_qty":col.get("qty"),
		                    "picked_qty":row.get("required_qty"),
		                    "batch_no": col.get("name"),
							"uom":row.get("uom"), 
							"uom_conversion_factor":row.get("uom_conversion_factor"), 
							"stock_uom":row.get("stock_uom")
		                })
		                col.update({"qty":batch_qty})
		                allocated_item_dict.update({wo_item:wo_item})
		            else:
		                reqd_qty = row.get("required_qty")-col.get("qty")
		                if col.get("qty") > 0:
		                    new_list.append({
		                        "item_code":row.get("item_code"), 
		                        "warehouse":row.get("warehouse"), 
		                        "required_qty":row.get("required_qty"),
		                        "work_order":row.get("work_order"),
		                        "stock_qty":col.get("qty"),
		                        "picked_qty":col.get("qty"),
		                        "batch_no": col.get("name"),
		                        "uom":row.get("uom"), 
								"uom_conversion_factor":row.get("uom_conversion_factor"), 
								"stock_uom":row.get("stock_uom")
		                    })
		                elif war_item not in allocated_war_item_dict:
		                    new_list.append({
		                        "item_code":row.get("item_code"), 
		                        "warehouse":row.get("warehouse"), 
		                        "required_qty":row.get("required_qty"),
		                        "work_order":row.get("work_order"),
		                        "stock_qty": 0,
		                        "picked_qty": 0,
		                        "batch_no": "",
		                        "uom":row.get("uom"), 
								"uom_conversion_factor":row.get("uom_conversion_factor"), 
								"stock_uom":row.get("stock_uom")
		                    })
		                    allocated_war_item_dict.update({war_item:war_item})
		                col.update({"qty":0})		
		return new_list

	def on_submit(self):
		allocated_batch_data = frappe.db.sql("""SELECT b.item_code,b.picked_qty,b.batch_no from `tabWork Order Pick List Item` b join `tabWork Order Pick List` wo on wo.name = b.parent""",as_dict=1,debug=0)
		allocated_batch_data_dict = {item.batch_no:item.picked_qty for item in allocated_batch_data}
		allocated_batch_list = []
		for row in self.work_order_pick_list_item:
			if row.batch_no in allocated_batch_data_dict:
				allocated_batch_list.append(row.idx)
		# if len(allocated_batch_list) > 0:
		# 	frappe.throw("Kindly Review batches for rows {0}".format(allocated_batch_list))

def calculate_remaining_batch_qty(row):
	used_qty = frappe.db.sql("""SELECT item_code, sum(picked_qty) as picked_qty From `tabWork Order Pick List Item` where batch_no='{0}' and item_code='{1}' and docstatus=1""".format(row.name, row.item), as_dict=1)
	if used_qty:
		row["qty"] = abs(row.qty - flt(used_qty[0].get('picked_qty')))

def get_serial_nos(item_code,batch_id):
	serial_nos = frappe.db.sql("""SELECT name from `tabSerial No` where batch_no = '{0}' and item_code = '{1}'""".format(batch_id,item_code),as_dict=1)
	serial_nos = [item.name for item in serial_nos]
	return serial_nos
def get_serial_no_batchwise(item_code,batch_no,warehouse,qty):
	if frappe.db.get_single_value("Stock Settings", "automatically_set_serial_nos_based_on_fifo"):
		return "\n".join(frappe.db.sql_list("""SELECT name from `tabSerial No`
			where item_code=%(item_code)s and warehouse=%(warehouse)s 
			and batch_no=IF(%(batch_no)s IS NULL, batch_no, %(batch_no)s) order
			by timestamp(purchase_date, purchase_time) asc limit %(qty)s""", {
				"item_code": item_code,
				"warehouse": warehouse,
				"batch_no": batch_no,
				"qty": abs(cint(qty))
			}))

def get_item_locations(self,item_list,company):
	item_locations_dict = dict()
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", "default_wip_warehouse")
	warehouses = [x.get('name') for x in frappe.get_list("Warehouse", {'company': company}, "name")]
	if wip_warehouse in warehouses:
		warehouses.remove(wip_warehouse)
	for item in item_list :
		item_locations = frappe.get_all('Bin',
							fields=['warehouse', 'actual_qty as qty'],
							filters={'item_code':item,
							'actual_qty': ['>', 0],
							'warehouse' :['in', warehouses]
							},
							order_by='creation')
		item_locations_dict[item] = item_locations
	return item_locations_dict
def get_item_locations_for_manufacture(self,item_list,company):
	item_locations_dict = dict()
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", "default_wip_warehouse")
	# warehouses = [x.get('name') for x in frappe.get_list("Warehouse", {'company': company}, "name")]
	# if wip_warehouse in warehouses:
	# 	warehouses.remove(wip_warehouse)
	for item in item_list :
		item_locations = frappe.get_all('Bin',
							fields=['warehouse', 'actual_qty as qty'],
							filters={'item_code':item,
							'actual_qty': ['>', 0],
							'warehouse' :['=', wip_warehouse]
							},
							order_by='creation')
		item_locations_dict[item] = item_locations
	return item_locations_dict	
def get_raw_material(bom_no, company, qty,work_order):
	try:
		work_order_doc = frappe.get_doc("Work Order",work_order)
		if work_order_doc.use_multi_level_bom == 1:
			raw_materials = get_bom_items_as_dict(bom_no,company,qty,fetch_exploded=1)
		else:
			raw_materials = get_bom_items_as_dict(bom_no,company,qty,fetch_exploded=0)
		for item in raw_materials:
			item_info = raw_materials.get(item)
			item_info['work_order'] = work_order
		return raw_materials
	except Exception as e:
		frappe.log_error(message = frappe.get_traceback() , title = "Get Raw Material")

@frappe.whitelist()
def get_work_orders(production_plan):
	if production_plan:
		work_order_data = frappe.db.sql("""SELECT wo.name,wo.qty,wo.produced_qty,(wo.qty-wo.produced_qty) as pending_qty from `tabWork Order` wo where production_plan = '{0}' and (wo.produced_qty < wo.qty)  and wo.status in ('In process','Not Started') and wo.docstatus = 1 order by wo.bom_level asc""".format(production_plan),as_dict =1,debug=1)
		return work_order_data

@frappe.whitelist()
def get_work_order_data(work_order):
	if work_order:
		work_order_data = frappe.db.sql("""SELECT wo.name,wo.qty,wo.produced_qty,(wo.qty-wo.produced_qty) as pending_qty from `tabWork Order` wo where wo.name = '{0}' and (wo.produced_qty < wo.qty)  and wo.status in ('In process','Not Started') and wo.docstatus = 1 order by wo.name asc""".format(work_order),as_dict =1,debug=1)
		doc = frappe.get_doc("Work Order",work_order)
		ohs = get_current_stock()
		qty_will_be_produced_list = []
		for item in doc.required_items:
			if item.item_code in ohs:
				if item.required_qty <= ohs.get(item.item_code):
					percent_stock = 100
					qty_will_be_produced = item.required_qty
					qty_will_be_produced_list.append(qty_will_be_produced)
				elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0: 
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					qty_will_be_produced = (percent_stock/100*item.required_qty)
					qty_will_be_produced_list.append(qty_will_be_produced)
				else : 
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					qty_will_be_produced = 0
					qty_will_be_produced_list.append(qty_will_be_produced)
		
		work_order_data[0]['qty_will_be_produced'] =min(qty_will_be_produced_list)
		return work_order_data
def get_current_stock():
	# 1.get wip warehouse
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", 'default_wip_warehouse')
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse != '{0}' group by item_code """.format(wip_warehouse),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict
def get_current_stock_for_manufacture():
	# 1.get wip warehouse
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", 'default_wip_warehouse')
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse = '{0}' group by item_code """.format(wip_warehouse),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict	
@frappe.whitelist()
def validate_picked_qty(work_order,required_qty,picked_qty,doc_name,row_name,item_code):
	picked_qty = frappe.db.sql("""SELECT sum(picked_qty) as picked_qty from `tabWork Order Pick List Item` where parent = '{0}' and work_order = '{1}' and item_code = '{2}' and name != '{3}'""".format(doc_name,work_order,item_code,row_name),as_dict=1)
	return picked_qty[0].get('picked_qty')

@frappe.whitelist()
def check_stock_entries(work_order,work_order_pick_list):
	if work_order and work_order_pick_list:
		stock_entry_list = frappe.db.sql("""SELECT name from `tabStock Entry` where work_order = '{0}' and work_order_pick_list = '{1}'""".format(work_order,work_order_pick_list),as_dict=1)
		if len(stock_entry_list)==0:
			return True
		else:
			frappe.msgprint("Stock Entries already created for work order {0}".format(work_order))
@frappe.whitelist()
def create_stock_entry(work_order,work_order_pick_list):
	if work_order and work_order_pick_list:
		work_order_doc = frappe.get_doc("Work Order",work_order)
		pick_list_doc = frappe.get_doc("Work Order Pick List",work_order_pick_list)
		qty_of_finish_good = frappe.db.get_value("Pick Orders",{'parent':work_order_pick_list,'work_order':work_order},'qty_of_finished_goods_to_pull')
		data =  frappe.db.sql("""SELECT item_code,warehouse as s_warehouse,picked_qty,work_order,stock_uom,engineering_revision,batch_no,serial_no from `tabWork Order Pick List Item` where parent = '{0}' and work_order = '{1}' and picked_qty > 0""".format(work_order_pick_list,work_order),as_dict =1)
		if len(data) > 0:
			stock_entry = frappe.new_doc("Stock Entry")
			if stock_entry:
				if pick_list_doc.purpose == 'Material Transfer for Manufacture':
					stock_entry.stock_entry_type = 'Material Transfer for Manufacture'
					stock_entry.company = work_order_doc.company
					stock_entry.work_order = work_order
					stock_entry.work_order_pick_list = work_order_pick_list
					stock_entry.from_bom = 1
					stock_entry.bom_no = work_order_doc.bom_no
					stock_entry.fg_completed_qty = qty_of_finish_good
					# stock_entry.fg_completed_qty = data.get("fg_completed_qty")
					# stock_entry.from_warehouse = data.get("selected_warehouse")
					stock_entry.to_warehouse = work_order_doc.wip_warehouse
					for row in data:
						stock_entry.append("items",{
								"item_code": row.get("item_code"),
									"qty": row.get("picked_qty"),
									"s_warehouse": row.get("s_warehouse"),
									"t_warehouse": work_order_doc.wip_warehouse,
									'batch_no' : row.get("batch_no"),
									'serial_no' : row.get("serial_no")

							})
					stock_entry.save()
				else:
					stock_entry.stock_entry_type = 'Manufacture'
					stock_entry.company = work_order_doc.company
					stock_entry.work_order = work_order
					stock_entry.work_order_pick_list = work_order_pick_list
					stock_entry.from_bom = 1
					stock_entry.bom_no = work_order_doc.bom_no
					stock_entry.fg_completed_qty = qty_of_finish_good
					# stock_entry.fg_completed_qty = data.get("fg_completed_qty")
					# stock_entry.from_warehouse = data.get("selected_warehouse")
					stock_entry.to_warehouse = work_order_doc.fg_warehouse
					for row in data:
						stock_entry.append("items",{
							"item_code": row.get("item_code"),
								"qty": row.get("picked_qty"),
								"s_warehouse": row.get("s_warehouse"),
								# "t_warehouse": work_order_doc.fg_warehouse,
								'batch_no' : row.get("batch_no"),
								'serial_no' : row.get("serial_no")

						})
					stock_entry.append("items",{
						"item_code":work_order_doc.production_item,
						"qty": qty_of_finish_good,
						# "s_warehouse": row.get("s_warehouse"),
						"t_warehouse": work_order_doc.fg_warehouse,
						# 'batch_no' : row.get("batch_no"),
						# 'serial_no' : row.get("serial_no")
					})
					stock_entry.save()
				return stock_entry.name
		else:
			frappe.throw("Please Pick Quantity For Atleast One Item")
@frappe.whitelist()
def get_pick_list_details(doc):
	data = json.loads(doc)
	header = ['Sr.No','Work Order','Total Qty of Finished Goods on Work Order','Qty of Finished Goods Already Completed','Qty of Finished Goods To Pull']
	header_2 = ['Sr.No','Item','Warehouse','Work Order','Required Qty','Stock Qty','Picked Qty','Item Name','Description','Item Group','Batch','Engineering Revision','UOM','Stock UOM','Conversion Factor']
	
	book = Workbook()
	sheet = book.active
	
	row = 1
	col = 1

	cell = sheet.cell(row=row,column=col)
	cell.value = 'Company'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+1)
	cell.value = data.get("company")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+1,column=col)
	cell.value = 'Purpose'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+1,column=col+1)
	cell.value = data.get("purpose")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+2,column=col)
	cell.value = 'Production Plan'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+2,column=col+1)
	cell.value = data.get("production_plan")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+4,column=col)
	cell.value = 'Pick List Work Order Table'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	for item in header:
		cell = sheet.cell(row=row+5,column=col)
		cell.value = item
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

		col+=1
	row = 7
	col = 1
	# last_row = row
	for item in data.get("work_orders"):
		cell = sheet.cell(row=row,column=col)
		cell.value = item.get('idx')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+1)
		cell.value = item.get('work_order')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = item.get('total_qty_to_of_finished_goods_on_work_order')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+3)
		cell.value = item.get('qty_of_finished_goods_already_completed')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+4)
		cell.value = item.get('qty_of_finished_goods_to_pull')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		row+=1

	cell = sheet.cell(row=row+1,column=col)
	cell.value = 'Work Order Pick List Item'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')


	for item in header_2:
		cell = sheet.cell(row=row+2,column=col)
		cell.value = item
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

		col+=1
	row = row+3
	col = 1
	for item in data.get('work_order_pick_list_item'):
		cell = sheet.cell(row=row,column=col)
		cell.value = item.get('idx')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+1)
		cell.value = item.get('item_code')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = item.get('warehouse')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value = item.get('work_order')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+4)
		cell.value = item.get('required_qty')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+5)
		cell.value = item.get('stock_qty')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+6)
		cell.value = item.get('picked_qty')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = sheet.cell(row=row,column=col+7)
		cell.value = item.get('item_name')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = sheet.cell(row=row,column=col+8)
		cell.value = item.get('description')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+9)
		cell.value = item.get('item_group')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+10)
		cell.value = item.get('batch_no')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = sheet.cell(row=row,column=col+11)
		cell.value = item.get('engineering_revision')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+12)
		cell.value = item.get('uom')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+13)
		cell.value = item.get('stock_uom')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+14)
		cell.value = item.get('conversion_factor')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		row+=1



	file_path = frappe.utils.get_site_path("public")
	now = datetime.now()
	fname = "WORK_ORDER_PICK_LIST" + nowdate() + ".xlsx"
	book.save(file_path+fname)


@frappe.whitelist()
def download_xlsx():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	now = datetime.now()
	fname = "WORK_ORDER_PICK_LIST" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname



@frappe.whitelist()
def get_batch_no(item_code, warehouse, qty=1, throw=False):
	"""
	Get batch number using First Expiring First Out method.
	:param item_code: `item_code` of Item Document
	:param warehouse: name of Warehouse to check
	:param qty: quantity of Items
	:return: String represent batch number of batch with sufficient quantity else an empty String
	"""

	batch_no = None
	batches = get_batches(item_code, warehouse, qty, throw)
	for batch in batches:
		if cint(qty) <= cint(batch.qty):
			batch_no = batch.name
			break

	# if not batch_no:
	# 	frappe.msgprint(_('Please select a Batch for Item {0}. Unable to find a single batch that fulfills this requirement').format(frappe.bold(item_code)))
	# 	if throw:
	# 		raise UnableToSelectBatchError

	return batch_no


def get_batches(item_code, warehouse, qty=1, throw=False):
	# from erpnext.stock.doctype.serial_no.serial_no import get_serial_nos
	cond = ''
	if frappe.get_cached_value('Item', item_code, 'has_batch_no'):
		# serial_nos = get_serial_nos(serial_no)
		batch = frappe.get_all("Serial No",
			fields = ["distinct batch_no"],
			filters= {
				"item_code": item_code,
				"warehouse": warehouse
			}
		)
		# if not batch:
		# 	validate_serial_no_with_batch(serial_nos, item_code)

		if batch and len(batch) > 1:
			return []

		cond = " and b.name = %s" %(frappe.db.escape(batch[0].batch_no))

	return frappe.db.sql("""
		SELECT b.name, sum(`tabStock Ledger Entry`.actual_qty) as qty
		from `tabBatch` b
			join `tabStock Ledger Entry` ignore index (item_code, warehouse)
				on (b.name = `tabStock Ledger Entry`.batch_no )
		where `tabStock Ledger Entry`.item_code = %s and `tabStock Ledger Entry`.warehouse = %s
			and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) {0}
		group by batch_id
		order by b.expiry_date ASC, b.creation ASC
	""".format(cond), (item_code, warehouse), as_dict=True)

def validate_serial_no_with_batch(serial_nos, item_code):
	if frappe.get_cached_value("Serial No", serial_nos[0], "item_code") != item_code:
		frappe.throw(_("The serial no {0} does not belong to item {1}")
			.format(get_link_to_form("Serial No", serial_nos[0]), get_link_to_form("Item", item_code)))

	serial_no_link = ','.join(get_link_to_form("Serial No", sn) for sn in serial_nos)

	message = "Serial Nos" if len(serial_nos) > 1 else "Serial No"
	frappe.throw(_("There is no batch found against the {0}: {1}")
		.format(message, serial_no_link))

@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_work_order_for_pick_list(doctype, txt, searchfield, start, page_len, filters):
	fg_item_groups = frappe.db.sql("""SELECT item_group from `tabTable For Item Group`""",as_dict=1)
	todays_date = datetime.date.today()
	fg_item_groups = [item.get('item_group') for item in fg_item_groups]
	if filters.get("purpose") == "Material Transfer for Manufacture":
		return frappe.db.sql(""" SELECT wo.name FROM `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and wo.planned_start_date >= '{0}' and i.item_group in {1}""".format(todays_date,tuple(fg_item_groups)),debug=1)
	else:
		return frappe.db.sql(""" SELECT name FROM `tabWork Order` where status in ('Not Started','In Process') and skip_transfer = 1 or material_transferred_for_manufacturing > 0 or produced_qty < qty and planned_start_date >= '{0}'""".format(todays_date),debug=1)

@frappe.whitelist()
def get_pending_qty(work_order,purpose):
	if purpose == "Material Transfer for Manufacture":
		qty = frappe.db.sql("""SELECT (qty-material_transferred_for_manufacturing) as qty from `tabWork Order` where name = '{0}'""".format(work_order),as_dict=1)
		doc = frappe.get_doc("Work Order",work_order)
		ohs = get_current_stock()
		qty_will_be_produced_list = []
		for item in doc.required_items:
			if item.item_code in ohs:
				if item.required_qty <= ohs.get(item.item_code):
					percent_stock = 100
					qty_will_be_produced = item.required_qty
					qty_will_be_produced_list.append(qty_will_be_produced)
				elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0: 
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					qty_will_be_produced = (percent_stock/100*item.required_qty)
					qty_will_be_produced_list.append(qty_will_be_produced)
				else : 
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					qty_will_be_produced = 0
					qty_will_be_produced_list.append(qty_will_be_produced)
		
		qty[0]['qty_will_be_produced'] =min(qty_will_be_produced_list)
		return qty
	else :
		qty = frappe.db.sql("""SELECT (qty-produced_qty) as qty from `tabWork Order` where name = '{0}'""".format(work_order))
		return qty

def get_sub_assembly_items(bom_no, bom_data, to_produce_qty, indent=0):
	data = get_children('BOM', parent = bom_no)
	for d in data:
		if d.expandable:
			parent_item_code = frappe.get_cached_value("BOM", bom_no, "item")
			stock_qty = (d.stock_qty / d.parent_bom_qty) * flt(to_produce_qty)
			bom_data.append(frappe._dict({
				'parent_item_code': parent_item_code,
				'description': d.description,
				'production_item': d.item_code,
				'item_name': d.item_name,
				'stock_uom': d.stock_uom,
				'uom': d.stock_uom,
				'bom_no': d.value,
				'is_sub_contracted_item': d.is_sub_contracted_item,
				'bom_level': indent,
				'indent': indent,
				'stock_qty': stock_qty
			}))

			if d.value:
				get_sub_assembly_items(d.value, bom_data, stock_qty, indent=indent+1)
@frappe.whitelist()
def print_label(data,doc):
	data = json.loads(data)
	data1 = json.loads(doc)
	print("================data1",data1)
	work_order_list = [item.get('work_order') for item in data1.get('work_order_table')]
	total_work_orders = len(work_order_list)
	url = frappe.db.get_value('URL Data',{'sourcedoctype_name':'Work Order Pick List'},'url')
	public_file_path = url.split('app')
	public_file_path = public_file_path[0] + 'files/work_order_pick_list.xlsx'
	final_url = url + data1.get('name')
	file_path = os.path.realpath(get_files_path(is_private=0))
	file = file_path + '/' + 'work_order_pick_list.xlsx'
	if os.path.exists(file):
		wb = openpyxl.load_workbook(filename=file)
		ws = wb['Sheet']
		row = ws.max_row +1
		# row = ws.get_highest_row() + 1
		col = 1
		cell = ws.cell(row=row,column=col)
		cell.value = ws.max_row -1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+1)
		cell.value = data1.get('name') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+2)
		cell.value = total_work_orders
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = ws.cell(row=row,column=col+3)
		cell.value =  ''
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+4)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = ws.cell(row=row,column=col+5)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = ws.cell(row=row,column=col+6)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		wb.save(file)
		frappe.msgprint(f''' Work Order Pick List Traveller Updated,You can download it from here
                        <a href = {public_file_path}><b>{public_file_path}</b></a>''')
		return public_file_path

	else:
		header = ['Record ID','Work Order Pick List Name','Work Orders Total','Work Orders','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1

		cell = sheet.cell(row=row,column=col)
		cell.value = 1 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+1)
		cell.value = data1.get('name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = total_work_orders
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value =  ''
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+4)
		cell.value = data.get('no_of_copies') 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+5)
		cell.value = data.get('printer_name')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+6)
		cell.value = final_url 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	
		file = frappe.utils.get_site_path("public")
		fname = file_path+'/' + 'work_order_pick_list'+'.xlsx'
		book.save(fname)
		frappe.msgprint("Work Order Pick Traveler Created,You can download it from here {0}".format(public_file_path))
		return public_file_path