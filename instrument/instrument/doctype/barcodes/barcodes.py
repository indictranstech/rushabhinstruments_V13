# Copyright (c) 2023, instrument and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import json
from datetime import datetime,timedelta,date
from frappe.utils import nowdate,nowtime, today, flt,cstr,get_link_to_form
from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl.utils.cell import get_column_letter
import datetime
import calendar
import time
import os
from frappe.utils import get_files_path
from pathlib import Path
from frappe.utils import getdate,time
import glob
import pandas as pd

from frappe import _
import json
import sys

class Barcodes(Document):
	pass

@frappe.whitelist()
def get_barcode_details(doc):
	data = json.loads(doc)
	print("================data",data)
	if data.get('doctype_name') == 'Batch':
		header = ['Record ID','Part Number','Part Description','Total Qty','Batch','Batch Name','Locations','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1
		for i in data.get('table_for_label_print'):
			cell = sheet.cell(row=row,column=col)
			cell.value = (row - 1)
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value =i.get('part_number')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('part_description')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('total_qty')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('batch_id')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('batch_name')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('locations')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('no_of_copies')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('printer')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('url')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			row +=1
			col = 1
		file_path = frappe.utils.get_site_path("public")
		fname = "Batch_Traveller" + nowdate() + ".xlsx"
		book.save(file_path+fname)
		return file_path + fname	
	elif data.get('doctype_name') == 'Item':
		header = ['Record ID','Part Number','Part Description','Item Code','Locations','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1
		for i in data.get('table_for_label_print'):
			cell = sheet.cell(row=row,column=col)
			cell.value = (row - 1)
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value =i.get('part_number')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('part_description')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('part_number')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('locations')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('no_of_copies')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('printer')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('url')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			row +=1
			col = 1
		file_path = frappe.utils.get_site_path("public")
		fname = "Item_Traveller" + nowdate() + ".xlsx"
		book.save(file_path+fname)
		return file_path + fname
	elif data.get('doctype_name') == 'Work Order':
		header = ['Record ID','Work Order Name','Item To Manufacture','Item Name','Qty To Manufacture','WIP Warehouse','Target Warehouse','Sales Order','Number of Copies','Printer','URL']

		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1
		for i in data.get('table_for_label_print'):
			cell = sheet.cell(row=row,column=col)
			cell.value = (row - 1)
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value =i.get('work_order')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('part_number')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('item_name')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('qty_to_manufacture')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('sales_order')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('wip_warehouse')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('target_warehouse')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('no_of_copies')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('printer')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('url')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			row +=1
			col = 1
		file_path = frappe.utils.get_site_path("public")
		fname = "Work Order_Traveller" + nowdate() + ".xlsx"
		book.save(file_path+fname)
		return file_path + fname
	elif data.get('doctype_name') == 'Stock Entry':
		header = ['Record ID','Stock Entry Name','Stock Entry Type','Work Order','Target Warehouse','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1
		for i in data.get('table_for_label_print'):
			cell = sheet.cell(row=row,column=col)
			cell.value = (row - 1)
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value =i.get('stock_entry')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('stock_entry_type')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('work_order')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('target_warehouse')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1


			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('no_of_copies')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('printer')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('url')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			row +=1
			col = 1
		file_path = frappe.utils.get_site_path("public")
		fname = "Stock_Entry_Traveller" + nowdate() + ".xlsx"
		book.save(file_path+fname)
		return file_path + fname

	elif data.get('doctype_name') == 'Job Card':
		header = ['Record ID','Job Card Name','Production Item','Item Name','Qty To Manufacture','Operation','Workstation','Work Order','Number of Copies','Printer','URL']
		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1
		for i in data.get('table_for_label_print'):
			cell = sheet.cell(row=row,column=col)
			cell.value = (row - 1)
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value =i.get('job_card')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('part_number')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('item_name')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('qty_to_manufacture')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('operation')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('workstation')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('work_order')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('no_of_copies')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('printer')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('url')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			row +=1
			col = 1
		file_path = frappe.utils.get_site_path("public")
		fname = "Job_Card_Traveller" + nowdate() + ".xlsx"
		book.save(file_path+fname)
		return file_path + fname
	elif data.get('doctype_name') == 'Work Order Pick List':
		header = ['Record ID','Work Order Pick List Name','Work Orders Total','Number of Copies','Printer','URL']
	else:
		header = ['Record ID','Pick List Name','Work Order','Qty of Finished Goods Items','Number of Copies','Printer','URL']

		book = Workbook()
		sheet = book.active
		row = 1
		col = 1
		for item in header:
			cell = sheet.cell(row=row,column=col)
			cell.value = item
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1
		row =2
		col =1
		for i in data.get('table_for_label_print'):
			cell = sheet.cell(row=row,column=col)
			cell.value = (row - 1)
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			
			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value =i.get('pick_list_name')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('work_order')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('qty_of_finish_good_items')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+= 1


			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('no_of_copies')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('printer')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			col+=1

			cell = sheet.cell(row=row,column=col)
			cell.value = i.get('url')
			# cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			# cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

			row +=1
			col = 1
		file_path = frappe.utils.get_site_path("public")
		fname = "Pick_List_Traveller" + nowdate() + ".xlsx"
		book.save(file_path+fname)
		return file_path + fname



	
@frappe.whitelist()
def download_xlsx_batch():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	fname = "Batch_Traveller" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

@frappe.whitelist()
def download_xlsx_item():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	fname = "Item_Traveller" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname

@frappe.whitelist()
def download_xlsx_wo():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	fname = "Work_Order_Traveller" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname


@frappe.whitelist()
def download_xlsx_se():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	fname = "Stock_Entry_Traveller" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname


@frappe.whitelist()
def download_xlsx_jc():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	fname = "Job_Card_Traveller" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname

@frappe.whitelist()
def download_xlsx_pl():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	fname = "Pick_List_Traveller" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname

