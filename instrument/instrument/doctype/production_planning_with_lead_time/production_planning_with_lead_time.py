# Copyright (c) 2022, instrument and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe import _, msgprint
from frappe.utils import (
	add_days,
	ceil,
	cint,
	comma_and,
	flt,
	get_link_to_form,
	getdate,
	now_datetime,
	nowdate,today,formatdate, get_first_day, get_last_day 
)
from dateutil.relativedelta import relativedelta
from frappe.utils import (
	cint,
	date_diff,
	flt,
	get_datetime,
	get_link_to_form,
	getdate,
	nowdate,
	time_diff_in_hours,
)
from erpnext.manufacturing.doctype.manufacturing_settings.manufacturing_settings import (
	get_mins_between_operations,
)
from datetime import date,timedelta
from erpnext.manufacturing.doctype.bom.bom import get_children, validate_bom_no
import datetime
from erpnext.stock.doctype.item.item import get_item_defaults, get_last_purchase_details
import openpyxl
from io import BytesIO
from openpyxl import Workbook
from frappe.utils import now
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from copy import deepcopy

class ProductionPlanningWithLeadTime(Document):
	def onload(self):
		mr_doc = frappe.db.get_value("Material Request", {"production_planning_with_lead_time":self.name}, "name")
		if mr_doc:
			update_mr_status_in_raw_materials_table(self, mr_doc)
	
	def on_submit(self):
		frappe.db.set_value("Production Planning With Lead Time",self.name,"status","Submitted")
		frappe.db.commit()
		self.reload()

	def on_cancel(self):
		frappe.db.set_value("Production Planning With Lead Time",self.name,"status","Cancelled")
		frappe.db.commit()
		self.reload()
	
	@frappe.whitelist()
	def get_open_sales_orders(self):
		""" Pull sales orders  which are pending to deliver based on criteria selected"""
		self.sales_order_table = ''
		open_so = get_sales_orders(self)
		open_mr = get_open_mr(self)
		
		if open_so or open_mr:
			self.add_so_in_table(open_so,open_mr)
		else:
			frappe.msgprint(_("Sales orders And Material Request are not available for production"))
	@frappe.whitelist()
	def add_so_in_table(self, open_so,open_mr):
		""" Add sales orders in the table"""
		self.set('sales_order_table', [])
		if open_so != []:
			for data in open_so:
				bom_data = frappe.db.sql("""SELECT name,makeup_days from `tabBOM` where item = '{0}' and is_active = 1 and is_default =1 and docstatus = 1""".format(data.get("item_code")),as_dict=1)
				dn_date = data.get('delivery_date')
				today_date = date.today()
				# Calculate Days to Deliver
				days_to_deliver = (dn_date - today_date)
				self.append('sales_order_table', {
					'sales_order': data.get("name"),
					'item':data.get('item_code'),
					'delivery_date':data.get('delivery_date'),
					'qty':data.get('qty'),
					'bom':bom_data[0].get('name'),
					'makeup_days':bom_data[0].get('makeup_days'),
					'days_to_deliver' : days_to_deliver.days,
					'sales_order_item':data.get('sales_order_item')
				})
		if open_mr != []:
			for data in open_mr:
				bom_data = frappe.db.sql("""SELECT name,makeup_days from `tabBOM` where item = '{0}' and is_active = 1 and is_default =1 and docstatus = 1""".format(data.get("item_code")),as_dict=1)
				dn_date = data.get('delivery_date')
				today_date = date.today()
				# Calculate Days to Deliver
				days_to_deliver = (dn_date - today_date)
				self.append('sales_order_table', {
					'material_request': data.get("name"),
					'item':data.get('item_code'),
					'delivery_date':data.get('delivery_date'),
					'qty':data.get('qty'),
					'bom':bom_data[0].get('name'),
					'makeup_days':bom_data[0].get('makeup_days'),
					'days_to_deliver' : days_to_deliver.days,
					'material_request_item':data.get('material_request_item')
				})
		# self.save()
		return self.sales_order_table
	@frappe.whitelist()
	def sort_so_data(self):
		# sort so data based on delivery_date and priority
		self.sorted_sales_order_table = ''
		if self.sales_order_table:
			so_data = []
			for row in self.sales_order_table:
				so_data.append({
					'sales_order':row.sales_order,
					'item':row.item,
					'qty':row.qty,
					'delivery_date':row.delivery_date,
					'priority':row.priority,
					'bom':row.bom,
					'days_to_deliver':row.days_to_deliver,
					'makeup_days':row.makeup_days,
					'sales_order_item':row.sales_order_item,
					'material_request':row.material_request,
					'material_request_item':row.material_request_item
					})
			# so_data = frappe.db.sql("""SELECT * from `tabSales Order Table` where parent = '{0}'""".format(self.name),as_dict=1)
			sorted_so_data = sorted(so_data, key = lambda x: (x["delivery_date"],x["priority"]))
			count = 1
			for row in sorted_so_data:
				row.update({'idx':count})
				self.append('sorted_sales_order_table',row)
				count = count + 1
		return self.sorted_sales_order_table
	@frappe.whitelist()
	def work_order_planning(self):
		self.fg_items_table = ''
		# fetch warehouse list from Rushabh settings
		warehouse_list = get_warehouses()
		# Get On hand stock
		ohs = get_ohs(warehouse_list)
		# Get allocated available_stock
		allocated_ohs = get_allocated_ohs_fg()
		# Get Planned Stock
		planned_data = self.get_planned_data_fg()
		# Get allocated Planned Stock
		allocated_planned_stock = get_allocated_planned_stock_fg()
		# Get actual available_stock
		ohs = get_actual_ohs(ohs,allocated_ohs)
		# Get actual planned data for FG
		planned_data = get_actual_planned_fg(planned_data,allocated_planned_stock)
		if self.sales_order_table:
			fg_data = []
			for row in self.sales_order_table:
				fg_data.append({
					'sales_order':row.sales_order,
					'material_request':row.material_request,
					'item':row.item,
					'qty':row.qty,
					'delivery_date':row.delivery_date,
					'priority':row.priority,
					'bom':row.bom,
					'days_to_deliver':row.days_to_deliver,
					'makeup_days':row.makeup_days,
					'sales_order_item':row.sales_order_item,
					'material_request_item':row.material_request_item
					})
		# fg_data = frappe.db.sql("""SELECT * from `tabSales Order Table` where parent = '{0}'""".format(self.name),as_dict=1)
		if fg_data:
			fg_data = sorted(fg_data, key = lambda x: (x["delivery_date"],x["priority"]))
			count = 1
			for row in fg_data:
				qty = flt(row.get("qty")) - flt(ohs.get(row.get("item"))) if flt(ohs.get(row.get("item"))) < flt(row.get("qty")) else 0
				row.update({'planned_qty':qty,'available_stock':ohs.get(row.get('item')),'already_planned_qty':planned_data.get(row.get('item')),'shortage':qty})
				remaining_qty = flt(ohs.get(row.get("item"))) - flt(row.get("qty")) if flt(ohs.get(row.get("item"))) > flt(row.get("qty")) else 0
				ohs.update({row.get('item'):remaining_qty})
				planned_allocate = flt(qty) - flt(planned_data.get(row.get("item"))) if flt(planned_data.get(row.get("item"))) < flt(qty) else 0
				row.update({'planned_qty':planned_allocate,'shortage':planned_allocate})
				remainingg_qty = flt(planned_data.get(row.get("item"))) - flt(qty) if flt(planned_data.get(row.get("item"))) > flt(qty) else 0
				planned_data.update({row.get('item'):remainingg_qty})
				operation_time = frappe.db.sql("""SELECT sum(time_in_mins) as operation_time from `tabBOM Operation` where parent = '{0}'""".format(row.get('bom')),as_dict=1)
				total_operation_time_in_mins = flt(operation_time[0].get('operation_time'))*row.get('planned_qty')
				total_operation_time_in_days = ceil(total_operation_time_in_mins/480)
				# Calculate date_to_be_ready
				delivery_date = datetime.datetime.strptime(row.get('delivery_date'), '%Y-%m-%d')
				delivery_date = delivery_date.date() 
				date_to_be_ready = (delivery_date-timedelta(total_operation_time_in_days)-timedelta(row.get('makeup_days')))

				#calculate partial stock and workstation availability
				partial_qty_dict=get_partial_qty(date_to_be_ready, row.get("item"), warehouse_list)

				partial_workstation_availability=check_partial_workstation_availability(date_to_be_ready,row.get('bom'), row.get('planned_qty'))
				partial_remark = "{0}{1}".format(partial_qty_dict.get("partial_remark"), partial_workstation_availability.get("remark") if partial_workstation_availability else "")
				
				date_dict = check_workstation_availability(date_to_be_ready,row.get('bom'),row.get('planned_qty'))
				row.update({'idx':count,'total_operation_time':total_operation_time_in_days,'date_to_be_ready':date_to_be_ready, 'partial_remark':partial_remark})

				if date_dict:
					row.update({'planned_start_date':date_dict.get('planned_start_date'),'remark':date_dict.get('remark')})
				else:
					row.update({'planned_start_date':date_to_be_ready})
				self.append('fg_items_table',row)
				count = count + 1
			return self.fg_items_table
	@frappe.whitelist()
	def sub_assembly_items(self):
		self.sub_assembly_items_table = ''
		warehouse_list = get_warehouses()
		# Get On hand stock
		ohs = get_ohs(warehouse_list)
		# Get allocated available_stock
		allocated_ohs = get_allocated_ohs_sfg()
		# Get Planned Stock
		planned_data = self.get_planned_data_fg()
		# Get actual available_stock
		ohs = get_actual_ohs(ohs,allocated_ohs)
		if self.fg_items_table:
			for row in self.fg_items_table:
				bom_data = []
				if row.get('planned_qty') > 0:
					date_to_be_ready = datetime.datetime.strptime(row.get('planned_start_date'), '%Y-%m-%d')
					planned_start_date = date_to_be_ready.date()
					get_sub_assembly_item(row.get("bom"), bom_data, row.get("planned_qty"),planned_start_date,row.name, warehouse_list)
					bom_data = sorted(bom_data, key = lambda x: x["bom_level"],reverse=1)
					for item in bom_data:
						final_row = dict()
						qty = flt(item.get("stock_qty")) - flt(ohs.get(item.get("production_item"))) if flt(ohs.get(item.get("production_item"))) < flt(item.get("stock_qty")) else 0
						final_row.update({'qty':qty,'available_stock':ohs.get(item.get('production_item')),'alreaady_planned_qty':planned_data.get(item.get('production_item')),'shortage':qty,'fg_item':item.get('parent_item_code'),'item':item.get('production_item'),'original_qty':item.get('stock_qty'),'bom':item.get('bom_no'),'sales_order':row.get('sales_order'),'material_request':row.get('material_request')})
						remaining_qty = flt(ohs.get(item.get("production_item"))) - flt(item.get("stock_qty")) if flt(ohs.get(item.get("production_item"))) > flt(item.get("stock_qty")) else 0
						ohs.update({item.get('production_item'):remaining_qty})
						planned_allocate = flt(qty) - flt(planned_data.get(item.get("production_item"))) if flt(planned_data.get(item.get("production_item"))) < flt(qty) else 0
						final_row.update({'qty':planned_allocate,'shortage':planned_allocate})
						remainingg_qty = flt(planned_data.get(item.get("production_item"))) - flt(qty) if flt(planned_data.get(item.get("production_item"))) > flt(qty) else 0
						planned_data.update({item.get('production_item'):remainingg_qty})
						# operation_time = frappe.db.sql("""SELECT sum(time_in_mins) as operation_time from `tabBOM Operation` where parent = '{0}'""".format(item.get('bom_no')),as_dict=1)
						# total_operation_time_in_mins = flt(operation_time[0].get('operation_time'))*final_row.get('qty')
						# total_operation_time_in_days = ceil(total_operation_time_in_mins/480)
						# makeup_days = frappe.db.get_value("BOM",{'name':item.get('bom_no')},'makeup_days')
						# date_to_be_ready = datetime.datetime.strptime(row.get('date_to_be_ready'), '%Y-%m-%d')
						# date_to_be_ready = date_to_be_ready.date()
						# date_to_be_ready = (date_to_be_ready-timedelta(total_operation_time_in_days)-timedelta(makeup_days))
						final_row.update({'total_operation_time':item.get('total_operation_time'),'date_to_be_ready':item.get('date_to_be_ready'),'planned_start_date':item.get('planned_start_date'),'fg_row_name':item.get('row_name'),'remark':item.get('remark'), 'partial_remark':item.get('partial_remark'),'parent_item_code':row.get('item')})
						self.append('sub_assembly_items_table',final_row)
			return self.sub_assembly_items_table
						
	@frappe.whitelist()
	def get_raw_materials(self):
		self.raw_materials_table = ''
		warehouse_list = get_warehouses()
		# Get On hand stock
		ohs = get_ohs(warehouse_list)
		# Get allocated available_stock
		allocated_ohs = get_allocated_ohs_raw()
		# Get actual available_stock
		ohs = get_actual_ohs(ohs,allocated_ohs)
		if self.sub_assembly_items_table:
			for row in self.sub_assembly_items_table:
				raw_data = []
				get_raw_items(row.bom,raw_data,row.qty)
				remaining_dict = dict()
				for item in raw_data:
					lead_time = frappe.db.get_value("Item",{'name':item.get('item')},'lead_time_days')
					item.update({'available_stock':ohs.get(item.get('item')),'lead_time':lead_time,'original_qty':item.get('qty')})
					rm_readiness_days = frappe.db.get_single_value("Rushabh Settings",'rm_readiness_days')
					date_to_be_ready = datetime.datetime.strptime(row.get('planned_start_date'), '%Y-%m-%d')
					date_to_be_ready = date_to_be_ready.date()
					required_date = (date_to_be_ready-timedelta(rm_readiness_days))
					item.update({'date_to_be_ready':required_date})
					today_date = date.today()
					# Calculate Order in days
					order_in_days = (required_date - today_date)
					item.update({'order_in_days':order_in_days.days})
					# Allocate from available_stock
					qty = flt(item.get("qty")) - flt(ohs.get(item.get("item"))) if flt(ohs.get(item.get("item"))) < flt(item.get("qty")) else 0
					remainingg_qty = flt(ohs.get(item.get("item"))) - flt(item.get("qty")) if flt(ohs.get(item.get("item"))) > flt(item.get("qty")) else 0
					ohs.update({item.get('item'):remainingg_qty})
					item.update({'shortage':qty,'qty':qty})
					# Allocate from planned_po nd mr of type Purchase
					if item.get('item') not in remaining_dict:
						on_order_stock,schedule_date_dict = get_on_order_stock(self,required_date)
						
						qty = flt(item.get("qty")) - flt(on_order_stock.get(item.get("item"))) if flt(on_order_stock.get(item.get("item"))) < flt(item.get("qty")) else 0
						if flt(item.get("qty")) == 0:
							pass
						elif qty == 0:
							item.update({'latest_date_availability':schedule_date_dict.get(item.get('item'))})
						else:
							latest_date_availability = today_date + timedelta(lead_time)
							item.update({'latest_date_availability':latest_date_availability})
						remaining_qty = flt(on_order_stock.get(item.get('item'))) - item.get('qty') 
						remaining_dict[item.get('item')] = remaining_qty if remaining_qty > 0 else 0
						item.update({'qty':qty,'on_order_stock':on_order_stock.get(item.get("item")),'shortage':qty})
					else:
						virtual_stock = remaining_dict.get(item.get('item'))
						qty = flt(item.get("qty")) - flt(virtual_stock) if flt(virtual_stock) < flt(item.get("qty")) else 0
						remaining_qty = virtual_stock - item.get('qty')

						remaining_dict[item.get('item')] = remaining_qty if remaining_qty > 0 else 0
						item.update({'qty':qty,'on_order_stock':virtual_stock,'shortage':qty})
					if item.get('latest_date_availability'):
						latest_date_availability_in_days = (item.get('latest_date_availability') - today_date).days
						readiness_status = (item.get('date_to_be_ready')-item.get('latest_date_availability'))
						if readiness_status.days > 0:
							item.update({'readiness_status':'#FF0000'})
						elif readiness_status.days < 0:
							item.update({'readiness_status':'#CD3700'})
						else:
							item.update({'readiness_status':'#FF8000'})
					else:
						item.update({'readiness_status':'#FF0000'})
					self.append('raw_materials_table',item)
			return self.raw_materials_table
	@frappe.whitelist()
	def prepare_final_work_orders(self):
		if self.fg_items_table:
			item_list = []
			for row in self.fg_items_table:
				fg_item_dict = dict()
				if row.planned_qty > 0:
					for item in self.sub_assembly_items_table:
						item_dict = dict()
						if item.sales_order == row.sales_order and item.qty >0 and item.fg_row_name == row.name:
							item_dict.update({
								'item':item.item,
								'qty':item.qty,
								'sales_order':item.sales_order,
								'material_request':item.material_request,
								'bom':item.bom,
								'total_operation_time':item.total_operation_time,
								'date_to_be_ready':item.planned_start_date,
								'is_subassembly' : 1,
								'parent_item_code':row.item
								})
							item_list.append(item_dict)
					fg_item_dict.update({
						'item':row.item,
						'qty':row.planned_qty,
						'sales_order':row.sales_order,
						'sales_order_item':row.sales_order_item,
						'material_request':row.material_request,
						'material_request_item':row.material_request_item,
						'bom':row.bom,
						'total_operation_time':row.total_operation_time,
						'date_to_be_ready':row.planned_start_date,
						'is_subassembly':0,
						'parent_item_code':row.item
						})					
					item_list.append(fg_item_dict)
			count = 1
			for row in item_list:
				row.update({'idx':count})
				self.append("final_work_orders",row)
				count = count + 1
			return self.final_work_orders
	@frappe.whitelist()
	def create_work_order(self):
		if self.final_work_orders:
			from erpnext.manufacturing.doctype.work_order.work_order import get_default_warehouse
			wo_list= []
			default_warehouses = get_default_warehouse()
			self.make_work_order_for_finished_goods(wo_list, default_warehouses)
			self.show_list_created_message("Work Order", wo_list)
			frappe.db.set_value("Production Planning With Lead Time",self.name,'status','In Progress')
			frappe.db.commit()
		else:
			frappe.msgprint("Please Prepare for Final Work Orders")
	def make_work_order_for_finished_goods(self, wo_list, default_warehouses):
		items_data = self.get_production_items()
		for key, item in items_data.items():
			set_default_warehouses(item, default_warehouses)
			wo= frappe.db.get_value("Work Order",{'production_item':item.get('production_item'),'so_reference':item.get('so_reference')}) or frappe.db.get_value("Work Order",{'production_item':item.get('production_item'),'mr_reference':item.get('mr_reference')})
			if not wo: 
				work_order = self.create_work_orders(item)
				if work_order:
					wo_list.append(work_order)
					#Add WO Status on Final Work Order Table 
					wo_status = frappe.db.get_value("Work Order", work_order, "status")
					frappe.db.set_value("Final Work Orders", {'item':item.get('production_item'), 'sales_order':item.get('so_reference')}, "wo_status", wo_status)
					frappe.db.set_value("Final Work Orders", {'item':item.get('production_item'), 'sales_order':item.get('so_reference')}, "work_order", work_order)

					frappe.db.set_value("Final Work Orders", {'item':item.get('production_item'), 'material_request':item.get('mr_reference')}, "wo_status", wo_status)
					frappe.db.set_value("Final Work Orders", {'item':item.get('production_item'), 'material_request':item.get('mr_reference')}, "work_order", work_order)

					
					frappe.db.commit()
			else:
				wo_status = frappe.db.get_value("Work Order", wo, "status")
				frappe.db.set_value("Final Work Orders", {'item':item.get('production_item'), 'sales_order':item.get('so_reference')}, "wo_status", wo_status)
				frappe.db.set_value("Final Work Orders", {'item':item.get('production_item'), 'sales_order':item.get('so_reference')}, "work_order", wo)
				frappe.db.set_value("Final Work Orders", {'item':item.get('production_item'), 'material_request':item.get('mr_reference')}, "wo_status", wo_status)
				frappe.db.set_value("Final Work Orders", {'item':item.get('production_item'), 'material_request':item.get('mr_reference')}, "work_order", wo)
				frappe.db.commit()
				wo = get_link_to_form("Work Order", wo)
				so = get_link_to_form("Sales Order",item.get('so_reference'))
				msgprint(_("Work Order {0} is already created for the item {1} and Sales Order {2} ").format(wo,item.get("production_item"),so))
				# frappe.msgprint("Work Order {0} is already created for this item {1}".format(wo,item.get("production_item")))
	def get_production_items(self):
		item_dict = {}
		default_company = frappe.db.get_single_value("Global Defaults", "default_company")
		for d in self.final_work_orders:
			item_data = get_item_defaults(d.item, default_company)
			item_details = {
				"production_item": d.item,
				"use_multi_level_bom": 0,
				"sales_order": d.sales_order if d.is_subassembly == 0 else None,
				"so_reference":d.sales_order,
				"sales_order_item": d.sales_order_item,
				"bom_no": d.bom,
				"description": item_data.get('description'),
				"stock_uom": item_data.get('stock_uom'),
				"company": default_company,
				"planned_start_date": d.date_to_be_ready,
				"qty":d.qty,
				"production_planning_with_lead_time":self.name,
				"material_request":d.material_request if d.is_subassembly == 0 else None,
				"material_request_item":d.material_request_item,
				"mr_reference":d.material_request,
				"parent_item_code":d.parent_item_code
			}
			if d.sales_order:
				item_dict[(d.item, d.sales_order)] = item_details
			else:
				item_dict[(d.item, d.material_request)] = item_details

		return item_dict

	def create_work_orders(self, item):
		from erpnext.manufacturing.doctype.work_order.work_order import OverProductionError
		wo = frappe.new_doc("Work Order")
		wo.update(item)
		wo.planned_start_date = item.get("planned_start_date") or item.get("schedule_date")
		# check item group to set default target warehouse
		item_group = frappe.db.get_value("Item",{'item_code':item.get('production_item')},'item_group')

		if item.get("warehouse"):
			wo.fg_warehouse = item.get("warehouse")
		# if item belong to subassembly assign target warehouse as wip warehosue
		if item_group != 'Product':
			wo.fg_warehouse = item.get('wip_warehouse')

		wo.set_work_order_operations()
		wo.set_required_items()

		try:
			wo.flags.ignore_mandatory = True
			wo.flags.ignore_validate = True
			wo.insert()
			return wo.name
		except OverProductionError:
			pass
	def show_list_created_message(self, doctype, doc_list=None):
		if not doc_list:
			return

		frappe.flags.mute_messages = False
		if doc_list:
			doc_list = [get_link_to_form(doctype, p) for p in doc_list]
			msgprint(_("{0} created").format(comma_and(doc_list)))

	def get_planned_data_fg(self):
		planned_data = frappe.db.sql("""SELECT i.item,sum(i.planned_qty) as qty from `tabFG Items Table` i join `tabProduction Planning With Lead Time` pp on pp.name = i.parent where pp.docstatus not in (0,2) and pp.status not in ('Completed') and i.material_request != '' group by i.item""",as_dict=1,debug=1)
		planned_data = {i.item : i.qty for i in planned_data}
		return planned_data
		# Get planned qty from material request for which production plan not in place and work order not in place
		# planned_mr = frappe.db.sql("""SELECT mri.item_code,sum(mri.qty) as qty from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.transaction_date < '{0}' and mr.transaction_date >= '{1}' and not exists(SELECT pp.name from `tabProduction Plan` pp join `tabProduction Plan Material Request` pp_item on pp_item.parent = pp.name where pp_item.material_request = mr.name) and not exists(SELECT wo.name from `tabWork Order` wo where wo.material_request = mr.name)""".format(self.to_date,self.from_date),as_dict=1)
		# # Manipulate in order to show in dict format
		# planned_data_dict = {item.item_code : item.qty for item in planned_mr if item.item_code != None and item.qty != None}
		# Get planned qty from production plan for which work order not in place
		# planned_data_dict = dict()
		# if self.to_date and self.from_date:
		# 	planned_pp = frappe.db.sql("""SELECT pp_item.item_code,sum(pp_item.planned_qty) as planned_qty from `tabProduction Plan` pp join `tabProduction Plan Item` pp_item on pp_item.parent = pp.name where pp.posting_date < {0} and pp.posting_date >= '{1}' and pp.docstatus = 1 and not exists(SELECT wo.name from `tabWork Order` wo where wo.production_plan = pp.name)""".format(self.to_date,self.from_date),as_dict=1)
		# else:
		# 	planned_pp = frappe.db.sql("""SELECT pp_item.item_code,sum(pp_item.planned_qty) as planned_qty from `tabProduction Plan` pp join `tabProduction Plan Item` pp_item on pp_item.parent = pp.name and pp.docstatus =1 and not exists(SELECT wo.name from `tabWork Order` wo where wo.production_plan = pp.name)""",as_dict=1)
		# # update planned_data_dict
		# if planned_pp:
		# 	for row in planned_pp:
		# 		if row.get('item_code') in planned_data_dict:
		# 			qty = flt(planned_data_dict.get(row.get('item_code'))) + row.get('planned_qty')
		# 			planned_data_dict.update({row.get('item_code'):qty})
		# 		else:
		# 			if row.item_code != None and row.planned_qty != None:
		# 				planned_data_dict.update({row.get('item_code'):row.get('planned_qty')})
		# # Get planned qty from work order
		# if self.to_date and self.from_date:
		# 	planned_wo = frappe.db.sql("""SELECT wo.production_item,(wo.qty-wo.produced_qty) as qty from `tabWork Order` wo where wo.planned_start_date < '{0}' and wo.planned_start_date >= '{1}' and wo.docstatus=1""".format(self.to_date,self.from_date),as_dict=1)
		# else:
		# 	planned_wo = frappe.db.sql("""SELECT wo.production_item,(wo.qty-wo.produced_qty) as qty from `tabWork Order` wo  where wo.docstatus=1""",as_dict=1)
		# # update planned_data_dict
		# if planned_wo:
		# 	for row in planned_wo:
		# 		if row.get('production_item') in planned_data_dict:
		# 			qty = flt(planned_data_dict.get(row.get('production_item'))) + row.get('qty')
		# 			planned_data_dict.update({row.get('production_item'):qty})
		# 		else:
		# 			if row.item_code != None and row.planned_qty != None:
		# 				planned_data_dict.update({row.get('production_item'):row.get('qty')})
		# 			else:
		# 				planned_data_dict = {}
		# return planned_data_dict
	
	@frappe.whitelist()
	def make_material_request(self):
		"""Create Material Requests grouped by Sales Order and Material Request Type"""
		material_request_list = []
		material_request_map = {}
		default_company = frappe.db.get_single_value("Global Defaults", "default_company")
		mr_doc = frappe.db.get_value("Material Request",{'production_planning_with_lead_time':self.name},'name')

		if not mr_doc:
			material_request_doc = frappe.new_doc("Material Request")
			if material_request_doc:
				material_request_doc.material_request_type = "Purchase"
				material_request_doc.transaction_date = nowdate()
				material_request_doc.company = default_company
				material_request_doc.production_planning_with_lead_time = self.name
				for row in self.raw_materials_table:
					if row.qty > 0:
						item_doc = get_item_defaults(row.item, default_company)
						# Make material requests for only stock items
						stock_item = frappe.db.get_value("Item",{'name':row.item},'is_stock_item')
						if stock_item:
							engineering_revision = frappe.db.get_value("Engineering Revision",{'item_code':row.item,'is_default':1,'is_active':1},'name')
							material_request_doc.append("items",{
								'item_code': row.item,
								'item_name': item_doc.get("item_name"),
								'engineering_revision':item_doc.get("engineering_revision") if item_doc.get("engineering_revision") else engineering_revision,
								'rfq_required':item_doc.get('rfq_required'),
								'schedule_date':row.get('date_to_be_ready'),
								'description':item_doc.get('description'),
								'item_group':item_doc.get('item_group'),
								'qty':row.qty,
								'uom':item_doc.get("stock_uom"),
								'stock_uom':item_doc.get("stock_uom"),
								'warehouse':item_doc.get('item_defaults')[0].get('default_warehouse') if item_doc.get('item_defaults') else ''
								})
				material_request_doc.flags.ignore_permissions = 1
				material_request_doc.run_method("set_missing_values")
				material_request_doc.save()

				frappe.flags.mute_messages = False

				if material_request_doc:
					update_mr_status_in_raw_materials_table(self, material_request_doc.name)
					mr = get_link_to_form("Material Request",material_request_doc.name)
					msgprint(_("Material Request {0} Created.").format(mr))
				else:
					msgprint(_("No material request created"))
		else:
			update_mr_status_in_raw_materials_table(self, mr_doc)
			mr = get_link_to_form("Material Request",mr_doc)
			msgprint(_("Material Request {0} is Already Created.").format(mr))

	@frappe.whitelist()
	def download_raw_material(self):
		#Add the headers
		headers = ['Item', 'Qty', 'Original Qty', 'Date To Be Ready', 'Available Stock', 'Shortage', 'Lead Time', 'On Order Stock', 'Order in Days', 'Latest Date Availability', 'Readiness Status', 'MR Status']
		final_data = []
		#raw material table is empty then throw the messsage 
		if not self.raw_materials_table:
			frappe.throw('Raw Material Table Is Empty')
		else:
			for i in self.raw_materials_table:
				#add the data
				dep_data = {}
				dep_data['Item'] = i.item
				dep_data['Qty'] = i.qty
				dep_data['Original Qty'] = i.original_qty
				dep_data['Date To Be Ready'] = i.date_to_be_ready
				dep_data['Available Stock'] = i.available_stock
				dep_data['Shortage'] = i.shortage
				dep_data['Lead Time'] = i.lead_time
				dep_data['On Order Stock'] = i.on_order_stock
				dep_data['Order in Days'] = i.order_in_days
				dep_data['Latest Date Availability'] = i.latest_date_availability
				dep_data['Readiness Status'] = i.readiness_status
				dep_data['MR Status'] = i.mr_status
				final_data.append(dep_data)
			if final_data:
				book = Workbook()	     
				sheet = book.active	
				row = 1
				col = 1
				for item in headers:
					cell = sheet.cell(row=row,column=col)				
					cell.font = Font(bold=True, color="FF0000")  
					cell.value = item
					col += 1
				row = 2
				col = 1

				for item in final_data:            
					cell = sheet.cell(row=row,column=col)
					cell.value = item['Item']

					cell = sheet.cell(row=row,column=col+1)
					cell.value = item['Qty']

					cell = sheet.cell(row=row,column=col+2)
					cell.value = item['Original Qty']                
					
					cell = sheet.cell(row=row,column=col+3)
					cell.value = item['Date To Be Ready']
					
					cell = sheet.cell(row=row,column=col+4)
					cell.value = item['Available Stock']
					
					cell = sheet.cell(row=row,column=col+5)
					cell.value = item['Shortage']
					
					cell = sheet.cell(row=row,column=col+6)
					cell.value = item['Lead Time']
					
					cell = sheet.cell(row=row,column=col+7)
					cell.value = item['On Order Stock']

					cell = sheet.cell(row=row,column=col+8)
					cell.value = item['Order in Days']

					cell = sheet.cell(row=row,column=col+9)
					cell.value = item['Latest Date Availability']

					cell = sheet.cell(row=row,column=col+10)
					cell.value = item['Readiness Status']

					cell = sheet.cell(row=row,column=col+11)
					cell.value = item['MR Status']
					row += 1
				file_path = frappe.utils.get_site_path("public")
				fname = self.name + '-' + now() + ".xlsx"
				# fname = self.name + ".xlsx"
				book.save(file_path+fname)
				return fname

	def validate(self):
		qty = []
		count = 1
		if self.hide_zero_qty_item == 1 and self.raw_materials_table:
			for i in self.raw_materials_table:
				#Get qty greater than zero  				
				if i.qty != 0:
					i.idx = count
					qty.append(deepcopy(i))
					count += 1
			if qty:
				self.raw_materials_table.clear()
				self.raw_materials_table = qty
				# self.reload()
	
@frappe.whitelist()
def download_xlsx(fname):
	file_path = frappe.utils.get_site_path("public")
	# fname = fname + '-' + now() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname
	
def set_default_warehouses(row, default_warehouses):
	for field in ["wip_warehouse", "fg_warehouse"]:
		if not row.get(field):
			row[field] = default_warehouses.get(field)

def make_work_order_for_finished_goods(production_plan_doc, wo_list):
	items_data = get_production_items(production_plan_doc)
	for key, item in items_data.items():
		if production_plan_doc.sub_assembly_items:
			item['use_multi_level_bom'] = 0

		work_order =create_work_order(production_plan_doc,item)
		if work_order:
			wo_list.append(work_order)
@frappe.whitelist()
def get_raw_items(bom,raw_data,qty):
	doc = frappe.get_doc("BOM",{'name':bom})
	if doc:
		for row in doc.items:
			stock_qty = (row.qty*qty)
			if not row.bom_no:
				raw_data.append({
					'item':row.item_code,
					'qty':stock_qty
				})
		return raw_data
@frappe.whitelist()
def get_sub_assembly_item(bom_no, bom_data, to_produce_qty,date_to_be_ready,row_name, warehouse_list, indent=0):
	data = get_children('BOM', parent = bom_no)
	for d in data:
		if d.expandable:
			operation_time = frappe.db.sql("""SELECT sum(time_in_mins) as operation_time from `tabBOM Operation` where parent = '{0}'""".format(d.value),as_dict=1)
			makeup_days = frappe.db.get_value("BOM",{'name':d.value},'makeup_days')
			parent_item_code = frappe.get_cached_value("BOM", bom_no, "item")
			stock_qty = (d.stock_qty / d.parent_bom_qty) * flt(to_produce_qty)
			total_operation_time_in_mins = flt(operation_time[0].get('operation_time'))*stock_qty
			total_operation_time_in_days = ceil(total_operation_time_in_mins/480)
			# Calculate date_to_be_ready
			date_to_be_ready = (date_to_be_ready-timedelta(total_operation_time_in_days)-timedelta(makeup_days))
			date_dict = check_workstation_availability(date_to_be_ready,d.value,stock_qty)

			#calculate partial stock and workstation availability
			partial_qty_dict=get_partial_qty(date_to_be_ready, d.item_code, warehouse_list)
			partial_workstation_availability=check_partial_workstation_availability(date_to_be_ready,d.value, stock_qty)
			partial_remark = "{0}{1}".format(partial_qty_dict.get("partial_remark"), partial_workstation_availability.get("remark") if partial_workstation_availability else "")

			bom_data.append(frappe._dict({
				'parent_item_code': parent_item_code,
				'description': d.description,
				'production_item': d.item_code,
				'item_name': d.item_name,
				'stock_uom': d.stock_uom,
				'uom': d.stock_uom,
				'bom_no': d.value,
				'is_sub_contracted_item': d.is_sub_contracted_item,
				'bom_level': indent,
				'indent': indent,
				'stock_qty': stock_qty,
				'date_to_be_ready':date_to_be_ready,
				'planned_start_date' : date_dict.get("planned_start_date") if date_dict else date_to_be_ready,
				'total_operation_time':total_operation_time_in_days,
				'row_name':row_name,
				'remark':date_dict.get("remark") if date_dict else "",
				"partial_remark": partial_remark
			}))
			if d.value:

				if date_dict:
					get_sub_assembly_item(d.value, bom_data, stock_qty,date_dict.get('planned_start_date'), row_name, warehouse_list, indent=indent+1)
				else:
					get_sub_assembly_item(d.value, bom_data, stock_qty,date_to_be_ready,row_name, warehouse_list, indent=indent+1)
def get_on_order_stock(self,required_date):
	planned_po = frappe.db.sql("""SELECT poi.item_code,if(sum(poi.qty-poi.received_qty)>0,sum(poi.qty-poi.received_qty),0) as qty,poi.schedule_date from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.schedule_date < '{0}' and qty > 0""".format(required_date),as_dict=1)
	# Manipulate in order to show in dict format
	if planned_po:
		on_order_stock = {item.item_code : item.qty for item in planned_po if item.item_code!=None}
	else:
		on_order_stock = dict()
	schedule_date_dict = {item.item_code : item.schedule_date for item in planned_po if item.item_code!=None}
	planned_mr = frappe.db.sql("""SELECT mri.item_code,if(sum(mri.qty)>0,sum(mri.qty),0) as qty ,mri.schedule_date from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.schedule_date <= '{0}' and mr.material_request_type = 'Purchase' and not exists (SELECT po.name from `tabPurchase Order` po join `tabPurchase Order Item` poi on poi.parent = po.name where poi.material_request = mr.name)""".format(required_date),as_dict=1)
	if planned_mr:
		for row in planned_mr:
			if row.get('item_code') != None:
				if row.get('item_code') in on_order_stock:
					qty = flt(on_order_stock.get(row.get('item_code'))) + flt(row.get('qty'))
					on_order_stock.update({row.get('item_code'):qty})
				else:
					on_order_stock.update({row.get('item_code'):flt(row.get('qty'))})
				if row.get('item_code') in schedule_date_dict:
					if row.get('schedule_date') > schedule_date_dict.get(row.get('item_code')):
						schedule_date_dict.update({row.get('item_code'):row.schedule_date})
				else:
					schedule_date_dict.update({row.get('item_code'):row.schedule_date})
	return on_order_stock,schedule_date_dict
# fetch warehouse from Rushabh settings
def get_warehouses():
	# warehouse list
	fg_warehouse = frappe.db.sql("SELECT warehouse from `tabWarehouse Table`", as_dict = 1)
	all_warehouses = frappe.db.sql("SELECT warehouse from `tabBin`",as_dict=1)
	from_warehouses = []

	if fg_warehouse:
		for row in fg_warehouse:
			warehouse_list = frappe.db.get_descendants('Warehouse', row.warehouse)
			if warehouse_list:
				for item in warehouse_list:
					from_warehouses.append(item)
			else:
				from_warehouses.append(row.warehouse)
		fg_warehouse_ll = ["'" + row + "'" for row in from_warehouses]
		fg_warehouse_list = ','.join(fg_warehouse_ll)
	else:
		for row in all_warehouses:
			from_warehouses.append(row.get('warehouse'))
		fg_warehouse_ll = ["'" + row + "'" for row in from_warehouses]
		fg_warehouse_list = ','.join(fg_warehouse_ll)
	    # fg_warehouse_list = "' '"
	
	return fg_warehouse_list

def get_ohs(fg_warehouse_list):
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse in ({0}) group by item_code """.format(fg_warehouse_list),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict

def get_allocated_ohs_fg():
	allocated_ohs = frappe.db.sql("""SELECT i.item,sum(i.available_stock) as qty from `tabFG Items Table` i join `tabProduction Planning With Lead Time` pp on pp.name = i.parent where pp.docstatus not in (0,2) and pp.status not in ('Completed') group by i.item""",as_dict=1,debug=1)
	allocated_ohs = {item.item : item.qty for item in allocated_ohs}

	return allocated_ohs

def get_allocated_planned_stock_fg():
	allocated_ohs = frappe.db.sql("""SELECT i.item,sum(i.already_planned_qty) as qty from `tabFG Items Table` i join `tabProduction Planning With Lead Time` pp on pp.name = i.parent where pp.docstatus not in (0,2) and pp.status not in ('Completed') group by i.item""",as_dict=1,debug=1)
	allocated_ohs = {item.item : item.qty for item in allocated_ohs}

	return allocated_ohs


def get_allocated_ohs_sfg():
	allocated_ohs = frappe.db.sql("""SELECT i.item,sum(i.available_stock) as qty from `tabSub Assembly Items Table` i join `tabProduction Planning With Lead Time` pp on pp.name = i.parent where pp.docstatus not in (0,2) and pp.status not in ('Completed') group by i.item""",as_dict=1,debug=1)
	allocated_ohs = {item.item : item.qty for item in allocated_ohs}
	return allocated_ohs

def get_allocated_planned_stock_sfg():
	allocated_ohs = frappe.db.sql("""SELECT i.item,sum(i.already_planned_qty) as qty from `tabSub Assembly Items Table` i join `tabProduction Planning With Lead Time` pp on pp.name = i.parent where pp.docstatus not in (0,2) and pp.status not in ('Completed') group by i.item""",as_dict=1,debug=1)
	allocated_ohs = {item.item : item.qty for item in allocated_ohs}

	return allocated_ohs

def get_allocated_ohs_raw():
	allocated_ohs = frappe.db.sql("""SELECT i.item,sum(i.available_stock) as qty from `tabRaw Materials Table` i join `tabProduction Planning With Lead Time` pp on pp.name = i.parent where pp.docstatus not in (0,2) and pp.status not in ('Completed') group by i.item""",as_dict=1,debug=1)
	allocated_ohs = {item.item : item.qty for item in allocated_ohs}

	return allocated_ohs

def get_actual_ohs(ohs,allocated_ohs):
	if ohs and allocated_ohs:
		for row in allocated_ohs:
			if row in ohs:
				qty = ohs.get(row) - allocated_ohs.get(row)
				ohs.update({row:qty})
	return ohs

def get_actual_planned_fg(planned_data,allocated_planned_stock):
	if planned_data and allocated_planned_stock:
		for row in planned_data:
			if row in planned_data:
				qty = planned_data.get(row) - allocated_planned_stock.get(row)
				planned_data.update({row:qty})
	return planned_data

def get_sales_orders(self):
	so_filter = item_filter = ""
	bom_item = "bom.item = so_item.item_code"

	date_field_mapper = {
		'from_date': ('>=', 'so.transaction_date'),
		'to_date': ('<=', 'so.transaction_date'),
		'from_delivery_date': ('>=', 'so_item.delivery_date'),
		'to_delivery_date': ('<=', 'so_item.delivery_date')
	}

	for field, value in date_field_mapper.items():
		if self.get(field):
			so_filter += f" and {value[1]} {value[0]} %({field})s"

	for field in ['customer']:
		if self.get(field):
			so_field = field
			so_filter += f" and so.{so_field} = %({field})s"

	if self.item and frappe.db.exists('Item', self.item):
		bom_item = get_bom_item(self) or bom_item
		item_code = self.item
		item_filter += " and so_item.item_code = '%s'" % self.item

	open_so = frappe.db.sql(f"""
		select distinct so.name, so.transaction_date, so.customer, date(so_item.delivery_date) as delivery_date,so_item.item_code,(so_item.qty-so_item.delivered_qty) as qty,so_item.name as sales_order_item
		from `tabSales Order` so, `tabSales Order Item` so_item
		where so_item.parent = so.name
			and so.docstatus = 1 and so.status not in ("Stopped", "Closed") and ((so_item.qty-so_item.delivered_qty)>0) {so_filter} {item_filter}
			and (exists (select name from `tabBOM` bom where bom.item = so_item.item_code
					and bom.is_active = 1)
				or exists (select name from `tabPacked Item` pi
					where pi.parent = so.name and pi.parent_item = so_item.item_code
						and exists (select name from `tabBOM` bom where bom.item=pi.item_code
							and bom.is_active = 1)))
		""", self.as_dict(), as_dict=1)

	open_sales_orders = [so.name for so in open_so]
	
	open_sales_orders = "', '".join(open_sales_orders)

	check_previous_planning = frappe.db.sql("""SELECT sot.sales_order from `tabSales Order Table` sot join `tabProduction Planning With Lead Time` pp  on pp.name = sot.parent where pp.docstatus in (0,1) and sot.sales_order in  ('{0}') """.format(open_sales_orders),as_dict=1)

	check_previous_planning = [sot.sales_order for sot in check_previous_planning]
	final_so_list = []
	if check_previous_planning:
		for row in open_so:
			if row.name in check_previous_planning:
				pass
			else:
				final_so_list.append(row)
		return final_so_list
	else:
		return open_so
def get_open_mr(self):
	mr_filter = item_filter = ""
	bom_item = "bom.item = mr_item.item_code"

	date_field_mapper = {
		'from_date': ('>=', 'mr.transaction_date'),
		'to_date': ('<=', 'mr.transaction_date'),
		'from_delivery_date': ('>=', 'mr_item.schedule_date'),
		'to_delivery_date': ('<=', 'mr_item.schedule_date')
	}

	for field, value in date_field_mapper.items():
		if self.get(field):
			mr_filter += f" and {value[1]} {value[0]} %({field})s"

	# for field in ['customer']:
	# 	if self.get(field):
	# 		so_field = field
	# 		so_filter += f" and so.{so_field} = %({field})s"

	if self.item and frappe.db.exists('Item', self.item):
		bom_item = get_bom_item(self) or bom_item
		item_filter += " and mr_item.item_code = '%s'" % self.item

	open_mr = frappe.db.sql(f"""
		select distinct mr.name, mr.transaction_date,date(mr_item.schedule_date) as delivery_date,mr_item.item_code,(mr_item.qty-mr_item.ordered_qty) as qty ,mr_item.name as material_request_item
		from `tabMaterial Request` mr, `tabMaterial Request Item` mr_item
		where mr_item.parent = mr.name
			and mr.docstatus = 1 and mr.status not in ("Stopped", "Cancelled") and mr.material_request_type = "Manufacture" and ((mr_item.qty-mr_item.ordered_qty)>0) {mr_filter} {item_filter}
			and (exists (select name from `tabBOM` bom where {bom_item}
					and bom.is_active = 1)
				)
		""", self.as_dict(), as_dict=1)
	open_material_request = [mr.name for mr in open_mr]

	open_material_request = "', '".join(open_material_request)

	check_previous_planning = frappe.db.sql("""SELECT sot.material_request from `tabSales Order Table` sot join `tabProduction Planning With Lead Time` pp  on pp.name = sot.parent where pp.docstatus in (0,1) and sot.material_request in  ('{0}') """.format(open_material_request),as_dict=1)

	check_previous_planning = [sot.material_request for sot in check_previous_planning]
	final_mr_list = []
	if check_previous_planning:
		for row in open_mr:
			if row.name in check_previous_planning:
				pass
			else:
				final_mr_list.append(row)
		return final_mr_list
	else:
		return open_mr
def get_bom_item(self):
	"""Check if Item or if its Template has a BOM."""
	bom_item = None
	has_bom = frappe.db.exists({'doctype': 'BOM', 'item': self.item, 'docstatus': 1})
	if not has_bom:
		template_item = frappe.db.get_value('Item', self.item, ['variant_of'])
		bom_item = "bom.item = {0}".format(frappe.db.escape(template_item)) if template_item else bom_item
	return bom_item

def get_overlap_for(self, args, check_next_available_slot=False):
		production_capacity = 1

		if self.workstation:
			production_capacity = (
				frappe.get_cached_value("Workstation", self.workstation, "production_capacity") or 1
			)
			validate_overlap_for = " and jc.workstation = %(workstation)s "

		if args.get("employee"):
			# override capacity for employee
			production_capacity = 1
			validate_overlap_for = " and jctl.employee = %(employee)s "

		extra_cond = ""
		if check_next_available_slot:
			extra_cond = " or (%(from_time)s <= jctl.from_time and %(to_time)s <= jctl.to_time)"

		existing = frappe.db.sql(
			"""select jc.name as name, jctl.to_time from
			`tabJob Card Time Log` jctl, `tabJob Card` jc where jctl.parent = jc.name and
			(
				(%(from_time)s > jctl.from_time and %(from_time)s < jctl.to_time) or
				(%(to_time)s > jctl.from_time and %(to_time)s < jctl.to_time) or
				(%(from_time)s <= jctl.from_time and %(to_time)s >= jctl.to_time) {0}
			)
			and jctl.name != %(name)s and jc.name != %(parent)s and jc.docstatus < 2 {1}
			order by jctl.to_time desc limit 1""".format(
				extra_cond, validate_overlap_for
			),
			{
				"from_time": args.from_time,
				"to_time": args.to_time,
				"name": args.name or "No Name",
				"parent": args.parent or "No Name",
				"employee": args.get("employee"),
				"workstation": self.workstation,
			},
			as_dict=True,
		)

		if existing and production_capacity > len(existing):
			return

		return existing[0] if existing else None
def check_workstation_availability(date_to_be_ready,bom,qty):
	bom_doc = frappe.get_doc("BOM",bom)
	date_dict = dict()
	if bom_doc.with_operations:
		time_dict = dict()
		for row in bom_doc.operations:
			time_in_mins = flt(row.time_in_mins*qty)
			time_in_days = (time_in_mins/480)
			if row.idx == 1:
				# first operation at planned_start date
				planned_start_time = date_to_be_ready
				time_dict.update({'planned_start_time':planned_start_time})
			else:
				planned_start_time = (
					get_datetime(time_dict.get('planned_end_time')) 
				)
				# + get_mins_between_operations()
			planned_end_time = get_datetime(planned_start_time) + timedelta(time_in_days)
			time_dict.update({'planned_end_time':planned_end_time})
			jc_data = frappe.db.sql("""SELECT jc.name,jc.workstation,date(jc_time.from_time) as from_date from `tabJob Card` jc join `tabJob Card Time Log` jc_time on jc_time.parent = jc.name where jc.workstation = '{0}' and jc.status in ('Open','Work In Progress','Material Transferred','Submitted') and date(jc_time.from_time) between '{1}' and '{2}' and date(jc_time.to_time) between '{1}' and '{2}'""".format(row.workstation,planned_start_time,planned_end_time),as_dict=1)
			if jc_data == []:
				if date_dict.get('planned_start_date'):
					if date_dict.get('planned_start_date') < date_to_be_ready:
						date_dict.update({'planned_start_date':date_to_be_ready,'remark':"All Workstations are available on time"})
				else:
					date_dict.update({'planned_start_date':date_to_be_ready,'remark':"All Workstations are available on time"})
			else:
				jc_data = frappe.db.sql("""SELECT jc.name,jc.workstation,date(jc_time.to_time) as to_time from `tabJob Card` jc join `tabJob Card Time Log` jc_time on jc_time.parent = jc.name where jc.workstation = '{0}' and jc.status in ('Open','Work In Progress','Material Transferred','Submitted') and date(jc_time.from_time) > '{1}' order by to_time desc""".format(row.workstation,planned_start_time),as_dict=1)
				planned_end_time=get_datetime(jc_data[0].get('to_time'))
				# planned_end_time=get_datetime(jc_data[0].get('to_time')) + get_mins_between_operations()
				remark = "Workstation {0} is not available for date {1} ".format(jc_data[0].get('workstation'),date_to_be_ready)
				date_dict.update({'planned_start_date':planned_end_time.date(),'remark':remark})
		return date_dict


def get_partial_qty(date_to_be_ready, item_code, fg_warehouse_list):
	partial_qty_dict = dict()
	current_stock = frappe.db.sql("""SELECT item_code, sum(actual_qty) as qty from `tabBin` where warehouse in ({0}) and item_code='{1}' group by item_code """.format(fg_warehouse_list, item_code),as_dict=1)

	# update planned_data_dict
	if current_stock:
		for row in current_stock:
			if row.get('item_code') in partial_qty_dict:
				qty = flt(partial_qty_dict.get(row.get('item_code'))) + row.get('qty')
				partial_qty_dict.update({row.get('item_code'):qty})
			else:
				partial_qty_dict.update({row.get('item_code'):row.get('qty')})

	# Get planned qty from purchase order
	planned_po = frappe.db.sql("""SELECT schedule_date, item_code, (qty-received_qty) as qty from `tabPurchase Order Item` where docstatus=1 and schedule_date >= '{0}' and schedule_date <= '{1}' and item_code='{2}' """.format(date.today(), date_to_be_ready, item_code), as_dict=1)

	# update partial_qty_dict
	if planned_po:
		for row in planned_po:
			if row.get('item_code') in partial_qty_dict:
				qty = flt(partial_qty_dict.get(row.get('production_item'))) + row.get('qty')
				partial_qty_dict.update({row.get('item_code'):qty})
				if partial_qty_dict.get("schedule_date") and partial_qty_dict.get("schedule_date") < row.get("schedule_date"):
					partial_qty_dict.update({"schedule_date":row.get("schedule_date")})
				else:
					partial_qty_dict.update({"schedule_date":row.get("schedule_date")})
			else:
				partial_qty_dict.update({row.get('item_code'):row.get('qty'), "schedule_date": row.get("schedule_date")})


	#Get planned qty from material request for which Purchase Order not in place
	planned_mr = frappe.db.sql("""SELECT mri.schedule_date, mri.item_code, sum(mri.qty) as qty from `tabMaterial Request` mr join `tabMaterial Request Item` mri on mri.parent = mr.name where mr.schedule_date >= '{0}' and mr.schedule_date <= '{1}' and mri.item_code='{2}' and not exists(SELECT po.name from `tabPurchase Order` po join `tabPurchase Order Item` po_item on po_item.parent = po.name where po_item.material_request = mr.name)""".format(date.today(), date_to_be_ready, item_code),as_dict=1)
	# update partial_qty_dict
	if planned_mr:
		for row in planned_mr:
			if row.get('item_code') in partial_qty_dict:
				qty = flt(partial_qty_dict.get(row.get('item_code'))) + row.get('qty')
				partial_qty_dict.update({row.get('item_code'):qty})
				if partial_qty_dict.get("schedule_date") and partial_qty_dict.get("schedule_date") < row.get("schedule_date"):
					partial_qty_dict.update({"schedule_date":row.get("schedule_date")})
				else:
					partial_qty_dict.update({"schedule_date":row.get("schedule_date")})

			else:
				partial_qty_dict.update({row.get('item_code'):row.get('qty'), "schedule_date": row.get("schedule_date")})
	if partial_qty_dict.get("schedule_date"):
		schedule_date = get_datetime(partial_qty_dict.get("schedule_date")) + timedelta(1)
		partial_remark = "{0} qty can be completed in planned/available inventory on date {1}.\n".format(partial_qty_dict.get(item_code), formatdate(schedule_date, "mm-dd-yyyy"))
		partial_qty_dict.update({"partial_remark":partial_remark})	
	elif partial_qty_dict.get(item_code):
		partial_remark = "{0} qty can be completed in planned/available inventory on date {1}.\n".format(partial_qty_dict.get(item_code), formatdate(date.today(), "mm-dd-yyyy"))
		partial_qty_dict.update({"partial_remark":partial_remark})
	else:
		partial_qty_dict.update({"partial_remark":""})
	return partial_qty_dict


def check_partial_workstation_availability(date_to_be_ready,bom, qty):
	bom_doc = frappe.get_doc("BOM",bom)
	workstation_availability = dict()
	if bom_doc.with_operations:
		time_dict = dict()
		for row in bom_doc.operations:
			jc_data = frappe.db.sql("""SELECT jc.name,jc.workstation,date(jc_time.from_time) as from_date, date(jc_time.to_time) as to_date from `tabJob Card` jc join `tabJob Card Time Log` jc_time on jc_time.parent = jc.name where jc.workstation = '{0}' and jc.status in ('Open','Work In Progress','Material Transferred','Submitted') and date(jc_time.from_time) >= '{1}' and date(jc_time.from_time) <= '{2}' and date(jc_time.to_time) >= '{1}' and date(jc_time.to_time) <= '{2}' order by jc_time.to_time desc """.format(row.workstation, date.today(), date_to_be_ready),as_dict=1)
			if jc_data == []:
				remark = "All Workstation is available for date {0}.".format(formatdate(date.today(), "mm-dd-yyyy"))
				workstation_availability.update({'remark':remark})
			elif jc_data[0].get('to_date') <  date_to_be_ready:
				ws_date = get_datetime(jc_data[0].get('to_date')) + timedelta(1)
				remark = "Workstation {0} is available for date {1} ".format(jc_data[0].get('workstation'),formatdate(ws_date, "mm-dd-yyyy"))
				workstation_availability.update({'remark':remark})
			else:
				workstation_availability.update({'remark':""})
		return workstation_availability

def update_mr_status_in_raw_materials_table(self, mr_doc):
	doc = frappe.get_doc("Material Request", mr_doc)
	for row in doc.items:
		frappe.db.set_value("Raw Materials Table", {'item':row.item_code}, "mr_status", doc.status)
		frappe.db.commit()
